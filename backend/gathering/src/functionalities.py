from datetime import datetime, date
from typing import Optional, List, Literal
from uuid import UUID
from sqlalchemy.orm import Session
from sqlalchemy import func, case, cast, Date

from .models.lots import LotModel
from .models.gatherers import GathererModel
from .models.gathering_centers import GatheringCenterModel
from .models.gatherer_gathering_center import GathererGatheringCenterModel
from .models.certifications import CertificationModel
from .models.lot_certifications import LotCertificationModel
from .models.purchases import PurchaseModel
from .models.balance_movements import BalanceMovementModel, BalanceMovementTypeEnum
from .models.lot_status_history import LotStatusHistoryModel
from .models.lot_process_history import LotProcessHistoryModel
from .models.lot_net_weight_history import LotNetWeightHistoryModel
# Import from warehouse module
from modules.warehouse.src.models.store_movement import StoreMovementModel, StoreMovementTypeEnum
# Importar modelos de farmers
from modules.farmers.src.models.farmers import FarmerModel
from modules.farmers.src.models.farms import FarmModel
# Importar modelo de auth
from modules.auth.src.models.identities import IdentityModel
from .schemas import (
    LotCreate, LotUpdate, LotResponse, LotListItemResponse, PaginatedLotListResponse, PurchaseItemResponse,
    GatheringCenterCreate, GatheringCenterUpdate, GatheringCenterResponse, PaginatedGatheringCenterResponse,
    GathererGatheringCenterCreate, GathererGatheringCenterUpdate, GathererGatheringCenterResponse, PaginatedGathererGatheringCenterResponse,
    LotCertificationWithDetailsResponse,
    PurchaseCreate, PurchaseUpdate, PurchaseResponse, PaginatedPurchaseResponse,
    BalanceMovementResponse, BalanceSummaryResponse, BalanceMovementCreate, PaginatedBalanceMovementResponse,
    GathererCreate, GathererUpdate, GathererResponse, PaginateGathererResponse, 
    GathererByGatheringCenterResponse, PaginateGathererByGatheringCenterResponse, GatheringSummaryResponse, BalanceMovementTypeEnum, BalanceSummaryGatherersResponse,
    # Store movement dispatch schemas
    DispatchLotsRequest, DispatchLotsResponse,
    # Schemas anidados
    GatheringCenterNested, GathererNested, FarmerNested, FarmNested, IdentityNested
)
from .resources import resolve_display_name

class Funcionalities:
    def __init__(self, container, database_key: str = "core_db"):
        self.container = container
        self.database_key = database_key
        
    def _get_db(self) -> Session:
        """Obtiene la sesión de base de datos desde el container"""
        return self.container.get(self.database_key, "databases")
    
    def _get_auth_service(self):
        """Obtiene el servicio de auth desde el container"""
        return self.container.get("auth")
    
    # ========== HELPER METHODS FOR LOADING RELATIONS ==========
    def _load_gathering_center(self, db: Session, gathering_center_id: Optional[UUID]) -> Optional[GatheringCenterNested]:
        """Carga los datos de un gathering center"""
        if not gathering_center_id:
            return None
        gc = db.query(GatheringCenterModel).filter(GatheringCenterModel.id == gathering_center_id).first()
        return GatheringCenterNested.model_validate(gc) if gc else None
    
    def _load_gatherer(self, db: Session, gatherer_id: Optional[UUID]) -> Optional[GathererNested]:
        """Carga los datos de un gatherer"""
        if not gatherer_id:
            return None
        gatherer = db.query(GathererModel).filter(GathererModel.id == gatherer_id).first()
        return GathererNested.model_validate(gatherer) if gatherer else None
    
    def _load_farmer(self, db: Session, farmer_id: Optional[UUID]) -> Optional[FarmerNested]:
        """Carga los datos de un farmer"""
        if not farmer_id:
            return None
        farmer = db.query(FarmerModel).filter(FarmerModel.id == farmer_id).first()
        return FarmerNested.model_validate(farmer) if farmer else None
    
    def _load_farm(self, db: Session, farm_id: Optional[UUID]) -> Optional[FarmNested]:
        """Carga los datos de un farm"""
        if not farm_id:
            return None
        farm = db.query(FarmModel).filter(FarmModel.id == farm_id).first()
        return FarmNested.model_validate(farm) if farm else None
    
    def _load_identity(self, db: Session, identity_id: Optional[UUID]) -> Optional[IdentityNested]:
        """Carga los datos de una identity"""
        if not identity_id:
            return None
        identity = db.query(IdentityModel).filter(IdentityModel.id == identity_id).first()
        return IdentityNested.model_validate(identity) if identity else None
    
    def _generar_numero_recibo_fecha_timestamp_aleatorio() -> str:
        fecha = datetime.now()

        anio = str(fecha.year)[-2:]
        mes = str(fecha.month).zfill(2)
        dia = str(fecha.day).zfill(2)

        fecha_formateada = f"{anio}{mes}{dia}"

        timestamp_str = str(int(time.time() * 1000))

        digitos_aleatorios = ""
        indices_aleatorios = set()

        while len(indices_aleatorios) < 6:
            indice = random.randint(0, len(timestamp_str) - 1)
            indices_aleatorios.add(indice)

        for indice in sorted(indices_aleatorios):
            digitos_aleatorios += timestamp_str[indice]

        return f"{fecha_formateada}{digitos_aleatorios.zfill(6)}"
    # ========== LOT METHODS ==========
    def create_lot(self, lot_data: LotCreate) -> LotResponse:
        """Crea un nuevo lote"""
        db = self._get_db()
        try:
            lot = LotModel(**lot_data.model_dump(exclude_none=True))
            db.add(lot)
            db.commit()
            db.refresh(lot)
            
            # Cargar relaciones
            lot_dict = {
                "id": lot.id,
                "name": lot.name,
                "gathering_center_id": lot.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, lot.gathering_center_id),
                "product_type": lot.product_type,
                "current_status": lot.current_status,
                "current_process": lot.current_process,
                "current_store_center_id": lot.current_store_center_id,
                "current_store_center": self._load_gathering_center(db, lot.current_store_center_id),
                "gatherer_id": lot.gatherer_id,
                "gatherer": self._load_gatherer(db, lot.gatherer_id),
                "net_weight": lot.net_weight,
                "created_at": lot.created_at,
                "updated_at": lot.updated_at,
                "disabled_at": lot.disabled_at
            }
            
            return LotResponse(**lot_dict)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_lots_paginated(self, page: int = 1, 
                           per_page: int = 10, 
                           sort_by: Optional[str] = None, 
                           order: Optional[str] = "asc", 
                           search: str = "",
                           status: Optional[Literal["activo", "en_stock", "despachado", "eliminado" ]] = None,
                           gathering_center_id: Optional[UUID] = None,
                           current_store_center_id: Optional[UUID] = None
                           ) -> PaginatedLotListResponse:
        """Obtiene lotes paginados con campos calculados (fresh_weight, cost, certificaciones y purchases)"""
        db = self._get_db()
        
        # Subconsulta para fresh_weight (suma de quantity de compras)
        fresh_weight_subq = db.query(
            PurchaseModel.lot_id,
            func.coalesce(func.sum(PurchaseModel.quantity), 0).label('fresh_weight')
        ).filter(
            PurchaseModel.disabled_at.is_(None)
        ).group_by(PurchaseModel.lot_id).subquery()
        
        # Subconsulta para cost (suma de quantity * price)
        cost_subq = db.query(
            PurchaseModel.lot_id,
            func.coalesce(func.sum(PurchaseModel.quantity * PurchaseModel.price), 0).label('cost')
        ).filter(
            PurchaseModel.disabled_at.is_(None)
        ).group_by(PurchaseModel.lot_id).subquery()
        
        # Query principal con joins
        query = db.query(
            LotModel,
            func.coalesce(fresh_weight_subq.c.fresh_weight, 0).label('fresh_weight'),
            func.coalesce(cost_subq.c.cost, 0).label('cost')
        ).outerjoin(
            fresh_weight_subq, LotModel.id == fresh_weight_subq.c.lot_id
        ).outerjoin(
            cost_subq, LotModel.id == cost_subq.c.lot_id
        )
        
        # Aplicar filtro de centro de acopio si se proporciona
        if gathering_center_id:
            query = query.filter(LotModel.gathering_center_id == gathering_center_id)
        
        # Aplicar filtro de centro de almacenamiento actual si se proporciona
        if current_store_center_id:
            query = query.filter(LotModel.current_store_center_id == current_store_center_id)
        
        # Aplicar filtro de estado
        if current_store_center_id is None:
            ## filtro de estado activo
            if status == "activo":
                query = query.filter(LotModel.current_status == "activo", LotModel.current_store_center_id == None, LotModel.disabled_at.is_(None))
            ## filtro de estado en stock
            elif status == "en_stock":
                query = query.filter(LotModel.current_status == "en stock", LotModel.current_store_center_id == None, LotModel.disabled_at.is_(None))
            ## filtro de estado despachado (cuando current_store_center_id no es nulo)
            elif status == "despachado":
                query = query.filter( LotModel.current_store_center_id != None, LotModel.disabled_at.is_(None))
            ## filtro de estado eliminado
            elif status == "eliminado":
                query = query.filter(LotModel.disabled_at != None)
        else:
            ## filtro de estado activo
            if status == "activo":
                query = query.filter(LotModel.current_status == "activo", LotModel.disabled_at.is_(None))
            ## filtro de estado en stock
            elif status == "en_stock":
                query = query.filter(LotModel.current_status == "en stock", LotModel.disabled_at.is_(None))
            ## filtro de estado despachado (cuando current_store_center_id no es nulo)
            elif status == "despachado":
                query = query.filter( LotModel.current_store_center_id != None, LotModel.disabled_at.is_(None))
            ## filtro de estado eliminado
            elif status == "eliminado":
                query = query.filter(LotModel.disabled_at != None)
            
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(
                LotModel.name.ilike(f"%{search}%")
            )
        
        # Aplicar ordenamiento
        if sort_by:
            if sort_by == 'fresh_weight':
                sort_column = fresh_weight_subq.c.fresh_weight
            elif sort_by == 'cost':
                sort_column = cost_subq.c.cost
            else:
                sort_column = getattr(LotModel, sort_by, None)
            
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(LotModel.created_at.desc())
        
        total = query.count()
        offset = (page - 1) * per_page
        total_pages = (total + per_page - 1) // per_page
        
        results = query.offset(offset).limit(per_page).all()
        
        # Construir respuesta con certificaciones y compras
        items = []
        for lot, fresh_weight, cost in results:
            # Obtener certificaciones del lote
            certifications = db.query(CertificationModel.name).join(
                LotCertificationModel, 
                CertificationModel.id == LotCertificationModel.certification_id
            ).filter(
                LotCertificationModel.lot_id == lot.id,
                LotCertificationModel.disabled_at.is_(None)
            ).all()
            
            cert_names = [cert[0] for cert in certifications]
            
            # Obtener compras del lote
            purchases = db.query(PurchaseModel).filter(
                PurchaseModel.lot_id == lot.id,
                PurchaseModel.disabled_at.is_(None)
            ).all()
            
            # Construir lista de compras con price_total y relaciones cargadas
            purchase_items = []
            for purchase in purchases:
                purchase_items.append(PurchaseItemResponse(
                    id=purchase.id,
                    farmer=self._load_farmer(db, purchase.farmer_id),
                    farm=self._load_farm(db, purchase.farm_id),
                    gatherer=self._load_gatherer(db, purchase.gatherer_id),
                    quantity=float(purchase.quantity),
                    price=float(purchase.price),
                    price_total=float(purchase.quantity * purchase.price),
                    presentation=purchase.presentation,
                    payment_method=purchase.payment_method,
                    purchase_date=purchase.purchase_date,
                    ticket_number=purchase.ticket_number,
                    gathering_center=self._load_gathering_center(db, purchase.gathering_center_id),
                    identity=self._load_identity(db, purchase.identity_id)
                ))
            
            items.append(LotListItemResponse(
                id=lot.id,
                name=lot.name,
                fresh_weight=float(fresh_weight),
                net_weight=float(lot.net_weight) if lot.net_weight else None,
                created_at=lot.created_at,
                gatherer=self._load_gatherer(db, lot.gatherer_id),
                current_store_center=self._load_gathering_center(db, lot.current_store_center_id),
                gathering_center=self._load_gathering_center(db, lot.gathering_center_id),
                product_type=lot.product_type,
                certifications=cert_names,
                cost=float(cost),
                current_process=lot.current_process,
                current_status=lot.current_status,
                purchases=purchase_items
            ))
        
        return PaginatedLotListResponse(
            items=items,
            total=total,
            page=page,
            page_size=per_page,
            total_pages=total_pages
        )
    
    def _get_lots_data(self, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "", status: Optional[Literal["activo", "en_stock", "despachado", "eliminado"]] = None, gathering_center_id: Optional[UUID] = None, current_store_center_id: Optional[UUID] = None) -> List[LotListItemResponse]:
        """Método interno que obtiene datos de lotes con la misma lógica que get_lots_paginated pero sin paginación"""
        db = self._get_db()
        
        # Subconsulta para fresh_weight (suma de quantity de compras)
        fresh_weight_subq = db.query(
            PurchaseModel.lot_id,
            func.coalesce(func.sum(PurchaseModel.quantity), 0).label('fresh_weight')
        ).filter(
            PurchaseModel.disabled_at.is_(None)
        ).group_by(PurchaseModel.lot_id).subquery()
        
        # Subconsulta para cost (suma de quantity * price)
        cost_subq = db.query(
            PurchaseModel.lot_id,
            func.coalesce(func.sum(PurchaseModel.quantity * PurchaseModel.price), 0).label('cost')
        ).filter(
            PurchaseModel.disabled_at.is_(None)
        ).group_by(PurchaseModel.lot_id).subquery()
        
        # Query principal con joins
        query = db.query(
            LotModel,
            func.coalesce(fresh_weight_subq.c.fresh_weight, 0).label('fresh_weight'),
            func.coalesce(cost_subq.c.cost, 0).label('cost')
        ).outerjoin(
            fresh_weight_subq, LotModel.id == fresh_weight_subq.c.lot_id
        ).outerjoin(
            cost_subq, LotModel.id == cost_subq.c.lot_id
        )
        
        # Aplicar filtro de centro de acopio si se proporciona
        if gathering_center_id:
            query = query.filter(LotModel.gathering_center_id == gathering_center_id)
        
        # Aplicar filtro de centro de almacenamiento actual si se proporciona
        if current_store_center_id:
            query = query.filter(LotModel.current_store_center_id == current_store_center_id)
        
        # Aplicar filtro de estado
        if status == "activo":
            query = query.filter(LotModel.current_status == "activo", LotModel.current_store_center_id == None, LotModel.disabled_at.is_(None))
        elif status == "en_stock":
            query = query.filter(LotModel.current_status == "en_stock", LotModel.current_store_center_id == None, LotModel.disabled_at.is_(None))
        elif status == "despachado":
            query = query.filter(LotModel.current_store_center_id != None, LotModel.disabled_at.is_(None))
        elif status == "eliminado":
            query = query.filter(LotModel.disabled_at != None)
            
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(LotModel.name.ilike(f"%{search}%"))
        
        # Aplicar ordenamiento
        if sort_by:
            if sort_by == 'fresh_weight':
                sort_column = fresh_weight_subq.c.fresh_weight
            elif sort_by == 'cost':
                sort_column = cost_subq.c.cost
            else:
                sort_column = getattr(LotModel, sort_by, None)
            
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            query = query.order_by(LotModel.created_at.desc())
        
        results = query.all()
        
        # Construir respuesta con certificaciones y compras (mismo código que get_lots_paginated)
        items = []
        for lot, fresh_weight, cost in results:
            # Obtener certificaciones del lote
            certifications = db.query(CertificationModel.name).join(
                LotCertificationModel, 
                CertificationModel.id == LotCertificationModel.certification_id
            ).filter(
                LotCertificationModel.lot_id == lot.id,
                LotCertificationModel.disabled_at.is_(None)
            ).all()
            
            cert_names = [cert[0] for cert in certifications]
            
            # Obtener compras del lote
            purchases = db.query(PurchaseModel).filter(
                PurchaseModel.lot_id == lot.id,
                PurchaseModel.disabled_at.is_(None)
            ).all()
            
            # Construir lista de compras con price_total y relaciones cargadas
            purchase_items = []
            for purchase in purchases:
                purchase_items.append(PurchaseItemResponse(
                    id=purchase.id,
                    farmer_id=purchase.farmer_id,
                    farmer=self._load_farmer(db, purchase.farmer_id),
                    farm_id=purchase.farm_id,
                    farm=self._load_farm(db, purchase.farm_id),
                    gatherer_id=purchase.gatherer_id,
                    gatherer=self._load_gatherer(db, purchase.gatherer_id),
                    quantity=float(purchase.quantity),
                    price=float(purchase.price),
                    price_total=float(purchase.quantity * purchase.price),
                    presentation=purchase.presentation,
                    payment_method=purchase.payment_method,
                    purchase_date=purchase.purchase_date,
                    ticket_number=purchase.ticket_number,
                    gathering_center_id=purchase.gathering_center_id,
                    gathering_center=self._load_gathering_center(db, purchase.gathering_center_id),
                    identity_id=purchase.identity_id,
                    identity=self._load_identity(db, purchase.identity_id)
                ))
            
            items.append(LotListItemResponse(
                id=lot.id,
                name=lot.name,
                fresh_weight=float(fresh_weight),
                net_weight=float(lot.net_weight) if lot.net_weight else None,
                created_at=lot.created_at,
                gatherer_id=lot.gatherer_id,
                gatherer=self._load_gatherer(db, lot.gatherer_id),
                product_type=lot.product_type,
                certifications=cert_names,
                cost=float(cost),
                current_process=lot.current_process,
                current_status=lot.current_status,
                purchases=purchase_items
            ))
        
        return items
    
    def export_lots_to_excel(self, type_download: Literal["lots", "purchases"], sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "", status: Optional[Literal["activo", "en_stock", "despachado", "eliminado"]] = None, gathering_center_id: Optional[UUID] = None, current_store_center_id: Optional[UUID] = None):
        """Exporta lotes a Excel con dos formatos posibles: 'lots' (una fila por lote) o 'purchases' (una fila por compra)"""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Reutilizar la lógica existente para obtener los datos
        lots_data = self._get_lots_data(sort_by=sort_by, order=order, search=search, status=status, gathering_center_id=gathering_center_id, current_store_center_id=current_store_center_id)
        
        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        
        # Estilo para el encabezado
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if type_download == "lots":
            ws.title = "Lotes"
            
            # Encabezados para formato 'lots' (SIN IDs, solo nombres)
            headers = [
                "Nombre Lote", "Peso Fresco (kg)", "Peso Neto (kg)", 
                "Costo Total", "Tipo Producto", "Proceso Actual", "Estado Actual",
                "Certificaciones", "Acopiador Responsable", "Centro de Acopio", 
                "Centro de Almacenamiento", "Fecha Creación"
            ]
            ws.append(headers)
            
            # Aplicar estilo al encabezado
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Agregar datos de lotes (SIN IDs)
            for lot in lots_data:
                ws.append([
                    lot.name,
                    lot.fresh_weight,
                    lot.net_weight if lot.net_weight else "",
                    lot.cost,
                    lot.product_type.value if hasattr(lot.product_type, 'value') else str(lot.product_type),
                    lot.current_process.value if hasattr(lot.current_process, 'value') else str(lot.current_process),
                    lot.current_status.value if hasattr(lot.current_status, 'value') else str(lot.current_status),
                    ", ".join(lot.certifications),
                    resolve_display_name(lot.gatherer) if lot.gatherer else "",
                    resolve_display_name(lot.gathering_center) if lot.gathering_center else "",
                    resolve_display_name(lot.current_store_center) if lot.current_store_center else "",
                    lot.created_at.strftime("%Y-%m-%d %H:%M:%S")
                ])
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 25  # Nombre Lote
            ws.column_dimensions['B'].width = 18  # Peso Fresco
            ws.column_dimensions['C'].width = 18  # Peso Neto
            ws.column_dimensions['D'].width = 15  # Costo Total
            ws.column_dimensions['E'].width = 18  # Tipo Producto
            ws.column_dimensions['F'].width = 18  # Proceso Actual
            ws.column_dimensions['G'].width = 18  # Estado Actual
            ws.column_dimensions['H'].width = 35  # Certificaciones
            ws.column_dimensions['I'].width = 30  # Acopiador
            ws.column_dimensions['J'].width = 30  # Centro de Acopio
            ws.column_dimensions['K'].width = 30  # Centro de Almacenamiento
            ws.column_dimensions['L'].width = 20  # Fecha Creación
            
        elif type_download == "purchases":
            ws.title = "Compras por Lote"
            
            # Encabezados para formato 'purchases' (SIN IDs, solo nombres y descripciones)
            headers = [
                # Información del Lote
                "Nombre Lote", "Peso Fresco Total (kg)", "Peso Neto (kg)", 
                "Costo Total Lote", "Tipo Producto", "Proceso Actual", "Estado Actual",
                "Certificaciones", "Acopiador Responsable Lote", "Centro de Acopio Lote", "Centro Almacenamiento",
                # Información de la Compra
                "Cantidad (kg)", "Precio Unitario", "Precio Total",
                "Presentación", "Método Pago", "Fecha Compra", "Número Ticket",
                # Información del Productor
                "Código Productor", "Productor", "DNI Productor",
                # Información de la Parcela
                "Parcela", "Área Total Parcela", "Área Cultivada Parcela",
                # Información del Acopiador de la Compra
                "Acopiador Compra", "DNI Acopiador",
                # Información del Centro de Acopio de la Compra
                "Centro Acopio Compra", "Código Centro Acopio",
                # Información de la Identidad
                "Identidad"
            ]
            ws.append(headers)
            
            # Aplicar estilo al encabezado
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Agregar datos: una fila por cada compra (SIN IDs)
            for lot in lots_data:
                for purchase in lot.purchases:
                    ws.append([
                        # Información del Lote (se repite para cada compra, SIN ID)
                        lot.name,
                        lot.fresh_weight,
                        lot.net_weight if lot.net_weight else "",
                        lot.cost,
                        lot.product_type.value if hasattr(lot.product_type, 'value') else str(lot.product_type),
                        lot.current_process.value if hasattr(lot.current_process, 'value') else str(lot.current_process),
                        lot.current_status.value if hasattr(lot.current_status, 'value') else str(lot.current_status),
                        ", ".join(lot.certifications),
                        resolve_display_name(lot.gatherer) if lot.gatherer else "",
                        resolve_display_name(lot.gathering_center) if lot.gathering_center else "",
                        resolve_display_name(lot.current_store_center) if lot.current_store_center else "",
                        # Información de la Compra (SIN ID)
                        purchase.quantity,
                        purchase.price,
                        purchase.price_total,
                        purchase.presentation.value if hasattr(purchase.presentation, 'value') else str(purchase.presentation),
                        purchase.payment_method or "",
                        purchase.purchase_date.strftime("%Y-%m-%d %H:%M:%S"),
                        purchase.ticket_number or "",
                        # Información del Productor
                        purchase.farmer.code if purchase.farmer else "",
                        resolve_display_name(purchase.farmer),
                        purchase.farmer.dni if purchase.farmer else "",
                        # Información de la Parcela
                        resolve_display_name(purchase.farm),
                        purchase.farm.total_area if purchase.farm else "",
                        purchase.farm.cultivated_area if purchase.farm else "",
                        # Información del Acopiador de la Compra
                        resolve_display_name(purchase.gatherer),
                        purchase.gatherer.dni if purchase.gatherer else "",
                        # Información del Centro de Acopio de la Compra
                        resolve_display_name(purchase.gathering_center),
                        purchase.gathering_center.code if purchase.gathering_center else "",
                        # Información de la Identidad
                        resolve_display_name(purchase.identity) if purchase.identity else ""
                    ])
            
            # Ajustar ancho de columnas para formato purchases
            column_widths = [25, 18, 18, 15, 18, 18, 18, 35, 30, 30, 30, 15, 15, 15, 18, 18, 20, 20, 20, 30, 18, 30, 20, 20, 30, 18, 30, 20, 30]
            for idx, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def get_lot_by_id(self, lot_id: UUID) -> Optional[LotResponse]:
        """Obtiene un lote específico por ID (solo si no está deshabilitado)"""
        db = self._get_db()
        lot = db.query(LotModel).filter(
            LotModel.id == lot_id,
            LotModel.disabled_at.is_(None)
        ).first()
        
        if not lot:
            return None
        
        # Cargar relaciones
        lot_dict = {
            "id": lot.id,
            "name": lot.name,
            "gathering_center_id": lot.gathering_center_id,
            "gathering_center": self._load_gathering_center(db, lot.gathering_center_id),
            "product_type": lot.product_type,
            "current_status": lot.current_status,
            "current_process": lot.current_process,
            "current_store_center_id": lot.current_store_center_id,
            "current_store_center": self._load_gathering_center(db, lot.current_store_center_id),
            "gatherer_id": lot.gatherer_id,
            "gatherer": self._load_gatherer(db, lot.gatherer_id),
            "net_weight": lot.net_weight,
            "created_at": lot.created_at,
            "updated_at": lot.updated_at,
            "disabled_at": lot.disabled_at
        }
        
        return LotResponse(**lot_dict)
    
    def update_lot(self, lot_id: UUID, lot_data: LotUpdate, identity_id: Optional[UUID] = None) -> Optional[LotResponse]:
        """
        Actualiza un lote existente.
        
        Genera registros de historial cuando cambian:
        - current_status → lot_status_history
        - current_process → lot_process_history
        - net_weight → lot_net_weight_history
        
        Args:
            lot_id: UUID del lote a actualizar
            lot_data: Datos de actualización
            identity_id: UUID de la identidad que realiza el cambio (del token)
            
        Returns:
            LotResponse con el lote actualizado o None si no existe
        """
        db = self._get_db()
        try:
            lot = db.query(LotModel).filter(
                LotModel.id == lot_id,
                LotModel.disabled_at.is_(None)
            ).first()
            
            if not lot:
                return None
            
            # Guardar valores anteriores para historial
            old_status = lot.current_status
            old_process = lot.current_process
            old_net_weight = lot.net_weight
            
            update_data = lot_data.model_dump(exclude_none=True)
            
            # Detectar cambios y crear registros de historial
            status_changed = 'current_status' in update_data and update_data['current_status'] != old_status
            process_changed = 'current_process' in update_data and update_data['current_process'] != old_process
            weight_changed = 'net_weight' in update_data and update_data['net_weight'] != old_net_weight
            
            # Aplicar cambios al lote
            for key, value in update_data.items():
                setattr(lot, key, value)
            
            lot.updated_at = datetime.utcnow()
            
            # Crear registro de historial de status
            if status_changed:
                status_history = LotStatusHistoryModel(
                    lot_id=lot_id,
                    status=update_data['current_status'].value if hasattr(update_data['current_status'], 'value') else str(update_data['current_status'])
                )
                db.add(status_history)
                print(f"✓ Registro de historial de status creado: {old_status} → {update_data['current_status']}")
            
            # Crear registro de historial de proceso
            if process_changed:
                process_history = LotProcessHistoryModel(
                    lot_id=lot_id,
                    process=update_data['current_process'].value if hasattr(update_data['current_process'], 'value') else str(update_data['current_process'])
                )
                db.add(process_history)
                print(f"✓ Registro de historial de proceso creado: {old_process} → {update_data['current_process']}")
            
            # Crear registro de historial de peso neto
            if weight_changed:
                weight_history = LotNetWeightHistoryModel(
                    lot_id=lot_id,
                    last_net_weight=old_net_weight,
                    new_net_weight=update_data['net_weight'],
                    identity_id=identity_id
                )
                db.add(weight_history)
                print(f"✓ Registro de historial de peso creado: {old_net_weight} kg → {update_data['net_weight']} kg")
            
            db.commit()
            db.refresh(lot)
            
            # Cargar relaciones
            lot_dict = {
                "id": lot.id,
                "name": lot.name,
                "gathering_center_id": lot.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, lot.gathering_center_id),
                "product_type": lot.product_type,
                "current_status": lot.current_status,
                "current_process": lot.current_process,
                "current_store_center_id": lot.current_store_center_id,
                "current_store_center": self._load_gathering_center(db, lot.current_store_center_id),
                "gatherer_id": lot.gatherer_id,
                "gatherer": self._load_gatherer(db, lot.gatherer_id),
                "net_weight": lot.net_weight,
                "created_at": lot.created_at,
                "updated_at": lot.updated_at,
                "disabled_at": lot.disabled_at
            }
            
            return LotResponse(**lot_dict)
        except Exception as e:
            db.rollback()
            print(f"❌ Error al actualizar lote: {e}")
            raise e
    
    def disable_lot(self, lot_id: UUID) -> Optional[LotResponse]:
        """Deshabilita un lote"""
        db = self._get_db()
        try:
            lot = db.query(LotModel).filter(
                LotModel.id == lot_id,
                LotModel.disabled_at.is_(None)
            ).first()
            
            if not lot:
                return None
            
            lot.disabled_at = datetime.now()
            lot.updated_at = datetime.utcnow()
            
            db.commit()
            db.refresh(lot)
            
            # Cargar relaciones
            lot_dict = {
                "id": lot.id,
                "name": lot.name,
                "gathering_center_id": lot.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, lot.gathering_center_id),
                "product_type": lot.product_type,
                "current_status": lot.current_status,
                "current_process": lot.current_process,
                "current_store_center_id": lot.current_store_center_id,
                "current_store_center": self._load_gathering_center(db, lot.current_store_center_id),
                "gatherer_id": lot.gatherer_id,
                "gatherer": self._load_gatherer(db, lot.gatherer_id),
                "net_weight": lot.net_weight,
                "created_at": lot.created_at,
                "updated_at": lot.updated_at,
                "disabled_at": lot.disabled_at
            }
            return LotResponse(**lot_dict)
        except Exception as e:
            db.rollback()
            raise e
    
    def restore_lot(self, lot_id: UUID) -> Optional[LotResponse]:
        """Restaura un lote deshabilitado"""
        db = self._get_db()
        try:
            lot = db.query(LotModel).filter(
                LotModel.id == lot_id,
                LotModel.disabled_at.isnot(None)
            ).first()
            
            if not lot:
                return None
            
            lot.disabled_at = None
            lot.updated_at = datetime.utcnow()
            
            db.commit()
            db.refresh(lot)
            
            # Cargar relaciones
            lot_dict = {
                "id": lot.id,
                "name": lot.name,
                "gathering_center_id": lot.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, lot.gathering_center_id),
                "product_type": lot.product_type,
                "current_status": lot.current_status,
                "current_process": lot.current_process,
                "current_store_center_id": lot.current_store_center_id,
                "current_store_center": self._load_gathering_center(db, lot.current_store_center_id),
                "gatherer_id": lot.gatherer_id,
                "gatherer": self._load_gatherer(db, lot.gatherer_id),
                "net_weight": lot.net_weight,
                "created_at": lot.created_at,
                "updated_at": lot.updated_at,
                "disabled_at": lot.disabled_at
            }
            
            return LotResponse(**lot_dict)
        except Exception as e:
            db.rollback()
            raise e
    
    # ========== GATHERING CENTER METHODS ==========
    def create_gathering_center(self, center_data: GatheringCenterCreate) -> GatheringCenterResponse:
        """Crea un nuevo centro de acopio"""
        db = self._get_db()
        try:
            center = GatheringCenterModel(**center_data.model_dump(exclude_none=True))
            db.add(center)
            db.commit()
            db.refresh(center)
            return GatheringCenterResponse.model_validate(center)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_gathering_centers_paginated(self, page: int = 1, per_page: int = 10, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "") -> PaginatedGatheringCenterResponse:
        """Obtiene centros de acopio paginados (solo los no deshabilitados)"""
        db = self._get_db()

        gatherers_count_sq = (
            db.query(
                GathererGatheringCenterModel.gathering_center_id.label("gc_id"),
                func.count(GathererGatheringCenterModel.gatherer_id).label("gatherers_count"),
            )
            .join(
                GathererModel,
                GathererModel.id == GathererGatheringCenterModel.gatherer_id
            )
            .filter(
                GathererGatheringCenterModel.disabled_at.is_(None),
                GathererModel.disabled_at.is_(None)
            )
            .group_by(GathererGatheringCenterModel.gathering_center_id)
            .subquery()
        )

        lots_count_sq = (
            db.query(
                LotModel.gathering_center_id.label("gc_id"),
                func.count(LotModel.id).label("lots_count"),
            )
            .filter(LotModel.disabled_at.is_(None))
            .group_by(LotModel.gathering_center_id)
            .subquery()
        )

        balance_sq = (
            db.query(
                BalanceMovementModel.gathering_center_id.label("gc_id"),
                func.coalesce(
                    func.sum(
                        case(
                            (BalanceMovementModel.type_movement == BalanceMovementTypeEnum.RECHARGE, BalanceMovementModel.ammount),
                            (BalanceMovementModel.type_movement == BalanceMovementTypeEnum.PURCHASE, -BalanceMovementModel.ammount),
                            else_=0,
                        )
                    ),
                    0,
                ).label("balance"),
            )
            .group_by(BalanceMovementModel.gathering_center_id)
            .subquery()
        )

        query = (
            db.query(
                GatheringCenterModel,
                func.coalesce(gatherers_count_sq.c.gatherers_count, 0).label("gatherers_count"),
                func.coalesce(lots_count_sq.c.lots_count, 0).label("lots_count"),
                func.coalesce(balance_sq.c.balance, 0).label("balance"),
            )
            .outerjoin(
                gatherers_count_sq,
                gatherers_count_sq.c.gc_id == GatheringCenterModel.id,
            )
            .outerjoin(
                lots_count_sq,
                lots_count_sq.c.gc_id == GatheringCenterModel.id,
            )
            .outerjoin(
                balance_sq,
                balance_sq.c.gc_id == GatheringCenterModel.id,
            )
            .filter(GatheringCenterModel.disabled_at.is_(None))
        )

        if search:
            query = query.filter(
                GatheringCenterModel.name.ilike(f"%{search}%") |
                (GatheringCenterModel.code.isnot(None) & GatheringCenterModel.code.ilike(f"%{search}%")) |
                (GatheringCenterModel.description.isnot(None) & GatheringCenterModel.description.ilike(f"%{search}%"))
            )

        if sort_by == "gatherers_count":
            sort_column = gatherers_count_sq.c.gatherers_count
        elif sort_by == "lots_count":
            sort_column = lots_count_sq.c.lots_count
        elif sort_by:
            sort_column = getattr(GatheringCenterModel, sort_by, None)
        else:
            sort_column = GatheringCenterModel.created_at
        
        if sort_column is not None:
            query = query.order_by(
                sort_column.desc() if order == "desc" else sort_column.asc()
            )
        total = (
            db.query(func.count(GatheringCenterModel.id))
            .filter(GatheringCenterModel.disabled_at.is_(None))
            .scalar()
        )
        offset = (page - 1) * per_page
        total_pages = (total + per_page - 1) // per_page
        
        centers = query.offset(offset).limit(per_page).all()

        items = []
        for center, gatherers_count, lots_count, balance in centers:
            data = GatheringCenterResponse.model_validate(center)
            data.gatherers_count = gatherers_count
            data.lots_count = lots_count
            data.balance = balance
            items.append(data)
        
        return PaginatedGatheringCenterResponse(
            items=items,
            total=total,
            page=page,
            page_size=per_page,
            total_pages=total_pages
        )
    
    def export_gathering_centers_to_excel(self, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = ""):
        """Exporta todos los centros de acopio a un archivo Excel con los mismos filtros que la paginación"""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        db = self._get_db()

        # Construir las mismas subqueries que en get_gathering_centers_paginated
        gatherers_count_sq = (
            db.query(
                GathererGatheringCenterModel.gathering_center_id.label("gc_id"),
                func.count(GathererGatheringCenterModel.gatherer_id).label("gatherers_count"),
            )
            .join(
                GathererModel,
                GathererModel.id == GathererGatheringCenterModel.gatherer_id
            )
            .filter(
                GathererGatheringCenterModel.disabled_at.is_(None),
                GathererModel.disabled_at.is_(None)
            )
            .group_by(GathererGatheringCenterModel.gathering_center_id)
            .subquery()
        )

        lots_count_sq = (
            db.query(
                LotModel.gathering_center_id.label("gc_id"),
                func.count(LotModel.id).label("lots_count"),
            )
            .filter(LotModel.disabled_at.is_(None))
            .group_by(LotModel.gathering_center_id)
            .subquery()
        )

        balance_sq = (
            db.query(
                BalanceMovementModel.gathering_center_id.label("gc_id"),
                func.coalesce(
                    func.sum(
                        case(
                            (BalanceMovementModel.type_movement == BalanceMovementTypeEnum.RECHARGE, BalanceMovementModel.ammount),
                            (BalanceMovementModel.type_movement == BalanceMovementTypeEnum.PURCHASE, -BalanceMovementModel.ammount),
                            else_=0,
                        )
                    ),
                    0,
                ).label("balance"),
            )
            .group_by(BalanceMovementModel.gathering_center_id)
            .subquery()
        )

        query = (
            db.query(
                GatheringCenterModel,
                func.coalesce(gatherers_count_sq.c.gatherers_count, 0).label("gatherers_count"),
                func.coalesce(lots_count_sq.c.lots_count, 0).label("lots_count"),
                func.coalesce(balance_sq.c.balance, 0).label("balance"),
            )
            .outerjoin(
                gatherers_count_sq,
                gatherers_count_sq.c.gc_id == GatheringCenterModel.id,
            )
            .outerjoin(
                lots_count_sq,
                lots_count_sq.c.gc_id == GatheringCenterModel.id,
            )
            .outerjoin(
                balance_sq,
                balance_sq.c.gc_id == GatheringCenterModel.id,
            )
            .filter(GatheringCenterModel.disabled_at.is_(None))
        )

        # Aplicar búsqueda si existe
        if search:
            query = query.filter(
                GatheringCenterModel.name.ilike(f"%{search}%") |
                (GatheringCenterModel.code.isnot(None) & GatheringCenterModel.code.ilike(f"%{search}%")) |
                (GatheringCenterModel.description.isnot(None) & GatheringCenterModel.description.ilike(f"%{search}%"))
            )

        # Aplicar ordenamiento
        if sort_by == "gatherers_count":
            sort_column = gatherers_count_sq.c.gatherers_count
        elif sort_by == "lots_count":
            sort_column = lots_count_sq.c.lots_count
        elif sort_by:
            sort_column = getattr(GatheringCenterModel, sort_by, None)
        else:
            sort_column = GatheringCenterModel.created_at
        
        if sort_column is not None:
            query = query.order_by(
                sort_column.desc() if order == "desc" else sort_column.asc()
            )
        
        # Obtener todos los resultados (sin paginación)
        centers = query.all()

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Centros de Acopio"

        # Estilo para el encabezado
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = ["Nombre", "Código", "Descripción", "Lotes", "Cantidad Acopiadores", "Nombres Acopiadores", "Saldo", "Fecha de Creación"]
        ws.append(headers)

        # Aplicar estilo al encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Agregar datos
        for center, gatherers_count, lots_count, balance in centers:
            # Obtener los nombres de los acopiadores asociados a este centro
            gatherers = db.query(GathererModel).join(
                GathererGatheringCenterModel,
                GathererModel.id == GathererGatheringCenterModel.gatherer_id
            ).filter(
                GathererGatheringCenterModel.gathering_center_id == center.id,
                GathererGatheringCenterModel.disabled_at.is_(None),
                GathererModel.disabled_at.is_(None)
            ).all()
            
            # Construir lista de nombres completos
            gatherer_names = [
                resolve_display_name(gatherer) for gatherer in gatherers
            ]
            gatherers_names_str = ", ".join(gatherer_names) if gatherer_names else ""
            
            ws.append([
                center.name,
                center.code or "",
                center.description or "",
                lots_count,
                gatherers_count,
                gatherers_names_str,
                float(balance),
                center.created_at.strftime("%Y-%m-%d %H:%M:%S") if center.created_at else ""
            ])

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30  # Nombre
        ws.column_dimensions['B'].width = 15  # Código
        ws.column_dimensions['C'].width = 40  # Descripción
        ws.column_dimensions['D'].width = 10  # Lotes
        ws.column_dimensions['E'].width = 15  # Acopiadores
        ws.column_dimensions['F'].width = 15  # Saldo
        ws.column_dimensions['G'].width = 20  # Fecha de Creación

        # Formato de número para la columna de saldo
        for row in range(2, ws.max_row + 1):
            ws[f'G{row}'].number_format = '#,##0.00'

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def get_gathering_center_by_id(self, center_id: UUID) -> Optional[GatheringCenterResponse]:
        """Obtiene un centro de acopio específico por ID (solo si no está deshabilitado)"""
        db = self._get_db()
        center = db.query(GatheringCenterModel).filter(
            GatheringCenterModel.id == center_id,
            GatheringCenterModel.disabled_at.is_(None)
        ).first()
        
        if not center:
            return None
        
        return GatheringCenterResponse.model_validate(center)
    
    def update_gathering_center(self, center_id: UUID, center_data: GatheringCenterUpdate) -> Optional[GatheringCenterResponse]:
        """Actualiza un centro de acopio existente"""
        db = self._get_db()
        try:
            center = db.query(GatheringCenterModel).filter(
                GatheringCenterModel.id == center_id,
                GatheringCenterModel.disabled_at.is_(None)
            ).first()
            
            if not center:
                return None
            
            update_data = center_data.model_dump(exclude_none=True)
            for key, value in update_data.items():
                setattr(center, key, value)
            
            center.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(center)
            return GatheringCenterResponse.model_validate(center)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_gathering_summary(
        self,
        gathering_center_id: Optional[UUID] = None
    ) -> GatheringSummaryResponse:

        db = self._get_db()
        today = date.today()

        base_filters = [
            BalanceMovementModel.type_movement == BalanceMovementTypeEnum.PURCHASE,
            BalanceMovementModel.disabled_at.is_(None),
        ]

        if gathering_center_id:
            base_filters.append(
                BalanceMovementModel.gathering_center_id == gathering_center_id
            )

        # 🔹 Subquery: última PURCHASE
        last_purchase_sq = (
            db.query(BalanceMovementModel.ammount)
            .filter(*base_filters)
            .order_by(BalanceMovementModel.created_at.desc())
            .limit(1)
            .scalar_subquery()
        )

        # 🔹 Query principal
        result = (
            db.query(
                func.coalesce(last_purchase_sq, 0).label("last_purchase_amount"),
                func.coalesce(
                    func.sum(
                        case(
                            (
                                cast(BalanceMovementModel.created_at, Date) == today,
                                BalanceMovementModel.ammount,
                            ),
                            else_=0,
                        )
                    ),
                    0,
                ).label("today_expense"),
                func.coalesce(
                    func.sum(
                        case(
                            (
                                func.date_trunc("month", BalanceMovementModel.created_at)
                                == func.date_trunc("month", func.now()),
                                BalanceMovementModel.ammount,
                            ),
                            else_=0,
                        )
                    ),
                    0,
                ).label("month_expense"),
            )
            .filter(*base_filters)
            .one()
        )

        return GatheringSummaryResponse(
            gathering_center_id=gathering_center_id,
            last_purchase_amount=float(result.last_purchase_amount),
            today_expense=float(result.today_expense),
            month_expense=float(result.month_expense),
        )
        
    # ========== GATHERERS METHODS ==========
    def create_gatherer(self, gatherer_data: GathererCreate) -> GathererResponse:
        """Crea un nuevo gatherer y su identidad asociada"""
        db = self._get_db()
        auth_service = self._get_auth_service()
        
        try:
            # Crear la identidad en el módulo auth
            from modules.auth.src.schemas import IdentityCreate
            
            identity_data = IdentityCreate(
                username=gatherer_data.username,
                eid=gatherer_data.dni,
                email=gatherer_data.email,
                first_name=gatherer_data.first_name,
                last_name=gatherer_data.last_name,
                sms_number=gatherer_data.sms_number
            )
            
            identity = auth_service.create_identity(identity_data)
            
            # Crear el gatherer con el identity_id devuelto
            gatherer_dict = gatherer_data.model_dump(exclude_none=True)
            gatherer_dict['identity_id'] = identity.id
            # Remover username del dict ya que no se guarda en gatherers
            gatherer_dict.pop('username', None)
            
            gatherer = GathererModel(**gatherer_dict)
            db.add(gatherer)
            db.commit()
            db.refresh(gatherer)
            
            # Para un gatherer recién creado, no hay compras, por lo tanto es inactivo
            gatherer_response_dict = {
                'id': gatherer.id,
                'identity_id': gatherer.identity_id,
                'first_name': gatherer.first_name,
                'last_name': gatherer.last_name,
                'dni': gatherer.dni,
                'sms_number': gatherer.sms_number,
                'wsp_number': gatherer.wsp_number,
                'call_number': gatherer.call_number,
                'email': gatherer.email,
                'last_purchase_date': None,
                'status': 'inactivo',
                'created_at': gatherer.created_at,
                'updated_at': gatherer.updated_at,
                'disabled_at': gatherer.disabled_at
            }
            
            return GathererResponse(**gatherer_response_dict)
        except Exception as e:
            db.rollback()
            raise e
    
    def update_gatherer(self, gatherer_id: UUID, gatherer_data: GathererUpdate) -> Optional[GathererResponse]:
        """Actualiza un gatherer existente"""
        db = self._get_db()
        try:
            gatherer = db.query(GathererModel).filter(
                GathererModel.id == gatherer_id,
                GathererModel.disabled_at.is_(None)
            ).first()
            
            if not gatherer:
                return None
            
            # Actualizar solo los campos que no son None
            update_data = gatherer_data.model_dump(exclude_none=True)
            for key, value in update_data.items():
                setattr(gatherer, key, value)
            
            gatherer.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(gatherer)
            
            # Obtener la fecha de última compra y calcular el estado
            from datetime import datetime, timedelta
            last_purchase_subquery = db.query(
                PurchaseModel.gatherer_id,
                func.max(PurchaseModel.purchase_date).label('last_purchase_date')
            ).filter(
                PurchaseModel.disabled_at.is_(None),
                PurchaseModel.gatherer_id == gatherer_id
            ).group_by(PurchaseModel.gatherer_id).subquery()
            
            result = db.query(
                GathererModel,
                last_purchase_subquery.c.last_purchase_date
            ).outerjoin(
                last_purchase_subquery,
                GathererModel.id == last_purchase_subquery.c.gatherer_id
            ).filter(
                GathererModel.id == gatherer_id
            ).first()
            
            last_purchase_date = result[1] if result else None
            
            # Calcular estado
            now = datetime.utcnow()
            fifteen_days_ago = now - timedelta(days=15)
            status = None
            if last_purchase_date:
                status = "activo" if last_purchase_date >= fifteen_days_ago else "inactivo"
            else:
                status = "inactivo"
            
            gatherer_response_dict = {
                'id': gatherer.id,
                'identity_id': gatherer.identity_id,
                'first_name': gatherer.first_name,
                'last_name': gatherer.last_name,
                'dni': gatherer.dni,
                'sms_number': gatherer.sms_number,
                'wsp_number': gatherer.wsp_number,
                'call_number': gatherer.call_number,
                'email': gatherer.email,
                'last_purchase_date': last_purchase_date,
                'status': status,
                'created_at': gatherer.created_at,
                'updated_at': gatherer.updated_at,
                'disabled_at': gatherer.disabled_at
            }
            
            return GathererResponse(**gatherer_response_dict)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_gatherers_paginated(self, page: int = 1, page_size: int = 10, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "", status: Optional[Literal["activo", "inactivo"]] = None, is_disabled: Optional[bool] = None) -> PaginateGathererResponse:
        """Obtiene gatherers paginados con fecha última compra y estado"""
        db = self._get_db()
        from datetime import datetime, timedelta
        
        # Subquery para obtener la fecha de la última compra por gatherer
        last_purchase_subquery = db.query(
            PurchaseModel.gatherer_id,
            func.max(PurchaseModel.purchase_date).label('last_purchase_date')
        ).filter(
            PurchaseModel.disabled_at.is_(None)
        ).group_by(PurchaseModel.gatherer_id).subquery()
        
        # Query principal con LEFT JOIN a la subquery
        query = db.query(
            GathererModel,
            last_purchase_subquery.c.last_purchase_date
        ).outerjoin(
            last_purchase_subquery,
            GathererModel.id == last_purchase_subquery.c.gatherer_id
        )
        
        # Aplicar filtro de is_disabled
        if is_disabled is not None:
            if is_disabled:
                # Mostrar solo los deshabilitados
                query = query.filter(GathererModel.disabled_at.isnot(None))
            else:
                # Mostrar solo los no deshabilitados
                query = query.filter(GathererModel.disabled_at.is_(None))
        # Si is_disabled es None, mostrar todos (sin filtro)
        
        # Aplicar filtro de estado si se proporciona
        if status:
            now = datetime.utcnow()
            fifteen_days_ago = now - timedelta(days=15)
            
            if status == "activo":
                # Activo: última compra dentro de los últimos 15 días
                query = query.filter(
                    last_purchase_subquery.c.last_purchase_date.isnot(None),
                    last_purchase_subquery.c.last_purchase_date >= fifteen_days_ago
                )
            elif status == "inactivo":
                # Inactivo: sin compras o última compra hace más de 15 días
                query = query.filter(
                    (last_purchase_subquery.c.last_purchase_date.is_(None)) |
                    (last_purchase_subquery.c.last_purchase_date < fifteen_days_ago)
                )
        
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(
                GathererModel.first_name.ilike(f"%{search}%") |
                GathererModel.last_name.ilike(f"%{search}%") |
                GathererModel.dni.ilike(f"%{search}%")
            )
        
        # Aplicar ordenamiento
        if sort_by:
            if sort_by == "last_purchase_date":
                # Ordenar por la fecha de última compra
                if order and order.lower() == "desc":
                    query = query.order_by(last_purchase_subquery.c.last_purchase_date.desc().nullslast())
                else:
                    query = query.order_by(last_purchase_subquery.c.last_purchase_date.asc().nullslast())
            elif sort_by == "status":
                # Ordenar por estado (calculado)
                # Primero calculamos el estado y luego ordenamos
                pass  # Se manejará después de obtener los resultados
            else:
                sort_column = getattr(GathererModel, sort_by, None)
                if sort_column is not None:
                    if order and order.lower() == "desc":
                        query = query.order_by(sort_column.desc())
                    else:
                        query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(GathererModel.created_at.desc())
        
        total = query.count()
        offset = (page - 1) * page_size
        total_pages = (total + page_size - 1) // page_size
        
        results = query.offset(offset).limit(page_size).all()
        
        # Construir respuestas con campos calculados
        gatherer_responses = []
        now = datetime.utcnow()
        fifteen_days_ago = now - timedelta(days=15)
        
        for gatherer, last_purchase_date in results:
            # Calcular estado
            status = None
            if last_purchase_date:
                status = "activo" if last_purchase_date >= fifteen_days_ago else "inactivo"
            else:
                status = "inactivo"  # Si no tiene compras, es inactivo
            
            gatherer_dict = {
                'id': gatherer.id,
                'identity_id': gatherer.identity_id,
                'first_name': gatherer.first_name,
                'last_name': gatherer.last_name,
                'dni': gatherer.dni,
                'sms_number': gatherer.sms_number,
                'wsp_number': gatherer.wsp_number,
                'call_number': gatherer.call_number,
                'email': gatherer.email,
                'last_purchase_date': last_purchase_date,
                'status': status,
                'created_at': gatherer.created_at,
                'updated_at': gatherer.updated_at,
                'disabled_at': gatherer.disabled_at
            }
            gatherer_responses.append(GathererResponse(**gatherer_dict))
        
        # Si se ordenó por status, ordenar los resultados
        if sort_by == "status":
            reverse = order and order.lower() == "desc"
            gatherer_responses.sort(key=lambda x: x.status or "", reverse=reverse)
        
        return PaginateGathererResponse(
            items=gatherer_responses,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=total_pages
        )
    
    def get_gatherer_by_id(self, gatherer_id: UUID) -> Optional[GathererResponse]:
        """Obtiene un gatherer específico por ID (solo si no está deshabilitado) con fecha última compra y estado"""
        db = self._get_db()
        from datetime import datetime, timedelta
        
        # Subquery para obtener la fecha de la última compra
        last_purchase_subquery = db.query(
            PurchaseModel.gatherer_id,
            func.max(PurchaseModel.purchase_date).label('last_purchase_date')
        ).filter(
            PurchaseModel.disabled_at.is_(None),
            PurchaseModel.gatherer_id == gatherer_id
        ).group_by(PurchaseModel.gatherer_id).subquery()
        
        # Query principal con LEFT JOIN
        result = db.query(
            GathererModel,
            last_purchase_subquery.c.last_purchase_date
        ).outerjoin(
            last_purchase_subquery,
            GathererModel.id == last_purchase_subquery.c.gatherer_id
        ).filter(
            GathererModel.id == gatherer_id,
            GathererModel.disabled_at.is_(None)
        ).first()
        
        if not result:
            return None
        
        gatherer, last_purchase_date = result
        
        # Calcular estado
        now = datetime.utcnow()
        fifteen_days_ago = now - timedelta(days=15)
        status = None
        if last_purchase_date:
            status = "activo" if last_purchase_date >= fifteen_days_ago else "inactivo"
        else:
            status = "inactivo"
        
        gatherer_dict = {
            'id': gatherer.id,
            'identity_id': gatherer.identity_id,
            'first_name': gatherer.first_name,
            'last_name': gatherer.last_name,
            'dni': gatherer.dni,
            'sms_number': gatherer.sms_number,
            'wsp_number': gatherer.wsp_number,
            'call_number': gatherer.call_number,
            'email': gatherer.email,
            'last_purchase_date': last_purchase_date,
            'status': status,
            'created_at': gatherer.created_at,
            'updated_at': gatherer.updated_at,
            'disabled_at': gatherer.disabled_at
        }
        
        return GathererResponse(**gatherer_dict)
    
    def disable_gatherer(self, gatherer_id: UUID) -> Optional[GathererResponse]:
        """Deshabilita un gatherer"""
        db = self._get_db()
        try:
            gatherer = db.query(GathererModel).filter(
                GathererModel.id == gatherer_id,
                GathererModel.disabled_at.is_(None)
            ).first()
            
            if not gatherer:
                return None
            
            gatherer.disabled_at = datetime.now()
            gatherer.updated_at = datetime.utcnow()
            
            db.commit()
            db.refresh(gatherer)
            return GathererResponse.model_validate(gatherer)
        except Exception as e:
            db.rollback()
            raise e
    
    def restore_gatherer(self, gatherer_id: UUID) -> Optional[GathererResponse]:
        """Restaura un gatherer deshabilitado"""
        db = self._get_db()
        try:
            gatherer = db.query(GathererModel).filter(
                GathererModel.id == gatherer_id,
                GathererModel.disabled_at.isnot(None)
            ).first()
            
            if not gatherer:
                return None
            
            gatherer.disabled_at = None
            gatherer.updated_at = datetime.utcnow()
            
            db.commit()
            db.refresh(gatherer)
            return GathererResponse.model_validate(gatherer)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_gatherers_by_gathering_center(self, gathering_center_id: UUID, page: int = 1, per_page: int = 10, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "") -> PaginateGathererByGatheringCenterResponse:
        """Obtiene gatherers filtrados por gathering_center_id con información adicional"""
        db = self._get_db()
        
        # Query base: gatherers que pertenecen al gathering center específico
        # Modificado para obtener también el ID de la relación
        query = db.query(
            GathererModel,
            GathererGatheringCenterModel.id.label('relation_id')
        ).join(
            GathererGatheringCenterModel,
            GathererModel.id == GathererGatheringCenterModel.gatherer_id
        ).filter(
            GathererGatheringCenterModel.gathering_center_id == gathering_center_id,
            GathererGatheringCenterModel.disabled_at.is_(None),
            GathererModel.disabled_at.is_(None)
        )
        
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(
                GathererModel.first_name.ilike(f"%{search}%") |
                GathererModel.last_name.ilike(f"%{search}%") |
                GathererModel.dni.ilike(f"%{search}%")
            )
        
        # Aplicar ordenamiento
        if sort_by:
            sort_column = getattr(GathererModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(GathererModel.first_name.asc())
        
        total = query.count()
        offset = (page - 1) * per_page
        total_pages = (total + per_page - 1) // per_page
        
        results = query.offset(offset).limit(per_page).all()
        
        # Construir respuesta con otros centros de acopio
        items = []
        for gatherer, relation_id in results:
            # Obtener otros centros de acopio (excluyendo el actual)
            other_centers = db.query(GatheringCenterModel.name).join(
                GathererGatheringCenterModel,
                GatheringCenterModel.id == GathererGatheringCenterModel.gathering_center_id
            ).filter(
                GathererGatheringCenterModel.gatherer_id == gatherer.id,
                GathererGatheringCenterModel.gathering_center_id != gathering_center_id,
                GathererGatheringCenterModel.disabled_at.is_(None),
                GatheringCenterModel.disabled_at.is_(None)
            ).all()
            
            other_center_names = [center[0] for center in other_centers]
            
            items.append(GathererByGatheringCenterResponse(
                id=gatherer.id,
                gatherer_gathering_center_id=relation_id,
                first_name=gatherer.first_name,
                last_name=gatherer.last_name,
                dni=gatherer.dni,
                call_number=gatherer.call_number,
                other_gathering_centers=other_center_names
            ))
        
        return PaginateGathererByGatheringCenterResponse(
            items=items,
            total=total,
            page=page,
            page_size=per_page,
            total_pages=total_pages
        )
    
    def export_gatherers_by_gathering_center_to_excel(self, gathering_center_id: UUID, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = ""):
        """Exporta todos los gatherers de un centro de acopio a Excel con los mismos datos que la paginación"""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        db = self._get_db()
        
        # Query base: gatherers que pertenecen al gathering center específico
        query = db.query(
            GathererModel,
            GathererGatheringCenterModel.id.label('relation_id')
        ).join(
            GathererGatheringCenterModel,
            GathererModel.id == GathererGatheringCenterModel.gatherer_id
        ).filter(
            GathererGatheringCenterModel.gathering_center_id == gathering_center_id,
            GathererGatheringCenterModel.disabled_at.is_(None),
            GathererModel.disabled_at.is_(None)
        )
        
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(
                GathererModel.first_name.ilike(f"%{search}%") |
                GathererModel.last_name.ilike(f"%{search}%") |
                GathererModel.dni.ilike(f"%{search}%")
            )
        
        # Aplicar ordenamiento
        if sort_by:
            sort_column = getattr(GathererModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(GathererModel.first_name.asc())
        
        # Obtener todos los resultados (sin paginación)
        results = query.all()
        
        # Obtener nombre del centro de acopio
        gathering_center = db.query(GatheringCenterModel).filter(
            GatheringCenterModel.id == gathering_center_id
        ).first()
        center_name = gathering_center.name if gathering_center else "Centro de Acopio"
        
        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Acopiadores - {center_name}"[:31]  # Límite de 31 caracteres para nombres de hojas
        
        # Estilo para el encabezado
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados (SIN IDs, solo nombres)
        headers = [
            "Nombres",
            "Apellidos",
            "DNI",
            "Teléfono",
            "Otros Centros de Acopio"
        ]
        ws.append(headers)
        
        # Aplicar estilo al encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Agregar datos (SIN IDs)
        for gatherer, relation_id in results:
            # Obtener otros centros de acopio (excluyendo el actual)
            other_centers = db.query(GatheringCenterModel.name).join(
                GathererGatheringCenterModel,
                GatheringCenterModel.id == GathererGatheringCenterModel.gathering_center_id
            ).filter(
                GathererGatheringCenterModel.gatherer_id == gatherer.id,
                GathererGatheringCenterModel.gathering_center_id != gathering_center_id,
                GathererGatheringCenterModel.disabled_at.is_(None),
                GatheringCenterModel.disabled_at.is_(None)
            ).all()
            
            other_center_names = ", ".join([center[0] for center in other_centers]) if other_centers else ""
            
            ws.append([
                gatherer.first_name,
                gatherer.last_name,
                gatherer.dni,
                gatherer.call_number or "",
                other_center_names
            ])
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    # ========== GATHERER GATHERING CENTER METHODS ==========
    def create_gatherer_gathering_center(self, data: GathererGatheringCenterCreate) -> GathererGatheringCenterResponse:
        """Crea una nueva relación acopiador-centro de acopio"""
        db = self._get_db()
        try:
            # Verificar que no exista una relación activa
            existing = db.query(GathererGatheringCenterModel).filter(
                GathererGatheringCenterModel.gatherer_id == data.gatherer_id,
                GathererGatheringCenterModel.gathering_center_id == data.gathering_center_id,
                GathererGatheringCenterModel.disabled_at.is_(None)
            ).first()
            
            if existing:
                raise ValueError("Ya existe una relación activa entre este acopiador y centro de acopio")
            
            relation = GathererGatheringCenterModel(**data.model_dump())
            db.add(relation)
            db.commit()
            db.refresh(relation)
            return GathererGatheringCenterResponse.model_validate(relation)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_gatherer_gathering_centers_paginated(self, page: int = 1, per_page: int = 10, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "") -> PaginatedGathererGatheringCenterResponse:
        """Obtiene relaciones acopiador-centro de acopio paginadas (solo las no deshabilitadas)"""
        db = self._get_db()
        
        query = db.query(GathererGatheringCenterModel).filter(GathererGatheringCenterModel.disabled_at.is_(None))
        
        # Aplicar búsqueda si se proporciona (búsqueda por IDs)
        if search:
            # Intentar buscar por UUID si el search es un UUID válido
            try:
                search_uuid = UUID(search)
                query = query.filter(
                    (GathererGatheringCenterModel.gatherer_id == search_uuid) |
                    (GathererGatheringCenterModel.gathering_center_id == search_uuid)
                )
            except ValueError:
                # Si no es un UUID válido, no aplicar filtro de búsqueda
                pass
        
        # Aplicar ordenamiento
        if sort_by:
            sort_column = getattr(GathererGatheringCenterModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(GathererGatheringCenterModel.created_at.desc())
        
        total = query.count()
        offset = (page - 1) * per_page
        total_pages = (total + per_page - 1) // per_page
        
        relations = query.offset(offset).limit(per_page).all()
        
        return PaginatedGathererGatheringCenterResponse(
            items=[GathererGatheringCenterResponse.model_validate(rel) for rel in relations],
            total=total,
            page=page,
            page_size=per_page,
            total_pages=total_pages
        )
    
    def get_gatherer_gathering_center_by_id(self, relation_id: UUID) -> Optional[GathererGatheringCenterResponse]:
        """Obtiene una relación acopiador-centro de acopio específica por ID (solo si no está deshabilitada)"""
        db = self._get_db()
        relation = db.query(GathererGatheringCenterModel).filter(
            GathererGatheringCenterModel.id == relation_id,
            GathererGatheringCenterModel.disabled_at.is_(None)
        ).first()
        
        if not relation:
            return None
        
        return GathererGatheringCenterResponse.model_validate(relation)
    
    def update_gatherer_gathering_center(self, relation_id: UUID, data: GathererGatheringCenterUpdate) -> Optional[GathererGatheringCenterResponse]:
        """Actualiza una relación acopiador-centro de acopio existente"""
        db = self._get_db()
        try:
            relation = db.query(GathererGatheringCenterModel).filter(
                GathererGatheringCenterModel.id == relation_id,
                GathererGatheringCenterModel.disabled_at.is_(None)
            ).first()
            
            if not relation:
                return None
            
            # Si se actualiza gatherer_id o gathering_center_id, verificar que no exista otra relación activa
            update_data = data.model_dump(exclude_none=True)
            if 'gatherer_id' in update_data or 'gathering_center_id' in update_data:
                new_gatherer_id = update_data.get('gatherer_id', relation.gatherer_id)
                new_center_id = update_data.get('gathering_center_id', relation.gathering_center_id)
                
                existing = db.query(GathererGatheringCenterModel).filter(
                    GathererGatheringCenterModel.gatherer_id == new_gatherer_id,
                    GathererGatheringCenterModel.gathering_center_id == new_center_id,
                    GathererGatheringCenterModel.id != relation_id,
                    GathererGatheringCenterModel.disabled_at.is_(None)
                ).first()
                
                if existing:
                    raise ValueError("Ya existe una relación activa entre este acopiador y centro de acopio")
            
            for key, value in update_data.items():
                setattr(relation, key, value)
            
            db.commit()
            db.refresh(relation)
            return GathererGatheringCenterResponse.model_validate(relation)
        except Exception as e:
            db.rollback()
            raise e
    
    def disable_gatherer_gathering_center(self, relation_id: UUID) -> Optional[GathererGatheringCenterResponse]:
        """Deshabilita una relación acopiador-centro de acopio"""
        db = self._get_db()
        try:
            relation = db.query(GathererGatheringCenterModel).filter(
                GathererGatheringCenterModel.id == relation_id,
                GathererGatheringCenterModel.disabled_at.is_(None)
            ).first()
            
            if not relation:
                return None
            
            from datetime import datetime
            relation.disabled_at = datetime.now()
            
            db.commit()
            db.refresh(relation)
            return GathererGatheringCenterResponse.model_validate(relation)
        except Exception as e:
            db.rollback()
            raise e
    
    # ========== BALANCE MOVEMENT METHODS ==========
    def create_balance(self, data: BalanceMovementCreate) -> BalanceMovementResponse:
        """Crea un nuevo movimiento de balance"""
        db = self._get_db()
        
        try:
            balance = BalanceMovementModel(**data.dict())
            db.add(balance)
            db.commit()
            db.refresh(balance)
            
            # Cargar relaciones
            balance_dict = {
                "id": balance.id,
                "gathering_center_id": balance.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, balance.gathering_center_id),
                "gatherer_id": balance.gatherer_id,
                "gatherer": self._load_gatherer(db, balance.gatherer_id),
                "type_movement": balance.type_movement,
                "purchase_id": balance.purchase_id,
                "ammount": balance.ammount,
                "identity_id": balance.identity_id,
                "identity": self._load_identity(db, balance.identity_id),
                "created_at": balance.created_at,
                "disabled_at": balance.disabled_at
            }
            
            return BalanceMovementResponse(**balance_dict)
        except Exception as e:
            db.rollback()
            raise e
        
    def get_balances(self, gathering_center_id: Optional[UUID] = None, gatherer_id: Optional[UUID] = None, type_movement: Optional[BalanceMovementTypeEnum] = None, page: int = 1, page_size: int = 10) -> PaginatedBalanceMovementResponse:
        """Obtiene movimientos de balance paginados con filtros opcionales"""
        db = self._get_db()
        
        query = db.query(BalanceMovementModel).filter(BalanceMovementModel.disabled_at.is_(None))
        
        if gathering_center_id:
            query = query.filter(BalanceMovementModel.gathering_center_id == gathering_center_id)
        
        if gatherer_id:
            query = query.filter(BalanceMovementModel.gatherer_id == gatherer_id)
        
        if type_movement:
            query = query.filter(BalanceMovementModel.type_movement == type_movement)
        
        query = query.order_by(BalanceMovementModel.created_at.desc())
        
        total = query.count()
        offset = (page - 1) * page_size
        total_pages = (total + page_size - 1) // page_size
        
        movements = query.offset(offset).limit(page_size).all()
        
        # Construir respuestas con relaciones cargadas
        movement_responses = []
        for movement in movements:
            movement_dict = {
                "id": movement.id,
                "gathering_center_id": movement.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, movement.gathering_center_id),
                "gatherer_id": movement.gatherer_id,
                "gatherer": self._load_gatherer(db, movement.gatherer_id),
                "type_movement": movement.type_movement,
                "purchase_id": movement.purchase_id,
                "ammount": movement.ammount,
                "identity_id": movement.identity_id,
                "identity": self._load_identity(db, movement.identity_id),
                "created_at": movement.created_at,
                "disabled_at": movement.disabled_at
            }
            movement_responses.append(BalanceMovementResponse(**movement_dict))
        
        return PaginatedBalanceMovementResponse(
            items=movement_responses,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=total_pages
        )
    
    def export_balances_to_excel(self, gathering_center_id: Optional[UUID] = None, gatherer_id: Optional[UUID] = None, type_movement: Optional[BalanceMovementTypeEnum] = None):
        """Exporta todos los movimientos de balance a Excel con los mismos datos que la paginación"""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        db = self._get_db()
        
        query = db.query(BalanceMovementModel).filter(BalanceMovementModel.disabled_at.is_(None))
        
        if gathering_center_id:
            query = query.filter(BalanceMovementModel.gathering_center_id == gathering_center_id)
        
        if gatherer_id:
            query = query.filter(BalanceMovementModel.gatherer_id == gatherer_id)
        
        if type_movement:
            query = query.filter(BalanceMovementModel.type_movement == type_movement)
        
        query = query.order_by(BalanceMovementModel.created_at.desc())
        
        # Obtener todos los resultados (sin paginación)
        movements = query.all()
        
        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos de Caja"
        
        # Estilo para el encabezado
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados (SIN IDs, solo nombres)
        headers = [
            "Centro de Acopio",
            "Acopiador",
            "Tipo de Movimiento",
            "Monto",
            "Identidad",
            "Fecha de Creación",
            "Número Ticket Compra"
        ]
        ws.append(headers)
        
        # Aplicar estilo al encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Agregar datos (SIN IDs)
        for movement in movements:
            # Obtener nombres de las relaciones
            gathering_center = db.query(GatheringCenterModel).filter(
                GatheringCenterModel.id == movement.gathering_center_id
            ).first() if movement.gathering_center_id else None
            
            gatherer = db.query(GathererModel).filter(
                GathererModel.id == movement.gatherer_id
            ).first() if movement.gatherer_id else None
            
            identity = db.query(IdentityModel).filter(
                IdentityModel.id == movement.identity_id
            ).first() if movement.identity_id else None
            
            # Obtener número de ticket si hay purchase_id
            ticket_number = ""
            if movement.purchase_id:
                purchase = db.query(PurchaseModel).filter(
                    PurchaseModel.id == movement.purchase_id
                ).first()
                if purchase and purchase.ticket_number:
                    ticket_number = purchase.ticket_number
            
            ws.append([
                resolve_display_name(gathering_center) if gathering_center else "",
                resolve_display_name(gatherer) if gatherer else "",
                movement.type_movement.value if movement.type_movement else "",
                float(movement.ammount) if movement.ammount else 0.0,
                resolve_display_name(identity) if identity else "",
                movement.created_at.strftime("%Y-%m-%d %H:%M:%S") if movement.created_at else "",
                ticket_number
            ])
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def get_balance_summary(self, gathering_center_id: UUID, gatherer_id: Optional[UUID]) -> BalanceSummaryResponse:
        """Obtiene el resumen de balance para un acopiador en un centro de acopio"""
        db = self._get_db()
        
        query = db.query(BalanceMovementModel).filter(
            BalanceMovementModel.gathering_center_id == gathering_center_id,
            BalanceMovementModel.disabled_at.is_(None)
        )

        if gatherer_id:
            query = query.filter(
                BalanceMovementModel.gatherer_id == gatherer_id
            )

        movements = query.order_by(
            BalanceMovementModel.created_at.desc()
        ).all()
        
        # Calcular balance total
        total_balance = sum(float(m.ammount) if m.type_movement == BalanceMovementTypeEnum.RECHARGE else -float(m.ammount) for m in movements)
        
        # Construir respuestas con relaciones cargadas
        movement_responses = []
        for m in movements:
            movement_dict = {
                "id": m.id,
                "gathering_center_id": m.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, m.gathering_center_id),
                "gatherer_id": m.gatherer_id,
                "gatherer": self._load_gatherer(db, m.gatherer_id),
                "type_movement": m.type_movement,
                "purchase_id": m.purchase_id,
                "ammount": m.ammount,
                "identity_id": m.identity_id,
                "identity": self._load_identity(db, m.identity_id),
                "created_at": m.created_at,
                "disabled_at": m.disabled_at
            }
            movement_responses.append(BalanceMovementResponse(**movement_dict))
        
        return BalanceSummaryResponse(
            gathering_center_id=gathering_center_id,
            gathering_center=self._load_gathering_center(db, gathering_center_id),
            gatherer_id=gatherer_id,
            gatherer=self._load_gatherer(db, gatherer_id) if gatherer_id else None,
            total_balance=total_balance,
            movements=movement_responses
        )
    
    def get_balance_gatherers_summary(self, gatherer_id: Optional[UUID]) -> BalanceSummaryResponse:
        """Obtiene el resumen de balance para un acopiador en un centro de acopio"""
        db = self._get_db()
        
        query = db.query(BalanceMovementModel).filter(
            BalanceMovementModel.gatherer_id.isnot(None),
            BalanceMovementModel.disabled_at.is_(None)
        )

        if gatherer_id:
            query = query.filter(
                BalanceMovementModel.gatherer_id == gatherer_id,
                BalanceMovementModel.type_movement == BalanceMovementTypeEnum.PURCHASE
            )

        movements = query.order_by(
            BalanceMovementModel.created_at.desc()
        ).all()
        
        # Calcular balance total
        total_balance = sum(float(m.ammount) if m.type_movement == BalanceMovementTypeEnum.RECHARGE else -float(m.ammount) for m in movements)
        today = datetime.today().date()
        first_day_month = today.replace(day=1)
        return BalanceSummaryGatherersResponse(
                    total_balance=total_balance,
                    average_balance=total_balance / len(movements) if movements else 0,
                    daily_amount=sum(
                        float(m.ammount) if m.type_movement == BalanceMovementTypeEnum.RECHARGE else -float(m.ammount)
                        for m in movements
                        if m.created_at.date() == today
                    ),
                    monthly_amount=sum(
                        float(m.ammount) if m.type_movement == BalanceMovementTypeEnum.RECHARGE else -float(m.ammount)
                        for m in movements
                        if m.created_at.date() >= first_day_month
                    )
                )
    
    # ========== LOT CERTIFICATION METHODS ==========
    def get_certifications_by_lot(self, lot_id: UUID) -> List[LotCertificationWithDetailsResponse]:
        """Obtiene las certificaciones de un lote (solo las activas)"""
        db = self._get_db()
        
        lot_certs = db.query(LotCertificationModel).join(
            CertificationModel, LotCertificationModel.certification_id == CertificationModel.id
        ).filter(
            LotCertificationModel.lot_id == lot_id,
            LotCertificationModel.disabled_at.is_(None)
        ).all()
        
        result = []
        for lot_cert in lot_certs:
            cert = db.query(CertificationModel).filter(CertificationModel.id == lot_cert.certification_id).first()
            result.append(LotCertificationWithDetailsResponse(
                id=lot_cert.id,
                lot_id=lot_cert.lot_id,
                certification_id=lot_cert.certification_id,
                certification_name=cert.name if cert else "",
                certification_code=cert.code if cert else None,
                created_at=lot_cert.created_at,
                disabled_at=lot_cert.disabled_at
            ))
        
        return result
    
    # ========== PURCHASE METHODS ==========
    def create_purchase(self, purchase_data: PurchaseCreate) -> PurchaseResponse:
        """Crea una nueva compra (el balance_movement se genera automáticamente por trigger)"""
        db = self._get_db()
        try:
            ticket_number = self._generar_numero_recibo_fecha_timestamp_aleatorio()
            purchase_data.ticket_number = ticket_number
            purchase = PurchaseModel(**purchase_data.model_dump(exclude_none=True))
            db.add(purchase)
            db.commit()
            db.refresh(purchase)
            
            # Cargar relaciones
            purchase_dict = {
                "id": purchase.id,
                "lot_id": purchase.lot_id,
                "farmer_id": purchase.farmer_id,
                "farmer": self._load_farmer(db, purchase.farmer_id),
                "farm_id": purchase.farm_id,
                "farm": self._load_farm(db, purchase.farm_id),
                "gatherer_id": purchase.gatherer_id,
                "gatherer": self._load_gatherer(db, purchase.gatherer_id),
                "quantity": purchase.quantity,
                "price": purchase.price,
                "presentation": purchase.presentation,
                "payment_method": purchase.payment_method,
                "purchase_date": purchase.purchase_date,
                "ticket_number": purchase.ticket_number,
                "gathering_center_id": purchase.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, purchase.gathering_center_id),
                "identity_id": purchase.identity_id,
                "identity": self._load_identity(db, purchase.identity_id),
                "created_at": purchase.created_at,
                "updated_at": purchase.updated_at,
                "disabled_at": purchase.disabled_at
            }
            
            return PurchaseResponse(**purchase_dict)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_purchases_paginated(self, page: int = 1, per_page: int = 10, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "") -> PaginatedPurchaseResponse:
        """Obtiene compras paginadas (solo las no deshabilitadas)"""
        db = self._get_db()
        
        query = db.query(PurchaseModel).filter(PurchaseModel.disabled_at.is_(None))
        
        # Aplicar búsqueda si se proporciona (búsqueda por IDs o payment_method)
        if search:
            # Intentar buscar por UUID si el search es un UUID válido
            try:
                search_uuid = UUID(search)
                query = query.filter(
                    (PurchaseModel.lot_id == search_uuid) |
                    (PurchaseModel.farmer_id == search_uuid) |
                    (PurchaseModel.farm_id == search_uuid) |
                    (PurchaseModel.gatherer_id == search_uuid)
                )
            except ValueError:
                # Si no es un UUID válido, buscar por payment_method
                query = query.filter(
                    (PurchaseModel.payment_method.isnot(None) & PurchaseModel.payment_method.ilike(f"%{search}%"))
                )
        
        # Aplicar ordenamiento
        if sort_by:
            sort_column = getattr(PurchaseModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(PurchaseModel.created_at.desc())
        
        total = query.count()
        offset = (page - 1) * per_page
        total_pages = (total + per_page - 1) // per_page
        
        purchases = query.offset(offset).limit(per_page).all()
        
        # Construir respuestas con relaciones cargadas
        purchase_responses = []
        for purchase in purchases:
            purchase_dict = {
                "id": purchase.id,
                "lot_id": purchase.lot_id,
                "farmer_id": purchase.farmer_id,
                "farmer": self._load_farmer(db, purchase.farmer_id),
                "farm_id": purchase.farm_id,
                "farm": self._load_farm(db, purchase.farm_id),
                "gatherer_id": purchase.gatherer_id,
                "gatherer": self._load_gatherer(db, purchase.gatherer_id),
                "quantity": purchase.quantity,
                "price": purchase.price,
                "presentation": purchase.presentation,
                "payment_method": purchase.payment_method,
                "purchase_date": purchase.purchase_date,
                "ticket_number": purchase.ticket_number,
                "gathering_center_id": purchase.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, purchase.gathering_center_id),
                "identity_id": purchase.identity_id,
                "identity": self._load_identity(db, purchase.identity_id),
                "created_at": purchase.created_at,
                "updated_at": purchase.updated_at,
                "disabled_at": purchase.disabled_at
            }
            purchase_responses.append(PurchaseResponse(**purchase_dict))
        
        return PaginatedPurchaseResponse(
            items=purchase_responses,
            total=total,
            page=page,
            page_size=per_page,
            total_pages=total_pages
        )
    
    def get_purchase_by_id(self, purchase_id: UUID) -> Optional[PurchaseResponse]:
        """Obtiene una compra específica por ID (solo si no está deshabilitada)"""
        db = self._get_db()
        purchase = db.query(PurchaseModel).filter(
            PurchaseModel.id == purchase_id,
            PurchaseModel.disabled_at.is_(None)
        ).first()
        
        if not purchase:
            return None
        
        # Cargar relaciones
        purchase_dict = {
            "id": purchase.id,
            "lot_id": purchase.lot_id,
            "farmer_id": purchase.farmer_id,
            "farmer": self._load_farmer(db, purchase.farmer_id),
            "farm_id": purchase.farm_id,
            "farm": self._load_farm(db, purchase.farm_id),
            "gatherer_id": purchase.gatherer_id,
            "gatherer": self._load_gatherer(db, purchase.gatherer_id),
            "quantity": purchase.quantity,
            "price": purchase.price,
            "presentation": purchase.presentation,
            "payment_method": purchase.payment_method,
            "purchase_date": purchase.purchase_date,
            "ticket_number": purchase.ticket_number,
            "gathering_center_id": purchase.gathering_center_id,
            "gathering_center": self._load_gathering_center(db, purchase.gathering_center_id),
            "identity_id": purchase.identity_id,
            "identity": self._load_identity(db, purchase.identity_id),
            "created_at": purchase.created_at,
            "updated_at": purchase.updated_at,
            "disabled_at": purchase.disabled_at
        }
        
        return PurchaseResponse(**purchase_dict)
    
    def update_purchase(self, purchase_id: UUID, purchase_data: PurchaseUpdate) -> Optional[PurchaseResponse]:
        """Actualiza una compra existente"""
        db = self._get_db()
        try:
            purchase = db.query(PurchaseModel).filter(
                PurchaseModel.id == purchase_id,
                PurchaseModel.disabled_at.is_(None)
            ).first()
            
            if not purchase:
                return None
            
            update_data = purchase_data.model_dump(exclude_none=True)
            for key, value in update_data.items():
                setattr(purchase, key, value)
            
            purchase.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(purchase)
            
            # Cargar relaciones
            purchase_dict = {
                "id": purchase.id,
                "lot_id": purchase.lot_id,
                "farmer_id": purchase.farmer_id,
                "farmer": self._load_farmer(db, purchase.farmer_id),
                "farm_id": purchase.farm_id,
                "farm": self._load_farm(db, purchase.farm_id),
                "gatherer_id": purchase.gatherer_id,
                "gatherer": self._load_gatherer(db, purchase.gatherer_id),
                "quantity": purchase.quantity,
                "price": purchase.price,
                "presentation": purchase.presentation,
                "payment_method": purchase.payment_method,
                "purchase_date": purchase.purchase_date,
                "ticket_number": purchase.ticket_number,
                "gathering_center_id": purchase.gathering_center_id,
                "gathering_center": self._load_gathering_center(db, purchase.gathering_center_id),
                "identity_id": purchase.identity_id,
                "identity": self._load_identity(db, purchase.identity_id),
                "created_at": purchase.created_at,
                "updated_at": purchase.updated_at,
                "disabled_at": purchase.disabled_at
            }
            
            return PurchaseResponse(**purchase_dict)
        except Exception as e:
            db.rollback()
            raise e
    
    # ========== STORE MOVEMENT METHODS ==========
    def dispatch_lots(self, dispatch_data: DispatchLotsRequest, identity_id: Optional[UUID] = None) -> DispatchLotsResponse:
        """
        Despacha lotes a un centro de almacenamiento.
        
        Para cada lote:
        1. Actualiza current_store_center_id del lote
        2. Crea un store_movement de tipo INGRESO en la tabla del módulo warehouse
        
        Args:
            dispatch_data: Datos del despacho (lot_ids, store_center_id)
            identity_id: ID de la identidad que registra el movimiento (del token)
            
        Returns:
            DispatchLotsResponse con el mensaje y cantidad despachada
        """
        db = self._get_db()
        
        try:
            # Obtener TODOS los lotes en una sola consulta (optimizado para listas largas)
            lots = db.query(LotModel).filter(
                LotModel.id.in_(dispatch_data.lot_ids),
                LotModel.disabled_at.is_(None)
            ).all()
            
            # Crear diccionario de lotes encontrados para acceso rápido
            lots_dict = {lot.id: lot for lot in lots}
            
            # Reportar lotes no encontrados o deshabilitados
            not_found_ids = set(dispatch_data.lot_ids) - set(lots_dict.keys())
            if not_found_ids:
                print(f"⚠️  {len(not_found_ids)} lote(s) no encontrado(s) o deshabilitado(s): {list(not_found_ids)[:5]}{'...' if len(not_found_ids) > 5 else ''}")
            
            dispatched_count = 0
            skipped_no_weight = []
            
            # Procesar cada lote encontrado
            for lot_id, lot in lots_dict.items():
                # Verificar que el lote tenga peso
                if not lot.net_weight or lot.net_weight <= 0:
                    skipped_no_weight.append(lot_id)
                    continue
                
                # Actualizar el current_store_center_id del lote
                lot.current_store_center_id = dispatch_data.store_center_id
                
                # Crear el store_movement de tipo INGRESO
                store_movement = StoreMovementModel(
                    lot_id=lot_id,
                    store_center_id=dispatch_data.store_center_id,
                    type_movement=StoreMovementTypeEnum.INGRESO,
                    weight_kg=lot.net_weight,
                    identity_id=identity_id
                )
                db.add(store_movement)
                
                dispatched_count += 1
            
            # Reportar lotes sin peso
            if skipped_no_weight:
                print(f"⚠️  {len(skipped_no_weight)} lote(s) sin peso (net_weight) omitido(s): {skipped_no_weight[:5]}{'...' if len(skipped_no_weight) > 5 else ''}")
            
            # Commit de todos los cambios
            db.commit()
            
            print(f"✓ {dispatched_count} lote(s) despachado(s) exitosamente a centro {dispatch_data.store_center_id}")
            
            message = f"{dispatched_count} lote(s) despachado(s) exitosamente"
            
            return DispatchLotsResponse(
                message=message,
                dispatched_lots=dispatched_count
            )
            
        except Exception as e:
            db.rollback()
            print(f"❌ Error al despachar lotes: {e}")
            raise e

