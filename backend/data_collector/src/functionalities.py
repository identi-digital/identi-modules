import logging
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any, TypedDict, Union

logger = logging.getLogger(__name__)

logger = logging.getLogger(__name__)
from uuid import UUID
from sqlalchemy.orm import Session
from sqlalchemy import func, inspect
# Usar importaciones relativas para evitar problemas durante la inicialización del módulo
from .models.forms import FormModel, FormPurpose
from .models.action_tools import ActionToolModel
from .models.core_registers import CoreRegisterModel, RegisterStatus
from .models.schema_forms import SchemaFormModel
from .models.referencable_entities import ReferencableEntityModel
from .schemas import (
    FormCreate, FormUpdate, FormResponse, PaginatedFormResponse, FormWithSchemaResponse,
    ActionToolCreate, ActionToolUpdate, ActionToolResponse, PaginatedActionToolResponse,
    CoreRegisterCreate, CoreRegisterUpdate, CoreRegisterResponse, PaginatedCoreRegisterResponse,
    ReferencableEntityResponse, PaginatedReferencableEntityResponse,
    EntityDataItemResponse, PaginatedEntityDataResponse,
    UniqueFieldValidationResponse, UniqueFieldComplementaryValidationResponse
)
from .models.schema_forms import SchemaFormModel
from .resources.schema_processor import process_schema_add_data_input_to_metadata
from .resources.register_processor import (
    find_model_by_entity_name,
    extract_entity_data_from_detail,
    process_register_to_entity
)
from .resources.form_auto_creator import (
    get_logical_identifier_field,
    _get_logical_identifier_fields,
    get_schema_columns_for_template,
)
from .resources import resolve_display_name
import logging

logger = logging.getLogger(__name__)


# Type definitions for detail structure
class EntityValue(TypedDict, total=False):
    """Estructura de value cuando type='entity'"""
    id: str
    display_name: str


class DetailItem(TypedDict, total=False):
    """
    Estructura de cada item en el array detail.
    
    Campos principales:
    - name: Nombre del campo
    - value: Valor capturado
      * Si type='entity': EntityValue o List[EntityValue]
      * Otros tipos: Any (string, number, boolean, etc)
    - type: Tipo del campo ('entity', 'text', 'number', 'date', 'media', etc)
    
    Campos opcionales:
    - display_name: Nombre legible para mostrar
    - is_unique: Indica si el campo es único en la entidad
    - entity_display_value: Valor de display de entidad (legacy)
    """
    name: str                           # Nombre del campo
    value: Union[Any, EntityValue, List[EntityValue]]  # Valor capturado
    type: Optional[str]                 # Tipo del campo (entity, text, number, etc)
    type_value: Optional[str]           # Tipo de valor (legacy)
    type_list_value: Optional[str]      # Tipo para listas
    display_name: Optional[str]         # Nombre para mostrar
    is_unique: Optional[bool]           # Si el campo es único
    entity_display_value: Optional[str] # Valor de display de entidad (legacy)


DetailArray = List[DetailItem]


class Funcionalities:
    def __init__(self, container):
        self.container = container
        
    def _get_db(self) -> Session:
        """Obtiene la sesión de base de datos desde el container"""
        # El Container ahora maneja automáticamente la verificación y limpieza de transacciones abortadas
        return self.container.get("core_db", "databases")
    
    # Form methods
    def get_forms(self, page: int = 1, per_page: int = 10, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "") -> PaginatedFormResponse:
        """Lista todos los formularios paginados (solo los no archivados)"""
        db = self._get_db()
        
        query = db.query(FormModel).filter(FormModel.disabled_at.is_(None))
        
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(
                FormModel.name.ilike(f"%{search}%") |
                (FormModel.description.isnot(None) & FormModel.description.ilike(f"%{search}%"))
            )
        
        # Aplicar ordenamiento
        if sort_by:
            sort_column = getattr(FormModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(FormModel.created_at.desc())
        
        total = query.count()
        offset = (page - 1) * per_page
        
        forms = query.offset(offset).limit(per_page).all()
        
        return PaginatedFormResponse(
            items=[FormResponse.model_validate(form) for form in forms],
            total=total,
            page=page,
            per_page=per_page
        )
    
    def create_form(self, form_data: FormCreate) -> FormResponse:
        """Crea un nuevo formulario"""
        db = self._get_db()
        try:
            # Convertir enums a sus valores string para SQLAlchemy
            form_dict = form_data.model_dump(exclude_none=True, mode='python')
            # Convertir enums a valores string
            for key, value in form_dict.items():
                if hasattr(value, 'value'):  # Es un enum
                    form_dict[key] = value.value
                elif isinstance(value, list) and value:
                    # Para arrays de enums (como viewer)
                    form_dict[key] = [v.value if hasattr(v, 'value') else v for v in value]
            
            # Si se proporciona un ID, usarlo; si no, se generará automáticamente
            form = FormModel(**form_dict)
            db.add(form)
            db.commit()
            db.refresh(form)
            return FormResponse.model_validate(form)
        except Exception as e:
            db.rollback()
            raise e
    
    def update_form(self, form_id: UUID, form_data: FormUpdate) -> Optional[FormResponse]:
        """
        Actualiza un formulario existente.
        
        IMPORTANTE: No actualiza channel_name ni schema_id.
        - channel_name: No se puede cambiar después de crear el form
        - schema_id: Para actualizar el schema, usa la API POST /forms/{form_id}/schema
        """
        db = self._get_db()
        try:
            form = db.query(FormModel).filter(
                FormModel.id == form_id,
                FormModel.disabled_at.is_(None)
            ).first()
            
            if not form:
                return None
            
            # Excluir schema_id y channel_name del update
            # - schema_id debe actualizarse solo a través de la API de schema
            # - channel_name no se puede cambiar después de crear el form
            update_data = form_data.model_dump(exclude_none=True, exclude={'schema_id', 'channel_name'}, mode='python')
            # Convertir enums a valores string para SQLAlchemy
            for key, value in update_data.items():
                if hasattr(value, 'value'):  # Es un enum
                    update_data[key] = value.value
                elif isinstance(value, list) and value:
                    # Para arrays de enums (como viewer)
                    update_data[key] = [v.value if hasattr(v, 'value') else v for v in value]
            
            for key, value in update_data.items():
                setattr(form, key, value)
            
            form.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(form)
            return FormResponse.model_validate(form)
        except Exception as e:
            db.rollback()
            raise e
    
    def archive_form(self, form_id: UUID) -> bool:
        """Archiva un formulario (deshabilitado lógico)"""
        db = self._get_db()
        try:
            form = db.query(FormModel).filter(
                FormModel.id == form_id,
                FormModel.disabled_at.is_(None)
            ).first()
            
            if not form:
                return False
            
            form.disabled_at = datetime.utcnow()
            db.commit()
            return True
        except Exception as e:
            db.rollback()
            raise e
    
    # Action Tool methods
    def get_tools(self, page: int = 1, per_page: int = 10, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "") -> PaginatedActionToolResponse:
        """Lista todas las tools paginadas (solo las no deshabilitadas)"""
        db = self._get_db()
        
        query = db.query(ActionToolModel).filter(ActionToolModel.disabled_at.is_(None))
        
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(
                ActionToolModel.name.ilike(f"%{search}%") |
                (ActionToolModel.description.isnot(None) & ActionToolModel.description.ilike(f"%{search}%"))
            )
        
        # Aplicar ordenamiento
        if sort_by:
            sort_column = getattr(ActionToolModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(ActionToolModel.created_at.desc())
        
        total = query.count()
        offset = (page - 1) * per_page
        
        tools = query.offset(offset).limit(per_page).all()
        
        return PaginatedActionToolResponse(
            items=[ActionToolResponse.model_validate(tool) for tool in tools],
            total=total,
            page=page,
            per_page=per_page
        )
    
    def create_tool(self, tool_data: ActionToolCreate) -> ActionToolResponse:
        """Crea una nueva tool"""
        db = self._get_db()
        try:
            tool = ActionToolModel(**tool_data.model_dump(exclude_none=True))
            db.add(tool)
            db.commit()
            db.refresh(tool)
            return ActionToolResponse.model_validate(tool)
        except Exception as e:
            db.rollback()
            raise e
    
    def update_tool(self, tool_id: UUID, tool_data: ActionToolUpdate) -> Optional[ActionToolResponse]:
        """Actualiza una tool existente"""
        db = self._get_db()
        try:
            tool = db.query(ActionToolModel).filter(
                ActionToolModel.id == tool_id,
                ActionToolModel.disabled_at.is_(None)
            ).first()
            
            if not tool:
                return None
            
            update_data = tool_data.model_dump(exclude_none=True)
            for key, value in update_data.items():
                setattr(tool, key, value)
            
            tool.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(tool)
            return ActionToolResponse.model_validate(tool)
        except Exception as e:
            db.rollback()
            raise e
    
    # Core Register methods
    def _get_entity_representative_value(self, entity_name: str, entity_id: Any, representative_value: str, db) -> Optional[str]:
        """
        Obtiene el valor representativo de una entidad usando su representative_value template.
        
        Args:
            entity_name: Nombre de la entidad
            entity_id: ID de la entidad
            representative_value: Template del representative_value (ej: "{{first_name}} {{last_name}}")
            db: Sesión de base de datos
            
        Returns:
            Valor representativo construido o None si no se puede obtener
        """
        try:
            # Importar la función que encuentra el modelo por entity_name
            from .resources.form_auto_creator import find_model_by_entity_name
            
            # Buscar el modelo correspondiente
            model_class = find_model_by_entity_name(entity_name)
            if not model_class:
                print(f"⚠️  No se encontró modelo para entity_name: {entity_name}")
                return None
            
            # Buscar el registro en la BD
            entity_record = db.query(model_class).filter(
                model_class.id == entity_id
            ).first()
            
            if not entity_record:
                print(f"⚠️  No se encontró registro con id {entity_id} en {entity_name}")
                return None
            
            # Construir el valor representativo usando el template
            import re
            result = representative_value
            
            # Buscar todos los placeholders {{campo}}
            placeholders = re.findall(r'\{\{(\w+)\}\}', representative_value)
            
            for placeholder in placeholders:
                # Obtener el valor del atributo
                attr_value = getattr(entity_record, placeholder, None)
                if attr_value is not None:
                    result = result.replace(f'{{{{{placeholder}}}}}', str(attr_value))
                else:
                    result = result.replace(f'{{{{{placeholder}}}}}', '')
            
            # Limpiar espacios múltiples
            result = ' '.join(result.split())
            
            return result.strip() if result.strip() else None
            
        except Exception as e:
            print(f"❌ Error al obtener valor representativo de {entity_name}: {e}")
            return None
    
    def _enrich_detail_with_display_names(
        self, 
        detail: Optional[DetailArray], 
        schema_form_id: UUID, 
        db: Session
    ) -> Optional[DetailArray]:
        """
        Enriquece el detail agregando display_name y valores representativos de entidades.
        
        Args:
            detail: Lista de diccionarios con los datos capturados
            schema_form_id: ID del schema_form asociado
            db: Sesión de base de datos
            
        Returns:
            Lista de diccionarios enriquecida con display_name y entity_display_value
        """
        if not detail:
            return detail
        
        # Obtener el schema_form
        schema_form = db.query(SchemaFormModel).filter(
            SchemaFormModel.id == schema_form_id
        ).first()
        
        if not schema_form or not schema_form.schema:
            return detail
        
        # Crear mapas desde las instructions del schema
        display_name_map = {}
        entity_info_map = {}  # name -> {entity_name, representative_value}
        instructions = schema_form.schema.get('instructions', [])
        
        for instruction_data in instructions:
            if isinstance(instruction_data, dict):
                name = instruction_data.get('name')
                display_name = instruction_data.get('display_name')
                
                if name and display_name:
                    display_name_map[name] = display_name
                
                # Verificar si es una entidad (type_value == "entity")
                schema_gather = instruction_data.get('schema_gather', {})
                type_value = schema_gather.get('type_value')
                
                if type_value == 'entity' and name:
                    # Buscar el entity_type en schema_input
                    schema_input = instruction_data.get('schema_input', [])
                    entity_type_value = None
                    
                    # Buscar el input con name "entity_type"
                    for input_item in schema_input:
                        if isinstance(input_item, dict) and input_item.get('name') == 'entity_type':
                            # El value puede estar como string directo o como lista de objetos
                            value = input_item.get('value')
                            if isinstance(value, list) and len(value) > 0:
                                # Si es lista, tomar el primer elemento
                                first_item = value[0]
                                if isinstance(first_item, dict):
                                    entity_type_value = first_item.get('value')
                                else:
                                    entity_type_value = first_item
                            elif isinstance(value, str):
                                entity_type_value = value
                            break
                    
                    if entity_type_value:
                        # Obtener el representative_value de la entidad referenciable
                        ref_entity = db.query(ReferencableEntityModel).filter(
                            ReferencableEntityModel.entity_name == entity_type_value,
                            ReferencableEntityModel.disabled_at.is_(None)
                        ).first()
                        
                        if ref_entity and ref_entity.representative_value:
                            entity_info_map[name] = {
                                'entity_name': entity_type_value,
                                'representative_value': ref_entity.representative_value
                            }
        
        # Enriquecer cada elemento del detail con su display_name y valores representativos
        enriched_detail = []
        for item in detail:
            if isinstance(item, dict):
                item_copy = item.copy()
                item_name = item.get('name')
                item_value = item.get('value')
                
                # Agregar display_name
                if item_name and item_name in display_name_map:
                    item_copy['display_name'] = display_name_map[item_name]
                
                # Agregar valor representativo para entidades
                if item_name and item_name in entity_info_map and item_value:
                    entity_info = entity_info_map[item_name]
                    entity_display = self._get_entity_representative_value(
                        entity_info['entity_name'],
                        item_value,
                        entity_info['representative_value'],
                        db
                    )
                    if entity_display:
                        item_copy['entity_display_value'] = entity_display
                
                enriched_detail.append(item_copy)
            else:
                enriched_detail.append(item)
        
        return enriched_detail
    
    def get_registers_by_form(
        self, 
        form_id: UUID, 
        page: int = 1, 
        per_page: int = 10, 
        sort_by: Optional[str] = None, 
        order: Optional[str] = "asc", 
        search: str = "",
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ) -> PaginatedCoreRegisterResponse:
        """
        Lista todos los registros de un formulario paginados (solo los no deshabilitados)
        
        Args:
            form_id: ID del formulario
            page: Número de página
            per_page: Registros por página
            sort_by: Campo para ordenar
            order: Orden ascendente o descendente
            search: Texto de búsqueda en detail
            start_date: Fecha inicial (filtro por created_at) en formato YYYY-MM-DD
            end_date: Fecha final (filtro por created_at) en formato YYYY-MM-DD
        """
        db = self._get_db()
        from datetime import datetime as dt
        
        # Query base con conversión de geometry a texto directamente
        base_query = db.query(CoreRegisterModel).filter(
            CoreRegisterModel.form_id == form_id,
            CoreRegisterModel.disabled_at.is_(None)
        )
        
        # Aplicar búsqueda si se proporciona (buscar en detail que es JSONB)
        if search:
            # Buscar en el JSON detail convirtiendo a texto
            from sqlalchemy import String
            base_query = base_query.filter(
                func.cast(CoreRegisterModel.detail, String).ilike(f"%{search}%")
            )
        
        # Filtro por rango de fechas
        if start_date:
            try:
                start_dt = dt.strptime(start_date, "%Y-%m-%d")
                base_query = base_query.filter(
                    func.date(CoreRegisterModel.created_at) >= start_dt.date()
                )
            except ValueError:
                raise ValueError(f"Formato de fecha inicio inválido: {start_date}. Use YYYY-MM-DD")
        
        if end_date:
            try:
                end_dt = dt.strptime(end_date, "%Y-%m-%d")
                base_query = base_query.filter(
                    func.date(CoreRegisterModel.created_at) <= end_dt.date()
                )
            except ValueError:
                raise ValueError(f"Formato de fecha fin inválido: {end_date}. Use YYYY-MM-DD")
        
        # Aplicar ordenamiento
        if sort_by:
            sort_column = getattr(CoreRegisterModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    base_query = base_query.order_by(sort_column.desc())
                else:
                    base_query = base_query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            base_query = base_query.order_by(CoreRegisterModel.created_at.desc())
        
        total = base_query.count()
        offset = (page - 1) * per_page
        
        # Query con conversión de geometry a texto para los resultados
        query = db.query(
            CoreRegisterModel,
            func.ST_AsText(CoreRegisterModel.location).label('location_text')
        ).filter(
            CoreRegisterModel.form_id == form_id,
            CoreRegisterModel.disabled_at.is_(None)
        )
        
        # Aplicar búsqueda también en la query de resultados
        if search:
            from sqlalchemy import String
            query = query.filter(
                func.cast(CoreRegisterModel.detail, String).ilike(f"%{search}%")
            )
        
        # Aplicar filtros de fecha también en la query de resultados
        if start_date:
            try:
                start_dt = dt.strptime(start_date, "%Y-%m-%d")
                query = query.filter(
                    func.date(CoreRegisterModel.created_at) >= start_dt.date()
                )
            except ValueError:
                pass  # Ya validado en base_query
        
        if end_date:
            try:
                end_dt = dt.strptime(end_date, "%Y-%m-%d")
                query = query.filter(
                    func.date(CoreRegisterModel.created_at) <= end_dt.date()
                )
            except ValueError:
                pass  # Ya validado en base_query
        
        # Aplicar ordenamiento también en la query de resultados
        if sort_by:
            sort_column = getattr(CoreRegisterModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            query = query.order_by(CoreRegisterModel.created_at.desc())
        
        results = query.offset(offset).limit(per_page).all()
        
        # Construir respuestas con location convertido a texto
        register_responses = []
        for register, location_text in results:
            # Enriquecer el detail con display_names
            enriched_detail = self._enrich_detail_with_display_names(
                register.detail,
                register.schema_form_id,
                db
            )
            
            register_dict = {
                'id': register.id,
                'form_id': register.form_id,
                'schema_form_id': register.schema_form_id,
                'detail': enriched_detail,
                'status': register.status,
                'error': register.error,
                'location': location_text,
                'entity_name': register.entity_name,
                'entity_id': register.entity_id,
                'identity_id': register.identity_id,
                'duration': register.duration,
                'created_at': register.created_at,
                'updated_at': register.updated_at,
                'disabled_at': register.disabled_at
            }
            register_responses.append(CoreRegisterResponse(**register_dict))
        
        return PaginatedCoreRegisterResponse(
            items=register_responses,
            total=total,
            page=page,
            per_page=per_page
        )
    
    def create_form_schema(self, form_id: UUID, schema: Dict[str, Any]) -> FormWithSchemaResponse:
        """
        Crea o actualiza el schema de un formulario.
        
        Crea un nuevo schema_form en la tabla schema_forms y actualiza
        el form.schema_id para que apunte al schema recién creado.
        
        Args:
            form_id: ID del formulario
            schema: Schema completo del formulario (debe contener instructions y instruction_start)
            
        Returns:
            FormWithSchemaResponse con el form actualizado y su schema
        """
        db = self._get_db()
        try:
            # Verificar que el form existe
            form = db.query(FormModel).filter(
                FormModel.id == form_id,
                FormModel.disabled_at.is_(None)
            ).first()
            
            if not form:
                raise ValueError(f"Form con id {form_id} no existe o está deshabilitado")
            
            # Procesar el schema completo para agregar data_input a metadata de cada instrucción
            import copy
            schema_copy = copy.deepcopy(schema) if schema else {}
            processed_schema = process_schema_add_data_input_to_metadata(schema_copy)
            
            # Validar identificador lógico: a lo sumo uno; si hay uno, debe ser único y no nulo en el modelo
            logical_fields = _get_logical_identifier_fields(processed_schema)
            if len(logical_fields) > 1:
                raise ValueError("Solo puede haber un campo marcado como identificador lógico.")
            if len(logical_fields) == 1:
                field_name = logical_fields[0]
                if form.entity_name and form.form_purpose == FormPurpose.entity:
                    model_class = find_model_by_entity_name(form.entity_name)
                    if not model_class:
                        raise ValueError(f"No se encontró modelo para la entidad '{form.entity_name}'.")
                    insp = inspect(model_class)
                    if field_name not in [c.key for c in insp.columns]:
                        raise ValueError(
                            f"El campo '{field_name}' marcado como identificador lógico no existe en la entidad '{form.entity_name}'."
                        )
                    col = insp.columns[field_name]
                    if col.nullable:
                        raise ValueError(
                            f"El campo '{field_name}' debe ser obligatorio (no nulo) para ser identificador lógico."
                        )
                    # Comprobar unicidad: Column(unique=True) o UniqueConstraint en la tabla
                    is_unique = getattr(col, "unique", False) is True
                    if not is_unique and hasattr(insp.table, "constraints"):
                        for c in insp.table.constraints:
                            if type(c).__name__ == "UniqueConstraint" and hasattr(c, "columns"):
                                if col in c.columns:
                                    is_unique = True
                                    break
                    if not is_unique:
                        raise ValueError(
                            f"El campo '{field_name}' debe ser único para ser identificador lógico."
                        )
            
            # Crear el schema_form
            schema_form = SchemaFormModel(
                form_id=form_id,
                schema=processed_schema
            )
            db.add(schema_form)
            db.commit()
            db.refresh(schema_form)
            
            # Actualizar el form para que apunte al schema_form creado
            form.schema_id = schema_form.id
            form.updated_at = datetime.utcnow()
            db.flush()
            db.commit()
            db.refresh(form)
            
            # Construir la respuesta con el form y el schema
            form_dict = {
                "id": form.id,
                "channel_name": form.channel_name,
                "flow_type": form.flow_type,
                "name": form.name,
                "image_path": form.image_path,
                "description": form.description,
                "schema_id": form.schema_id,
                "viewer": form.viewer,
                "gps_tracking": form.gps_tracking,
                "entity_name": form.entity_name,
                "form_purpose": form.form_purpose,
                "created_at": form.created_at,
                "updated_at": form.updated_at,
                "disabled_at": form.disabled_at,
                "schema": schema_form.schema
            }
            
            return FormWithSchemaResponse(**form_dict)
        except Exception as e:
            db.rollback()
            raise e
    
    # Internal method to create schema_form (used only internally, not exposed as API)
    def _create_schema_form_internal(self, form_id: UUID, schema: Dict[str, Any]) -> SchemaFormModel:
        """
        Método interno para crear un schema_form.
        Solo se usa internamente durante la creación automática de formularios.
        """
        db = self._get_db()
        try:
            # Procesar el schema completo para agregar data_input a metadata de cada instrucción
            import copy
            schema_copy = copy.deepcopy(schema) if schema else {}
            processed_schema = process_schema_add_data_input_to_metadata(schema_copy)
            
            # Crear el schema_form
            schema_form = SchemaFormModel(
                form_id=form_id,
                schema=processed_schema
            )
            db.add(schema_form)
            db.commit()
            db.refresh(schema_form)
            
            # Actualizar el form para que apunte al schema_form creado
            form = db.query(FormModel).filter(FormModel.id == form_id).first()
            if form:
                form.schema_id = schema_form.id
                db.flush()
                db.commit()
                db.refresh(form)
            
            return schema_form
        except Exception as e:
            db.rollback()
            raise e
    
    # Form methods
    def get_form_by_id(self, form_id: UUID) -> Optional[FormWithSchemaResponse]:
        """
        Obtiene un formulario por su ID incluyendo su schema.
        
        Busca el form por form_id, obtiene su schema_id, busca el schema_form
        y devuelve el form con el schema como atributo adicional.
        """
        db = self._get_db()
        
        # Buscar el form por id
        form = db.query(FormModel).filter(FormModel.id == form_id).first()
        if not form:
            return None
        
        # Obtener el schema si existe schema_id
        schema = None
        if form.schema_id:
            schema_form = db.query(SchemaFormModel).filter(SchemaFormModel.id == form.schema_id).first()
            if schema_form:
                schema = schema_form.schema
        
        # Construir la respuesta con el form y el schema
        form_dict = {
            "id": form.id,
            "channel_name": form.channel_name,
            "flow_type": form.flow_type,
            "name": form.name,
            "image_path": form.image_path,
            "description": form.description,
            "schema_id": form.schema_id,
            "viewer": form.viewer,
            "gps_tracking": form.gps_tracking,
            "entity_name": form.entity_name,
            "form_purpose": form.form_purpose,
            "created_at": form.created_at,
            "updated_at": form.updated_at,
            "disabled_at": form.disabled_at,
            "schema": schema
        }
        
        return FormWithSchemaResponse(**form_dict)

    def get_form_by_entity_name(self, entity_name: str) -> Optional[FormWithSchemaResponse]:
        """
        Obtiene el primer formulario de tipo entity con el entity_name dado (incluye schema).
        Útil para carga masiva cuando se conoce la entidad pero no el form_id.
        """
        if not entity_name or not entity_name.strip():
            return None
        db = self._get_db()
        entity_name_clean = entity_name.strip().lower()
        form = (
            db.query(FormModel)
            .filter(
                FormModel.disabled_at.is_(None),
                FormModel.form_purpose == FormPurpose.entity,
                func.lower(FormModel.entity_name) == entity_name_clean,
            )
            .first()
        )
        if not form:
            return None
        return self.get_form_by_id(form.id)

    def get_forms_entity(
        self,
        page: int = 1,
        per_page: int = 50,
        entity_name: Optional[str] = None,
    ) -> PaginatedFormResponse:
        """Lista formularios con form_purpose=entity (para carga masiva). Opcional: filtrar por entity_name."""
        db = self._get_db()
        query = db.query(FormModel).filter(
            FormModel.disabled_at.is_(None),
            FormModel.form_purpose == FormPurpose.entity,
        )
        if entity_name and entity_name.strip():
            query = query.filter(func.lower(FormModel.entity_name) == entity_name.strip().lower())
        total = query.count()
        offset = (page - 1) * per_page
        items = query.order_by(FormModel.entity_name.asc(), FormModel.name.asc()).offset(offset).limit(per_page).all()
        return PaginatedFormResponse(
            page=page,
            per_page=per_page,
            total=total,
            items=[FormResponse.model_validate(f) for f in items],
        )

    def get_schema_columns_for_template(self, schema: Optional[Dict[str, Any]], entity_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Devuelve columnas para plantilla Excel desde el schema.
        Para columnas tipo entity (FK): si la entidad referenciada tiene form con form_purpose=entity,
        usa {prefix}_{logical_id} (ej: farmer_dni) en vez de farmer_id.
        Cada form se crea desde una entidad: entity_name permite buscar display_name de cada attr en el modelo.
        """
        columns = get_schema_columns_for_template(schema)
        filtered = []
        for col in columns:
            if col.get("type_value") != "entity":
                filtered.append(col)
                continue
            fk_table = col.get("foreign_key_table")
            if not fk_table and col.get("name", "").endswith("_id"):
                prefix = col["name"][:-3]
                fk_table = f"{prefix}s"
                col["foreign_key_table"] = fk_table
            if not fk_table:
                continue
            related_form = self.get_form_by_entity_name(fk_table)
            if not related_form or not related_form.schema:
                continue
            lid_field = get_logical_identifier_field(related_form.schema)
            if not lid_field:
                continue
            original_name = col["name"]
            prefix = original_name[:-3] if original_name.endswith("_id") else original_name
            new_name = f"{prefix}_{lid_field}"
            col["schema_field_name"] = original_name
            col["name"] = new_name
            col["display_name"] = new_name
            try:
                rel_model = find_model_by_entity_name(fk_table)
                if rel_model and hasattr(rel_model, "__table__") and lid_field in rel_model.__table__.c:
                    sa_col = rel_model.__table__.c[lid_field]
                    if getattr(sa_col, "info", None) and isinstance(sa_col.info, dict) and sa_col.info.get("display_name"):
                        col["display_name"] = sa_col.info["display_name"]
            except Exception:
                pass
            filtered.append(col)
        columns = filtered
        # display_name desde el modelo (column.info["display_name"]) - usar sqlalchemy.inspect
        model_class = None
        if entity_name:
            model_class = find_model_by_entity_name(entity_name)
        if not model_class:
            try:
                from core.models.registry import get_registry
                non_entity_names = {c.get("name") for c in columns if c.get("type_value") != "entity" and c.get("name")}
                if non_entity_names:
                    best_match, best_count = None, 0
                    for _tablename, mod in get_registry().get_all_models().items():
                        if not hasattr(mod, "__table__"):
                            continue
                        table_names = set(mod.__table__.c.keys())
                        match_count = len(non_entity_names & table_names)
                        if match_count > best_count:
                            best_count, best_match = match_count, mod
                    if best_match:
                        model_class = best_match
            except Exception:
                pass
        if model_class and hasattr(model_class, "__table__"):
            table = model_class.__table__
            display_map = {}
            try:
                for key in table.c.keys():
                    try:
                        sa_col = table.c[key]
                    except (KeyError, TypeError):
                        continue
                    if getattr(sa_col, "info", None) and isinstance(sa_col.info, dict) and sa_col.info.get("display_name"):
                        display_map[key] = sa_col.info["display_name"]
            except Exception:
                pass
            for col in columns:
                if col.get("type_value") == "entity":
                    continue
                attr_name = col.get("name")
                if attr_name and attr_name in display_map:
                    col["display_name"] = display_map[attr_name]
        return columns
    
    # Core Register methods
    
    def create_register(self, register_data: CoreRegisterCreate) -> CoreRegisterResponse:
        """
        Crea un nuevo registro en core_registers.
        Si el formulario tiene form_purpose=ENTITY, también guarda los datos en la tabla correspondiente.
        
        Args:
            register_data: Datos del registro a crear
            
        Returns:
            CoreRegisterResponse con el registro creado
        """
        db = self._get_db()
        
        try:
            # Obtener el formulario para buscar entity_name
            form = db.query(FormModel).filter(FormModel.id == register_data.form_id).first()
            if not form:
                raise ValueError(f"Formulario con id {register_data.form_id} no encontrado")
            
            # Obtener entity_name desde el formulario
            entity_name = form.entity_name
            
            # Convertir location de WKT string a Geometry si existe
            from geoalchemy2 import WKTElement
            from sqlalchemy import func as sql_func
            
            location_geom = None
            if register_data.location:
                # Convertir WKT string a Geometry usando WKTElement
                location_geom = WKTElement(register_data.location, srid=4326)
            
            # Asegurar que identity_id existe en identities para no violar FK
            identity_id_to_use = register_data.identity_id
            if identity_id_to_use is not None:
                from modules.auth.src.models.identities import IdentityModel
                identity_query = db.query(IdentityModel).filter(
                    IdentityModel.sub == identity_id_to_use,
                    IdentityModel.disabled_at.is_(None)
                ).first()
                identity_exists = identity_query is not None
                if not identity_exists:
                    print(f"⚠️ Identity {identity_id_to_use} from token not found in identities table; saving register without identity_id")
                    raise ValueError("Identidad no encontrada")
                    # identity_id_to_use = None
            
            # Crear el registro
            # status se establece por defecto como success (el schema no lo incluye)
            # error se establece como None (el schema no lo incluye)
            # entity_id se establece como None inicialmente, se actualizará si se procesa la entidad
            register = CoreRegisterModel(
                form_id=register_data.form_id,
                schema_form_id=register_data.schema_form_id,
                detail=register_data.detail,
                status=RegisterStatus.success,  # Valor por defecto
                error=None,  # Valor por defecto
                location=location_geom,
                entity_name=entity_name,  # Obtenido del formulario
                entity_id=None,  # Se actualizará si se procesa la entidad
                identity_id=identity_query.id,  # Usuario quien registra (None si no existe en identities)
                duration=register_data.duration  # Tiempo que demoró el registro
            )
            db.add(register)
            db.commit()
            db.refresh(register)
            
            # Si el formulario tiene form_purpose=ENTITY, procesar y guardar en la tabla correspondiente
            # (misma lógica que antes ejecutaba el trigger vía API interna)
            if form.form_purpose == FormPurpose.entity:
                try:
                    entity_id = process_register_to_entity(register, db, FormModel, FormPurpose)
                    if entity_id:
                        register.status = RegisterStatus.success
                        register.error = None
                    else:
                        register.status = RegisterStatus.failed
                        register.error = {'message': 'No se pudo procesar el registro a entidad'}
                    db.commit()
                    db.refresh(register)
                except Exception as e:
                    db.rollback()
                    import traceback
                    register = db.query(CoreRegisterModel).filter(CoreRegisterModel.id == register.id).first()
                    if register:
                        register.status = RegisterStatus.failed
                        register.error = {'message': str(e), 'traceback': traceback.format_exc()}
                        db.commit()
                        db.refresh(register)
            
            # Convertir location a texto para la respuesta
            location_text = None
            if register.location:
                result = db.query(sql_func.ST_AsText(register.location)).filter(
                    CoreRegisterModel.id == register.id
                ).first()
                if result:
                    location_text = result[0]
            
            # Construir respuesta
            # Enriquecer el detail con display_names
            enriched_detail = self._enrich_detail_with_display_names(
                register.detail,
                register.schema_form_id,
                db
            )
            
            register_dict = {
                'id': register.id,
                'form_id': register.form_id,
                'schema_form_id': register.schema_form_id,
                'detail': enriched_detail,
                'status': register.status,
                'error': register.error,
                'location': location_text,
                'entity_name': register.entity_name,
                'entity_id': register.entity_id,
                'identity_id': register.identity_id,
                'duration': register.duration,
                'created_at': register.created_at,
                'updated_at': register.updated_at,
                'disabled_at': register.disabled_at
            }
            
            return CoreRegisterResponse(**register_dict)
            
        except Exception as e:
            db.rollback()
            print(f"❌ Error al crear registro: {e}")
            import traceback
            traceback.print_exc()
            raise e
    
    def generate_form_entities(self, config: Dict[str, Any]) -> None:
        """
        Genera formularios automáticamente basados en la configuración del config.yaml.
        
        Esta función es de uso interno del core, no se expone como API.
        Recorre todos los módulos en backend.modules y busca secciones 'forms'.
        Para cada form definido, crea automáticamente un formulario que recolecta
        los atributos de la entidad mencionada.
        
        Args:
            config: Configuración completa del config.yaml
        """
        # Importar funciones necesarias desde resources (específicas del módulo data_collector)
        from .resources.form_auto_creator import (
            find_model_by_entity_name,
            generate_schema_from_model,
            _get_attributes_signature,
            _get_schema_attributes_signature,
            get_model_attributes
        )
        from sqlalchemy.orm import Session
        from sqlalchemy import text
        
        backend_config = config.get("backend", {})
        modules_config = backend_config.get("modules", [])
        
        print(f"📋 Procesando formularios automáticos desde config.yaml...")
        
        # Crear un mapa GLOBAL de todas las entidades referenciables de TODOS los módulos
        # Esto permite que cualquier formulario pueda encontrar el representative_value
        # de cualquier entidad, sin importar en qué módulo esté definida
        global_referencable_entities_map = {}
        for module_config in modules_config:
            referencable_entities_config = module_config.get("referencable_entities", [])
            for ref_entity in referencable_entities_config:
                ref_entity_name = ref_entity.get("entity", "").strip().lower() if ref_entity.get("entity") else None
                if ref_entity_name:
                    # Si ya existe, mantener el primero encontrado (o podrías hacer merge)
                    if ref_entity_name not in global_referencable_entities_map:
                        global_referencable_entities_map[ref_entity_name] = ref_entity
                        print(f"  📋 Entidad referenciable global registrada: '{ref_entity_name}' -> representative_value: {ref_entity.get('representative_value', 'N/A')}")
        
        print(f"  ✅ Total de entidades referenciables encontradas: {len(global_referencable_entities_map)}")
        
        forms_created = 0
        forms_skipped = 0
        
        for module_config in modules_config:
            module_name = module_config.get("name")
            forms_config = module_config.get("forms", [])
            
            if not forms_config:
                continue
            
            print(f"  📦 Módulo {module_name}: {len(forms_config)} formulario(s) definido(s)")
            
            for form_config in forms_config:
                form_name = form_config.get("name")
                form_id = form_config.get("id")  # Leer ID del config.yaml si existe
                form_description = form_config.get("description", "")
                entity_name = form_config.get("entity")
                # IMPORTANTE: entityMap debe estar dentro de cada formulario en 'forms', NO en 'referencable_entities'
                # entityMap permite personalizar campos del formulario (displayName, description, inputs, etc.)
                entity_map_config = form_config.get("entityMap", {})
                # entityMap puede ser un dict con mode e items, o una lista (compatibilidad hacia atrás)
                if isinstance(entity_map_config, dict):
                    entity_map_mode = entity_map_config.get("mode", "merge")  # Por defecto merge
                    entity_map = entity_map_config.get("items", [])
                else:
                    # Compatibilidad hacia atrás: si es una lista, usar merge por defecto
                    entity_map_mode = "merge"
                    entity_map = entity_map_config if isinstance(entity_map_config, list) else []
                form_display_name = form_config.get("display_name")  # Obtener display_name del formulario
                
                if not form_name or not entity_name:
                    print(f"    ⚠️  Formulario sin nombre o entidad, saltando...")
                    forms_skipped += 1
                    continue
                
                # Convertir entity_name a minúsculas
                entity_name = entity_name.lower() if entity_name else None
                
                try:
                    # Buscar el modelo de la entidad
                    model_class = find_model_by_entity_name(entity_name, self.container)
                    if not model_class:
                        print(f"    ⚠️  No se encontró modelo para entidad '{entity_name}', saltando formulario '{form_name}'")
                        forms_skipped += 1
                        continue
                    
                    print(f"    🔍 Entidad '{entity_name}' encontrada, generando schema...")
                    if entity_map:
                        print(f"    📝 Usando entityMap con mode='{entity_map_mode}' y {len(entity_map)} campos personalizados")
                    if form_display_name:
                        print(f"    📝 Usando display_name del formulario: {form_display_name}")
                    
                    # Usar el mapa GLOBAL de entidades referenciables (de todos los módulos)
                    # Esto permite que cualquier ForeignKey encuentre su representative_value
                    # sin importar en qué módulo esté definida la entidad referenciada
                    referencable_entities_map = global_referencable_entities_map
                    
                    # Generar el schema desde el modelo (una instrucción por atributo usando tools)
                    schema = generate_schema_from_model(
                        model_class, 
                        form_name, 
                        self.container, 
                        entity_map=entity_map,
                        entity_map_mode=entity_map_mode,
                        form_display_name=form_display_name,
                        referencable_entities_map=referencable_entities_map
                    )
                    
                    # Verificar si el formulario ya existe
                    db = self._get_db()
                    if isinstance(db, Session):
                        # Verificar si la transacción está en un estado válido
                        try:
                            db.execute(text("SELECT 1"))
                        except Exception as trans_error:
                            print(f"    ⚠️  Transacción abortada, haciendo rollback: {trans_error}")
                            try:
                                db.rollback()
                            except Exception as rollback_error:
                                print(f"    ⚠️  Error al hacer rollback: {rollback_error}")
                        
                        # Obtener los atributos actuales de la entidad
                        current_attributes = get_model_attributes(model_class)
                        current_attr_signature = _get_attributes_signature(current_attributes)
                        
                        # Buscar formulario existente
                        # Prioridad: 1) Por ID si se proporciona, 2) Por nombre+entidad
                        try:
                            if form_id:
                                # Si se proporciona un ID en el config, buscar por ID primero
                                from uuid import UUID as UUID_TYPE
                                try:
                                    form_uuid = UUID_TYPE(form_id) if isinstance(form_id, str) else form_id
                                    existing_form = db.query(FormModel).filter(
                                        FormModel.id == form_uuid,
                                        FormModel.disabled_at.is_(None)
                                    ).first()
                                except (ValueError, AttributeError):
                                    print(f"    ⚠️  ID inválido en config: {form_id}, buscando por nombre/entidad")
                                    existing_form = db.query(FormModel).filter(
                                        FormModel.name == form_name,
                                        FormModel.entity_name == entity_name,
                                        FormModel.form_purpose == FormPurpose.entity,
                                        FormModel.disabled_at.is_(None)
                                    ).first()
                            else:
                                # Si no hay ID, buscar por nombre y entidad (comportamiento anterior)
                                existing_form = db.query(FormModel).filter(
                                    FormModel.name == form_name,
                                    FormModel.entity_name == entity_name,
                                    FormModel.form_purpose == FormPurpose.entity,
                                    FormModel.disabled_at.is_(None)
                                ).first()
                        except Exception as query_error:
                            print(f"    ⚠️  Error en consulta, haciendo rollback: {query_error}")
                            try:
                                db.rollback()
                                existing_form = db.query(FormModel).filter(
                                    FormModel.name == form_name,
                                    FormModel.entity_name == entity_name,
                                    FormModel.form_purpose == FormPurpose.entity,
                                    FormModel.disabled_at.is_(None)
                                ).first()
                            except Exception as retry_error:
                                print(f"    ❌ Error al reintentar consulta: {retry_error}")
                                raise retry_error
                        
                        if existing_form:
                            # Verificar si el schema necesita actualizarse
                            schema_form = db.query(SchemaFormModel).filter(
                                SchemaFormModel.form_id == existing_form.id
                            ).first()
                            
                            if schema_form and schema_form.schema:
                                # Comparar firmas de schema (incluye is_logical_identifier desde entityMap)
                                # Usar el schema generado (con entityMap aplicado) vs el existente en BD
                                current_schema_signature = _get_schema_attributes_signature(schema)
                                existing_schema_signature = _get_schema_attributes_signature(schema_form.schema)
                                
                                if current_schema_signature != existing_schema_signature:
                                    print(f"    🔄 Cambios detectados en entidad '{entity_name}', generando nuevo schema para formulario '{existing_form.name}'...")
                                    
                                    # Generar nuevo schema
                                    new_schema = generate_schema_from_model(
                                        model_class, 
                                        form_name, 
                                        self.container, 
                                        entity_map=entity_map,
                                        entity_map_mode=entity_map_mode,
                                        form_display_name=form_display_name,
                                        referencable_entities_map=referencable_entities_map
                                    )
                                    
                                    # Crear un nuevo schema_form
                                    try:
                                        new_schema_response = self._create_schema_form_internal(
                                            form_id=existing_form.id,
                                            schema=new_schema
                                        )
                                        print(f"    ✅ Nuevo schema creado y asignado al formulario '{existing_form.name}' (Schema ID: {new_schema_response.id})")
                                    except Exception as schema_error:
                                        print(f"    ❌ Error al crear schema para formulario '{existing_form.name}': {schema_error}")
                                        import traceback
                                        traceback.print_exc()
                                        try:
                                            db.rollback()
                                        except:
                                            pass
                                        raise schema_error
                                else:
                                    print(f"    ✓ Formulario '{existing_form.name}' ya existe y la entidad no ha cambiado, omitiendo actualización")
                            else:
                                # Si no hay schema, crear uno nuevo
                                print(f"    🔄 Formulario '{existing_form.name}' existe pero no tiene schema, creando schema...")
                                try:
                                    schema_response = self._create_schema_form_internal(
                                        form_id=existing_form.id,
                                        schema=schema
                                    )
                                    print(f"    ✅ Schema creado para '{existing_form.name}' (ID: {schema_response.id})")
                                except Exception as schema_error:
                                    print(f"    ❌ Error al crear schema para formulario '{existing_form.name}': {schema_error}")
                                    import traceback
                                    traceback.print_exc()
                                    try:
                                        db.rollback()
                                    except:
                                        pass
                                    raise schema_error
                            
                            forms_skipped += 1
                            continue
                    
                    # Crear el formulario usando el servicio
                    from .models.forms import ChannelName, FormType
                    
                    # Preparar datos del formulario, incluyendo ID si se proporcionó en config
                    form_data_dict = {
                        "channel_name": ChannelName.identi_connect,
                        "flow_type": FormType.linear,
                        "name": form_name,
                        "description": form_description,
                        "entity_name": entity_name,
                        "form_purpose": FormPurpose.entity
                    }
                    
                    # Agregar ID si se proporcionó en el config.yaml
                    if form_id:
                        from uuid import UUID as UUID_TYPE
                        try:
                            form_uuid = UUID_TYPE(form_id) if isinstance(form_id, str) else form_id
                            form_data_dict["id"] = form_uuid
                            print(f"    📝 Usando ID personalizado del config: {form_uuid}")
                        except (ValueError, AttributeError) as id_error:
                            print(f"    ⚠️  ID inválido en config: {form_id}, se generará automáticamente. Error: {id_error}")
                    
                    form_data = FormCreate(**form_data_dict)
                    
                    # Crear el formulario
                    try:
                        form_response = self.create_form(form_data)
                        print(f"    ✅ Formulario '{form_name}' creado (ID: {form_response.id})")
                    except Exception as form_error:
                        print(f"    ❌ Error al crear formulario '{form_name}': {form_error}")
                        import traceback
                        traceback.print_exc()
                        try:
                            db.rollback()
                        except:
                            pass
                        raise form_error
                    
                    # Crear el schema_form
                    try:
                        schema_response = self._create_schema_form_internal(
                            form_id=form_response.id,
                            schema=schema
                        )
                        print(f"    ✅ Schema creado para '{form_name}' (ID: {schema_response.id})")
                    except Exception as schema_error:
                        print(f"    ❌ Error al crear schema para formulario '{form_name}': {schema_error}")
                        import traceback
                        traceback.print_exc()
                        try:
                            db.rollback()
                        except:
                            pass
                        raise schema_error
                    
                    forms_created += 1
                    
                except Exception as e:
                    print(f"    ❌ Error al crear formulario '{form_name}': {e}")
                    import traceback
                    traceback.print_exc()
                    
                    # Hacer rollback de la transacción si hay un error
                    db = self._get_db()
                    if isinstance(db, Session):
                        try:
                            db.rollback()
                        except Exception as rollback_error:
                            print(f"    ⚠️  Error al hacer rollback: {rollback_error}")
                    
                    forms_skipped += 1
        
        print(f"✅ Formularios automáticos procesados: {forms_created} creados, {forms_skipped} omitidos/errores")
    
    # Referencable Entities methods
    def get_referencable_entities(
        self, 
        page: int = 1, 
        per_page: int = 10, 
        sort_by: Optional[str] = None, 
        order: Optional[str] = "asc", 
        search: str = "",
        module_name: Optional[str] = None
    ):
        """Lista todas las entidades referenciables paginadas. Devuelve id (entity_name), display_name y description"""
        from .schemas import PaginatedEntityListItemResponse, EntityListItemResponse
        
        db = self._get_db()
        
        try:
            from .models.referencable_entities import ReferencableEntityModel
            
            query = db.query(ReferencableEntityModel).filter(ReferencableEntityModel.disabled_at.is_(None))
            
            # Filtrar por módulo si se proporciona
            if module_name:
                query = query.filter(ReferencableEntityModel.module_name == module_name)
            
            # Aplicar búsqueda si se proporciona
            if search:
                query = query.filter(
                    ReferencableEntityModel.entity_name.ilike(f"%{search}%") |
                    (ReferencableEntityModel.display_name.isnot(None) & ReferencableEntityModel.display_name.ilike(f"%{search}%")) |
                    (ReferencableEntityModel.description.isnot(None) & ReferencableEntityModel.description.ilike(f"%{search}%"))
                )
            
            # Aplicar ordenamiento
            if sort_by:
                sort_column = getattr(ReferencableEntityModel, sort_by, None)
                if sort_column is not None:
                    if order and order.lower() == "desc":
                        query = query.order_by(sort_column.desc())
                    else:
                        query = query.order_by(sort_column.asc())
                else:
                    query = query.order_by(ReferencableEntityModel.entity_name.asc())
            else:
                query = query.order_by(ReferencableEntityModel.module_name.asc(), ReferencableEntityModel.entity_name.asc())
            
            # Contar total
            total = query.count()
            
            # Aplicar paginación
            offset = (page - 1) * per_page
            items = query.offset(offset).limit(per_page).all()
            
            # Convertir a schema simplificado (id es entity_name, display_name, description)
            simplified_items = [
                EntityListItemResponse(
                    id=item.entity_name,  # entity_name se devuelve como id
                    display_name=item.display_name,
                    description=item.description
                )
                for item in items
            ]
            
            return PaginatedEntityListItemResponse(
                page=page,
                per_page=per_page,
                total=total,
                items=simplified_items
            )
        except Exception as e:
            print(f"⚠️  Error al obtener entidades referenciables: {e}")
            import traceback
            traceback.print_exc()
            return PaginatedEntityListItemResponse(
                page=page,
                per_page=per_page,
                total=0,
                items=[]
            )
    
    def get_entity_data(
        self,
        entity_name: str,
        page: int = 1,
        per_page: int = 10,
        sort_by: Optional[str] = None,
        order: Optional[str] = "asc",
        search: str = "",
        representative_value: Optional[str] = None,
        filter: Optional[str] = None
    ) -> PaginatedEntityDataResponse:
        """
        Obtiene datos paginados de una entidad específica.
        Solo funciona si la entidad está registrada en referencable_entities.
        
        Args:
            entity_name: Nombre de la entidad
            page: Número de página
            per_page: Elementos por página
            sort_by: Campo por el cual ordenar
            order: Orden ('asc' o 'desc')
            search: Texto de búsqueda
            representative_value: Template opcional para generar el campo 'name'. 
                                 Si se proporciona, tiene prioridad sobre el representative_value de la BD.
            filter: Filtro(s) opcional en formato 'field=value' o 'entity.field=other_entity.field'. 
                   Múltiples filtros separados por coma (ej: 'country_id=PE,status=active' o 'farmers.value=countries.value,farmers.value2=departments.value3').
                   IMPORTANTE: Al menos una entidad en cada filtro debe ser la entidad actual que se está consultando.
        """
        db = self._get_db()
        
        try:
            from .models.referencable_entities import ReferencableEntityModel
            
            # Verificar que la entidad esté registrada como referenciable
            referencable = db.query(ReferencableEntityModel).filter(
                ReferencableEntityModel.entity_name == entity_name.lower(),
                ReferencableEntityModel.disabled_at.is_(None)
            ).first()
            
            if not referencable:
                raise ValueError(f"La entidad '{entity_name}' no está registrada como referenciable")
            
            # Buscar el modelo de la entidad
            from .resources.form_auto_creator import find_model_by_entity_name
            model_class = find_model_by_entity_name(entity_name.lower(), self.container)
            
            if not model_class:
                raise ValueError(f"No se encontró el modelo para la entidad '{entity_name}'")
            
            # Construir query
            # Convertir campos de geometría a WKT en la query para evitar problemas de serialización
            from sqlalchemy import func
            from geoalchemy2 import Geometry
            
            # Identificar columnas de geometría
            has_geometry = any(isinstance(col.type, Geometry) for col in model_class.__table__.columns)
            
            if has_geometry:
                # Si hay columnas de geometría, construir query con ST_AsText para convertirlas a WKT
                select_columns = []
                for column in model_class.__table__.columns:
                    if isinstance(column.type, Geometry):
                        # Convertir geometría a WKT string
                        select_columns.append(func.ST_AsText(column).label(column.name))
                    else:
                        select_columns.append(column)
                query = db.query(*select_columns).filter(model_class.disabled_at.is_(None))
            else:
                query = db.query(model_class).filter(model_class.disabled_at.is_(None))
            
            # Aplicar búsqueda si se proporciona (buscar en todos los campos de texto)
            if search:
                search_filters = []
                for column in model_class.__table__.columns:
                    # Verificar si es un tipo de texto (String, Text, VARCHAR, etc.)
                    column_type_str = str(column.type).lower()
                    is_text_type = (
                        'varchar' in column_type_str or 
                        'string' in column_type_str or 
                        'text' in column_type_str or
                        'char' in column_type_str
                    )
                    
                    # También verificar python_type si está disponible (manejar excepciones)
                    if not is_text_type:
                        try:
                            if hasattr(column.type, 'python_type'):
                                python_type = column.type.python_type
                                is_text_type = python_type == str
                        except (NotImplementedError, AttributeError, TypeError):
                            # Algunos tipos (como Geometry) no implementan python_type
                            # Continuar sin usar python_type
                            pass
                    
                    if is_text_type:
                        # Usar getattr para obtener el atributo del modelo, no la columna directamente
                        model_column = getattr(model_class, column.name, None)
                        if model_column is not None:
                            search_filters.append(model_column.ilike(f"%{search}%"))
                
                if search_filters:
                    from sqlalchemy import or_
                    query = query.filter(or_(*search_filters))
            
            # Aplicar filtro si se proporciona
            if filter:
                from .resources.query_filter import apply_filter
                filter_applied = apply_filter(query, model_class, filter, db, entity_name, self.container)
                if filter_applied is not None:
                    query = filter_applied
            
            # Aplicar ordenamiento
            if sort_by:
                sort_column = getattr(model_class, sort_by, None)
                if sort_column is not None:
                    if order and order.lower() == "desc":
                        query = query.order_by(sort_column.desc())
                    else:
                        query = query.order_by(sort_column.asc())
                else:
                    # Ordenar por id por defecto si el campo no existe
                    if hasattr(model_class, 'id'):
                        query = query.order_by(model_class.id.desc() if order and order.lower() == "desc" else model_class.id.asc())
            else:
                # Ordenar por id por defecto
                if hasattr(model_class, 'id'):
                    query = query.order_by(model_class.id.desc())
            
            # Contar total
            total = query.count()
            
            # Aplicar paginación
            offset = (page - 1) * per_page
            items = query.offset(offset).limit(per_page).all()
            
            # Obtener representative_value: prioridad al parámetro, luego al de la BD
            representative_value_template = representative_value or (referencable.representative_value if referencable else None)
            
            # Convertir items a diccionarios y procesar representative_value
            items_dict = []
            import re
            
            for item in items:
                item_dict = {}
                item_id = None
                
                # Convertir el item a diccionario
                # Si usamos select_columns con ST_AsText, los valores de geometría ya vienen como strings WKT
                for column in model_class.__table__.columns:
                    value = getattr(item, column.name)
                    # Guardar el ID para usarlo después
                    if column.primary_key:
                        item_id = str(value) if isinstance(value, UUID) else value
                    # Convertir UUIDs a strings
                    if isinstance(value, UUID):
                        item_dict[column.name] = str(value)
                    # Convertir datetime a strings ISO
                    elif isinstance(value, datetime):
                        item_dict[column.name] = value.isoformat()
                    # Si el valor es None, mantenerlo como None
                    elif value is None:
                        item_dict[column.name] = None
                    # Si es un WKBElement/WKTElement (puede pasar si no usamos ST_AsText), convertirlo
                    elif hasattr(value, '__class__') and value.__class__.__name__ in ('WKBElement', 'WKTElement'):
                        try:
                            # Intentar acceder al atributo 'desc' que contiene el WKT
                            if hasattr(value, 'desc'):
                                item_dict[column.name] = value.desc
                            else:
                                # Si no tiene 'desc', convertir a string (puede ser que ya sea WKT)
                                item_dict[column.name] = str(value) if value else None
                        except Exception:
                            item_dict[column.name] = None
                    else:
                        item_dict[column.name] = value
                
                # Si no se encontró ID, intentar obtenerlo del campo 'id'
                if item_id is None:
                    item_id = item_dict.get('id')
                    if isinstance(item_id, UUID):
                        item_id = str(item_id)
                
                # Procesar representative_value para generar el campo 'name'
                name_value = ""
                if representative_value_template:
                    # Reemplazar {{columna}} con los valores reales
                    name_value = representative_value_template
                    # Buscar todos los placeholders {{columna}}
                    placeholders = re.findall(r'\{\{(\w+)\}\}', representative_value_template)
                    for placeholder in placeholders:
                        # Obtener el valor de la columna (puede estar en item_dict o en el objeto item)
                        column_value = item_dict.get(placeholder)
                        if column_value is None:
                            # Intentar obtener directamente del objeto si no está en el dict
                            try:
                                column_value = getattr(item, placeholder, None)
                                if isinstance(column_value, UUID):
                                    column_value = str(column_value)
                                elif isinstance(column_value, datetime):
                                    column_value = column_value.isoformat()
                            except:
                                column_value = ""
                        # Reemplazar el placeholder con el valor (convertir a string)
                        name_value = name_value.replace(f"{{{{{placeholder}}}}}", str(column_value) if column_value is not None else "")
                else:
                    # Si no hay representative_value, usar el ID como name
                    name_value = str(item_id) if item_id else ""
                
                items_dict.append(EntityDataItemResponse(
                    id=item_id,
                    name=name_value
                ))
            
            return PaginatedEntityDataResponse(
                page=page,
                per_page=per_page,
                total=total,
                items=items_dict
            )
        except Exception as e:
            print(f"⚠️  Error al obtener datos de la entidad '{entity_name}': {e}")
            import traceback
            traceback.print_exc()
            raise ValueError(f"Error al obtener datos de la entidad '{entity_name}': {str(e)}")
    
    def export_core_registers_to_excel(
        self,
        form_id: UUID,
        sort_by: Optional[str] = None,
        order: Optional[str] = "asc",
        search: str = "",
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        max_file_size_mb: float = 50.0
    ) -> tuple[bytes, str]:
        """
        Exporta core_registers a Excel con los filtros aplicados.
        
        Manejo especial para campos tipo 'entity':
        - Si value es un objeto {id, display_name}, muestra el display_name
        - Si value es un array [{id, display_name}], muestra los display_names separados por comas
        
        Args:
            form_id: ID del formulario
            sort_by: Campo para ordenar
            order: Orden ascendente o descendente
            search: Texto de búsqueda en detail
            start_date: Fecha inicial (filtro por created_at) en formato YYYY-MM-DD
            end_date: Fecha final (filtro por created_at) en formato YYYY-MM-DD
            max_file_size_mb: Tamaño máximo del archivo en MB (default 50.0)
        
        Returns:
            tuple[bytes, str]: Tupla con (contenido del archivo Excel, nombre sugerido del archivo)
        
        Raises:
            ValueError: Si el tamaño del archivo excede max_file_size_mb
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        from datetime import datetime as dt
        from sqlalchemy import String
        
        db = self._get_db()
        
        # Obtener auth_service para consultar usernames
        auth_service = self.container.get("auth")
        
        # 1. Construir query con filtros (similar a get_registers_by_form)
        query = db.query(
            CoreRegisterModel,
            func.ST_AsText(CoreRegisterModel.location).label('location_text')
        ).filter(
            CoreRegisterModel.form_id == form_id,
            CoreRegisterModel.disabled_at.is_(None)
        )
        
        # Aplicar búsqueda en detail
        if search:
            query = query.filter(
                func.cast(CoreRegisterModel.detail, String).ilike(f"%{search}%")
            )
        
        # Filtro por rango de fechas
        if start_date:
            try:
                start_dt = dt.strptime(start_date, "%Y-%m-%d")
                query = query.filter(
                    func.date(CoreRegisterModel.created_at) >= start_dt.date()
                )
            except ValueError:
                raise ValueError(f"Formato de fecha inicio inválido: {start_date}. Use YYYY-MM-DD")
        
        if end_date:
            try:
                end_dt = dt.strptime(end_date, "%Y-%m-%d")
                query = query.filter(
                    func.date(CoreRegisterModel.created_at) <= end_dt.date()
                )
            except ValueError:
                raise ValueError(f"Formato de fecha fin inválido: {end_date}. Use YYYY-MM-DD")
        
        # Aplicar ordenamiento
        if sort_by:
            sort_column = getattr(CoreRegisterModel, sort_by, None)
            if sort_column:
                query = query.order_by(
                    sort_column.desc() if order == "desc" else sort_column.asc()
                )
        else:
            query = query.order_by(CoreRegisterModel.created_at.desc())
        
        # 2. Obtener todos los registros
        total_records = query.count()
        results = query.all()
        
        # 3. Obtener información del formulario
        form = db.query(FormModel).filter(FormModel.id == form_id).first()
        form_name = form.name if form else "Formulario"
        
        # 4. Analizar estructura del detail para determinar columnas dinámicas
        all_field_names = set()
        field_display_names = {}  # Mapeo name -> display_name
        
        for register, _ in results:
            if register.detail:
                for item in register.detail:
                    if isinstance(item, dict) and 'name' in item:
                        field_name = item['name']
                        all_field_names.add(field_name)
                        # Usar display_name si existe
                        if 'display_name' in item and field_name not in field_display_names:
                            field_display_names[field_name] = item['display_name']
        
        # Ordenar campos alfabéticamente
        detail_fields = sorted(list(all_field_names))
        
        # 5. Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"[:31]  # Max 31 caracteres
        
        # 6. Definir headers (columnas fijas + columnas dinámicas del detail)
        fixed_headers = [
            "ID Registro",
            "Fecha Creación",
            "Usuario Registro",
            "Estado",
            "Entidad",
            "ID Entidad",
            "Duración (seg)",
            "Ubicación"
        ]
        
        # Usar display_name si existe, sino usar el nombre del campo
        detail_headers = [field_display_names.get(field, field) for field in detail_fields]
        headers = fixed_headers + detail_headers
        
        # Aplicar estilo a headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # 7. Llenar datos
        # Cache para usernames (evitar consultas repetidas)
        username_cache = {}
        
        for row_num, (register, location_text) in enumerate(results, 2):
            # Obtener username desde identidad
            username = ""
            if register.identity_id:
                if register.identity_id not in username_cache:
                    try:
                        identity = auth_service.get_identity(register.identity_id)
                        username_cache[register.identity_id] = identity.sub if identity else ""
                    except:
                        username_cache[register.identity_id] = ""
                username = username_cache[register.identity_id]
            
            # Crear diccionario de campos del detail para acceso rápido
            detail_dict = {}
            if register.detail:
                for item in register.detail:
                    if isinstance(item, dict) and 'name' in item:
                        value_data = item.get('value')
                        
                        # Detectar si value es una entidad por su estructura
                        # Caso 1: value es un objeto único {id, display_name}
                        if isinstance(value_data, dict) and 'id' in value_data and 'display_name' in value_data:
                            value = value_data.get('display_name', '')
                        
                        # Caso 2: value es un array [{id, display_name}, ...]
                        elif isinstance(value_data, list) and len(value_data) > 0:
                            # Verificar si es un array de entidades
                            if all(isinstance(obj, dict) and 'id' in obj and 'display_name' in obj for obj in value_data):
                                display_names = [obj.get('display_name', '') for obj in value_data]
                                value = ', '.join(display_names) if display_names else ''
                            else:
                                # Array de valores simples
                                value = ', '.join(str(v) for v in value_data)
                        
                        # Caso 3: value es un valor simple (string, number, etc.)
                        else:
                            # verificar si es un media
                            type_detail = item.get('type_value')
                            if type_detail and type_detail == 'media':
                                # obtengo el media usando la funcionalidad get_media_by_id
                                storage_service = self.container.get("storage_s3")
                                media_url = storage_service.get_url_by_media_id(value_data, 604800)
                                if media_url:
                                    value = media_url
                                else:
                                    value = ''
                            else:
                                value = value_data if value_data is not None else ''
                        
                        detail_dict[item['name']] = value
            
            # Escribir columnas fijas
            col = 1
            ws.cell(row=row_num, column=col, value=str(register.id)); col += 1
            ws.cell(row=row_num, column=col, value=register.created_at); col += 1
            ws.cell(row=row_num, column=col, value=username); col += 1
            ws.cell(row=row_num, column=col, value=register.status.value if register.status else ""); col += 1
            ws.cell(row=row_num, column=col, value=register.entity_name or ""); col += 1
            ws.cell(row=row_num, column=col, value=str(register.entity_id) if register.entity_id else ""); col += 1
            ws.cell(row=row_num, column=col, value=float(register.duration) if register.duration else None); col += 1
            ws.cell(row=row_num, column=col, value=location_text or ""); col += 1
            
            # Escribir columnas dinámicas del detail
            for field_name in detail_fields:
                value = detail_dict.get(field_name, "")
                # Convertir listas a texto (solo para tipos que no sean entity, que ya están procesados)
                if isinstance(value, list):
                    value = ", ".join([str(v) for v in value])
                # Convertir dicts a string legible
                elif isinstance(value, dict):
                    value = str(value)
                ws.cell(row=row_num, column=col, value=str(value) if value else "")
                col += 1
        
        # 8. Ajustar ancho de columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Límite entre 10 y 50 caracteres
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 9. Agregar hoja de información/metadatos
        ws_info = wb.create_sheet("Información")
        ws_info.append(["Informe de Exportación"])
        ws_info.append(["Formulario:", form_name])
        ws_info.append(["Total de registros:", total_records])
        ws_info.append(["Fecha de exportación:", dt.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_info.append(["Filtros aplicados:"])
        if search:
            ws_info.append(["  - Búsqueda:", search])
        if start_date:
            ws_info.append(["  - Fecha desde:", start_date])
        if end_date:
            ws_info.append(["  - Fecha hasta:", end_date])
        
        # Estilo para la hoja de información
        for row in ws_info.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=14)
        
        # 10. Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        wb.close()  # Cerrar el workbook explícitamente
        buffer.seek(0)
        excel_bytes = buffer.getvalue()
        buffer.close()  # Cerrar el buffer
        
        # 11. VALIDACIÓN: Verificar tamaño del archivo
        file_size_mb = len(excel_bytes) / (1024 * 1024)  # Convertir bytes a MB
        if file_size_mb > max_file_size_mb:
            raise ValueError(
                f"El tamaño del archivo generado ({file_size_mb:.2f} MB) excede el límite máximo "
                f"permitido ({max_file_size_mb:.2f} MB). "
                f"Total de registros: {total_records:,}. "
                f"Por favor, aplique filtros adicionales (como rango de fechas) para reducir el volumen de datos."
            )
        
        # 12. Generar nombre de archivo sugerido
        form_name_clean = form_name.replace(" ", "_").lower()
        timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{form_name_clean}_{timestamp}.xlsx"
        
        return excel_bytes, filename
    
    def get_registers_by_entity_mention(
        self, 
        entity_id: UUID, 
        page: int = 1, 
        per_page: int = 10, 
        sort_by: Optional[str] = None, 
        order: Optional[str] = "asc", 
        search: str = ""
    ) -> PaginatedCoreRegisterResponse:
        """
        Lista todos los registros donde una entidad específica ha sido mencionada.
        Busca en core_registers.detail donde value contenga el entity_id.
        Detecta entidades por su estructura (presencia de campos 'id' y 'display_name').
        
        Soporta dos estructuras de value:
        - Objeto único: {"id": "...", "display_name": "..."}
        - Array de objetos: [{"id": "...", "display_name": "..."}, ...]
        
        Args:
            entity_id: UUID de la entidad a buscar
            page: Número de página
            per_page: Elementos por página
            sort_by: Campo por el cual ordenar
            order: Orden 'asc' o 'desc'
            search: Texto de búsqueda adicional en detail
            
        Returns:
            PaginatedCoreRegisterResponse con los registros encontrados
        """
        from sqlalchemy import text, and_, String
        db = self._get_db()
        
        # Convertir entity_id a string para búsqueda en JSON
        entity_id_str = str(entity_id)
        
        # Query base para contar
        base_query = db.query(CoreRegisterModel).filter(
            CoreRegisterModel.disabled_at.is_(None)
        )
        
        # Subquery para buscar en el array JSONB detail
        # Busca elementos donde value contenga el entity_id (detecta entidades por estructura)
        # Considera dos casos:
        # 1. value es un objeto: {id, display_name}
        # 2. value es un array: [{id, display_name}, ...]
        # Nota: detail es ARRAY(JSONB), por eso usamos unnest() primero
        entity_filter = text("""
            EXISTS (
                SELECT 1 
                FROM unnest(core_registers.detail) AS item
                WHERE (
                    -- Caso 1: value es un objeto único con id
                    (item->'value' ? 'id' AND item->'value'->>'id' = :entity_id_str)
                    OR
                    -- Caso 2: value es un array de objetos
                    (jsonb_typeof(item->'value') = 'array' AND EXISTS (
                        SELECT 1
                        FROM jsonb_array_elements(item->'value') AS val_item
                        WHERE val_item ? 'id' AND val_item->>'id' = :entity_id_str
                    ))
                )
            )
        """)
        
        base_query = base_query.filter(entity_filter.bindparams(entity_id_str=entity_id_str))
        
        # Aplicar búsqueda adicional si se proporciona
        if search:
            base_query = base_query.filter(
                func.cast(CoreRegisterModel.detail, String).ilike(f"%{search}%")
            )
        
        # Aplicar ordenamiento para el conteo
        if sort_by:
            sort_column = getattr(CoreRegisterModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    base_query = base_query.order_by(sort_column.desc())
                else:
                    base_query = base_query.order_by(sort_column.asc())
        else:
            base_query = base_query.order_by(CoreRegisterModel.created_at.desc())
        
        # Contar total
        total = base_query.count()
        offset = (page - 1) * per_page
        
        # Query para obtener resultados con location convertido a texto
        query = db.query(
            CoreRegisterModel,
            func.ST_AsText(CoreRegisterModel.location).label('location_text')
        ).filter(
            CoreRegisterModel.disabled_at.is_(None)
        )
        
        # Aplicar el mismo filtro de entidad
        query = query.filter(entity_filter.bindparams(entity_id_str=entity_id_str))
        
        # Aplicar búsqueda adicional también en la query de resultados
        if search:
            query = query.filter(
                func.cast(CoreRegisterModel.detail, String).ilike(f"%{search}%")
            )
        
        # Aplicar ordenamiento también en la query de resultados
        if sort_by:
            sort_column = getattr(CoreRegisterModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            query = query.order_by(CoreRegisterModel.created_at.desc())
        
        # Obtener resultados paginados
        results = query.offset(offset).limit(per_page).all()
        
        # Construir respuestas con location convertido a texto
        register_responses = []
        for register, location_text in results:
            # Enriquecer el detail con display_names
            enriched_detail = self._enrich_detail_with_display_names(
                register.detail,
                register.schema_form_id,
                db
            )
            
            # Obtener información del formulario
            form = db.query(FormModel).filter(FormModel.id == register.form_id).first()
            form_info = None
            if form:
                form_info = {
                    'id': form.id,
                    'name': form.name
                }
            
            register_dict = {
                'id': register.id,
                'form_id': register.form_id,
                'form': form_info,
                'schema_form_id': register.schema_form_id,
                'detail': enriched_detail,
                'status': register.status,
                'error': register.error,
                'location': location_text,
                'entity_name': register.entity_name,
                'entity_id': register.entity_id,
                'identity_id': register.identity_id,
                'duration': register.duration,
                'created_at': register.created_at,
                'updated_at': register.updated_at,
                'disabled_at': register.disabled_at
            }
            
            register_responses.append(CoreRegisterResponse(**register_dict))
        
        return PaginatedCoreRegisterResponse(
            page=page,
            per_page=per_page,
            total=total,
            items=register_responses
        )
    
    def validate_unique_field(
        self, 
        entity_name: str, 
        entity_field: dict, 
        entity_exclude_id: Optional[UUID] = None
    ) -> UniqueFieldValidationResponse:
        """
        Valida si un valor es único para un campo específico en una entidad.
        
        Si el campo existe como columna en el modelo, busca directamente en la columna.
        Si NO existe como columna, busca en el campo JSONB 'entity_value'.
        
        Args:
            entity_name: Nombre de la entidad (ej: "farmers", "farms")
            entity_field: Diccionario con el campo y valor a validar (ej: {"dni": "12345678"})
            entity_exclude_id: ID de la entidad a excluir en la búsqueda (útil para updates)
            
        Returns:
            UniqueFieldValidationResponse con resultado de validación (exist: bool)
        """
        db = self._get_db()
        
        try:
            # Validar que entity_field tenga exactamente un campo
            if not entity_field or len(entity_field) != 1:
                raise ValueError("entity_field debe contener exactamente un campo")
            
            # Extraer el nombre del campo y su valor
            field_name = list(entity_field.keys())[0]
            field_value = entity_field[field_name]
            
            # Buscar el modelo por entity_name
            from .resources.register_processor import find_model_by_entity_name
            model_class = find_model_by_entity_name(entity_name)
            
            if not model_class:
                raise ValueError(f"No se encontró modelo para entidad '{entity_name}'")
            
            # Verificar si el campo existe en el modelo como columna
            if hasattr(model_class, field_name):
                # Caso 1: El campo existe como columna en el modelo
                query = db.query(model_class).filter(
                    getattr(model_class, field_name) == field_value,
                    model_class.disabled_at.is_(None)
                )
            else:
                # Caso 2: El campo NO existe como columna, buscar en entity_value (JSONB)
                # Verificar que el modelo tenga el campo entity_value
                if not hasattr(model_class, 'entity_value'):
                    raise ValueError(f"El campo '{field_name}' no existe en la entidad '{entity_name}' y no hay campo 'entity_value' para atributos dinámicos")
                
                # Buscar en el campo JSONB entity_value
                # Convertir el valor a string para comparación JSON
                if isinstance(field_value, str):
                    value_str = field_value
                else:
                    import json
                    value_str = json.dumps(field_value)
                
                # Query usando el operador ->> para acceder a la clave del JSONB
                from sqlalchemy import text
                
                query = db.query(model_class).filter(
                    text(f"entity_value->>:field_name = :field_value").bindparams(
                        field_name=field_name,
                        field_value=value_str
                    ),
                    model_class.disabled_at.is_(None)
                )
            
            # Excluir la entidad especificada si existe (para updates)
            if entity_exclude_id:
                query = query.filter(model_class.id != entity_exclude_id)
            
            # Ejecutar la query - verificar si existe algún registro
            existing_entity = query.first()
            
            # Retornar si existe o no
            return UniqueFieldValidationResponse(
                exist=existing_entity is not None
            )
                
        except Exception as e:
            print(f"Error al validar campo único: {e}")
            import traceback
            traceback.print_exc()
            raise

    def validate_unique_field_complementary(
        self, 
        entity_field: dict, 
        form_id: UUID
    ) -> UniqueFieldComplementaryValidationResponse:
        """
        Valida si un valor es único en un registro complementario.
        
        Busca en el campo JSONB 'detail' de core_registers.

        Args:
            entity_field: Diccionario con el campo y valor a validar (ej: {"dni": "12345678"})
            form_id: ID del formulario en el que se encuentra el registro complementario
            
        Returns:
            UniqueFieldComplementaryValidationResponse con resultado de validación (exist: bool)
        
        """
        db = self._get_db()
        
        try:
            # Validar que entity_field tenga exactamente un campo
            if not entity_field or len(entity_field) != 1:
                raise ValueError("entity_field debe contener exactamente un campo")
            
            # Extraer el nombre del campo y su valor
            field_name = list(entity_field.keys())[0]
            field_value = entity_field[field_name]
            
            # Caso 1: Buscar en detail (JSONB)
            if not hasattr(CoreRegisterModel, 'detail'):
                raise ValueError(f"El campo '{field_name}' no existe en el registro y no hay campo 'detail' para atributos dinámicos")
                
            # Buscar en el campo JSONB detail
            # Convertir el valor a string para comparación JSON
            if isinstance(field_value, str):
                value_str = field_value
            else:
                import json
                value_str = json.dumps(field_value)
                
            # Query usando el operador ->> para acceder a la clave del JSONB
            from sqlalchemy import text
                
            query = db.query(CoreRegisterModel).filter(
                CoreRegisterModel.form_id == form_id,
                CoreRegisterModel.disabled_at.is_(None),
                text("""
                    EXISTS (
                        SELECT 1
                        FROM unnest(detail) AS d
                        WHERE d->>'name' = :field_name
                        AND d->>'value' = :field_value
                    )
                """).bindparams(
                    field_name=field_name,
                    field_value=value_str
                )
            )
            
            # Ejecutar la query - verificar si existe algún registro
            existing_entity = query.first()
            
            # Retornar si existe o no
            return UniqueFieldComplementaryValidationResponse(
                exist=existing_entity is not None
            )
                
        except Exception as e:
            print(f"Error al validar campo único: {e}")
            import traceback
            traceback.print_exc()
            raise

