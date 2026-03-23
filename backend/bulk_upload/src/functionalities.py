from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from uuid import UUID
from io import BytesIO
from sqlalchemy.orm import Session

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models.bulk_upload_job import BulkUploadJobModel, BulkUploadStatus
from .models.bulk_upload_job_row import BulkUploadJobRowModel
from .schemas import (
    BulkUploadJobResponse,
    BulkUploadJobDetailResponse,
    PaginatedBulkUploadJobsResponse,
    BulkUploadErrorItem,
    PaginatedBulkUploadErrorsResponse,
    JobHeadersResponse,
    PaginatedJobRowsResponse,
)


class Funcionalities:
    def __init__(self, container, database_key: str = "core_db", **kwargs):
        self.container = container
        self.database_key = database_key

    def _get_db(self) -> Session:
        return self.container.get(self.database_key, "databases")

    def get_jobs(
        self,
        page: int = 1,
        per_page: int = 10,
        entity_name: Optional[str] = None,
        status: Optional[str] = None,
    ) -> PaginatedBulkUploadJobsResponse:
        db = self._get_db()
        query = db.query(BulkUploadJobModel)
        if entity_name:
            query = query.filter(BulkUploadJobModel.entity_name == entity_name)
        if status:
            try:
                status_enum = BulkUploadStatus(status)
                query = query.filter(BulkUploadJobModel.status == status_enum)
            except ValueError:
                pass
        total = query.count()
        offset = (page - 1) * per_page
        items = query.order_by(BulkUploadJobModel.created_at.desc()).offset(offset).limit(per_page).all()
        form_names = {}
        if items:
            try:
                from modules.data_collector.src.models.forms import FormModel
            except ImportError:
                from backend.modules.data_collector.src.models.forms import FormModel
            form_ids = list({j.form_id for j in items})
            forms = db.query(FormModel).filter(FormModel.id.in_(form_ids)).all()
            form_names = {f.id: (f.name or "") for f in forms}
        return PaginatedBulkUploadJobsResponse(
            page=page,
            per_page=per_page,
            total=total,
            items=[
                BulkUploadJobResponse(
                    id=j.id,
                    form_id=j.form_id,
                    form_name=form_names.get(j.form_id),
                    entity_name=j.entity_name,
                    file_name=j.file_name,
                    total_rows=j.total_rows,
                    success_count=j.success_count,
                    error_count=j.error_count,
                    status=j.status.value,
                    created_at=j.created_at,
                    created_by=j.created_by,
                    finished_at=j.finished_at,
                    processed_rows=(j.success_count or 0) + (j.error_count or 0),
                )
                for j in items
            ],
        )

    def _normalize_column_headers(self, raw: Any) -> List[Dict[str, str]]:
        """Convierte column_headers a lista de {name, display_name}. Compatible con legacy (lista de strings)."""
        if not raw or not isinstance(raw, list):
            return []
        result = []
        for h in raw:
            if isinstance(h, dict) and "name" in h:
                result.append({"name": h["name"], "display_name": h.get("display_name") or h["name"]})
            elif isinstance(h, str):
                result.append({"name": h, "display_name": h})
            else:
                result.append({"name": str(h), "display_name": str(h)})
        return result

    def get_job_headers(self, job_id: UUID) -> Optional[JobHeadersResponse]:
        """Devuelve los display_name de columnas del job (headers legibles, ej: 'Código' en vez de 'code')."""
        db = self._get_db()
        job = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
        if not job:
            return None
        normalized = self._normalize_column_headers(job.column_headers)
        headers = [h["display_name"] for h in normalized]
        return JobHeadersResponse(headers=headers)

    def get_job_rows(
        self,
        job_id: UUID,
        page: int = 1,
        per_page: int = 10,
        errors_only: bool = False,
    ) -> Optional[PaginatedJobRowsResponse]:
        """Filas del job paginadas; cada fila tiene para cada columna: valor y columna_error. Keys usan display_name (ej: 'Código'). Si errors_only=true solo filas con al menos un error. Paginación en BD."""
        db = self._get_db()
        job = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
        if not job:
            return PaginatedJobRowsResponse(page=page, per_page=per_page, total=0, total_pages=0, items=[])
        headers = self._normalize_column_headers(job.column_headers)
        query = db.query(BulkUploadJobRowModel).filter(BulkUploadJobRowModel.job_id == job_id)
        if errors_only:
            query = query.filter(BulkUploadJobRowModel.errors != {})
        total = query.count()
        total_pages = max(1, (total + per_page - 1) // per_page) if total > 0 else 0
        offset = (page - 1) * per_page
        page_rows = query.order_by(BulkUploadJobRowModel.row_index).offset(offset).limit(per_page).all()
        items = []
        for r in page_rows:
            values_raw = r.values or {}
            errors_raw = r.errors or {}
            row_values = {}
            row_errors = {}
            for h in headers:
                name = h["name"]
                display = h["display_name"]
                row_values[display] = values_raw.get(name)
                err = errors_raw.get(name)
                if err is not None:
                    row_errors[display] = err
            status = "error" if row_errors else "success"
            items.append({
                "row_index": r.row_index,
                "status": status,
                "values": row_values,
                "errors": row_errors if row_errors else None,
            })
        return PaginatedJobRowsResponse(
            page=page,
            per_page=per_page,
            total=total,
            total_pages=total_pages,
            items=items,
        )

    def get_job_by_id(self, job_id: UUID) -> Optional[BulkUploadJobDetailResponse]:
        db = self._get_db()
        job = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
        if not job:
            return None
        form_name = None
        try:
            from modules.data_collector.src.models.forms import FormModel
        except ImportError:
            from backend.modules.data_collector.src.models.forms import FormModel
        form = db.query(FormModel).filter(FormModel.id == job.form_id).first()
        if form:
            form_name = form.name or ""
        processed = (job.success_count or 0) + (job.error_count or 0)
        return BulkUploadJobDetailResponse(
            id=job.id,
            form_id=job.form_id,
            form_name=form_name,
            entity_name=job.entity_name,
            file_name=job.file_name,
            total_rows=job.total_rows,
            success_count=job.success_count,
            error_count=job.error_count,
            status=job.status.value,
            created_at=job.created_at,
            created_by=job.created_by,
            finished_at=job.finished_at,
            processed_rows=processed,
            errors=job.errors,
            column_headers=[h["display_name"] for h in self._normalize_column_headers(job.column_headers)] or None,
        )

    def get_job_errors(
        self,
        job_id: UUID,
        page: int = 1,
        per_page: int = 100,
    ) -> Optional[PaginatedBulkUploadErrorsResponse]:
        db = self._get_db()
        job = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
        if not job or not job.errors:
            return PaginatedBulkUploadErrorsResponse(page=page, per_page=per_page, total=0, items=[])
        all_errors = [BulkUploadErrorItem(**e) for e in job.errors]
        total = len(all_errors)
        offset = (page - 1) * per_page
        items = all_errors[offset : offset + per_page]
        return PaginatedBulkUploadErrorsResponse(page=page, per_page=per_page, total=total, items=items)

    def get_job_errors_excel(self, job_id: UUID) -> Tuple[BytesIO, str]:
        """Genera Excel solo con filas que tienen errores. Columnas: Fila, Col1, Col1_error, Col2, Col2_error, ..."""
        db = self._get_db()
        job = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
        if not job:
            raise ValueError("Job no encontrado")
        headers = self._normalize_column_headers(job.column_headers)
        if not headers:
            raise ValueError("Job sin columnas definidas")
        error_rows = (
            db.query(BulkUploadJobRowModel)
            .filter(
                BulkUploadJobRowModel.job_id == job_id,
                BulkUploadJobRowModel.errors != {},
            )
            .order_by(BulkUploadJobRowModel.row_index)
            .all()
        )
        if not error_rows:
            raise ValueError("No hay filas con errores para este job")
        wb = Workbook()
        ws = wb.active
        ws.title = "Errores"[:31]
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        col_num = 1
        ws.cell(row=1, column=col_num, value="Fila")
        ws.cell(row=1, column=col_num).fill = header_fill
        ws.cell(row=1, column=col_num).font = header_font
        col_num += 1
        for h in headers:
            display = h["display_name"]
            ws.cell(row=1, column=col_num, value=display)
            ws.cell(row=1, column=col_num).fill = header_fill
            ws.cell(row=1, column=col_num).font = header_font
            col_num += 1
            ws.cell(row=1, column=col_num, value=f"{display}_error")
            ws.cell(row=1, column=col_num).fill = header_fill
            ws.cell(row=1, column=col_num).font = header_font
            col_num += 1
        for row_idx, r in enumerate(error_rows, 2):
            values_raw = r.values or {}
            errors_raw = r.errors or {}
            col_num = 1
            ws.cell(row=row_idx, column=col_num, value=r.row_index)
            ws.cell(row=row_idx, column=col_num).border = border
            col_num += 1
            for h in headers:
                name = h["name"]
                display = h["display_name"]
                ws.cell(row=row_idx, column=col_num, value=values_raw.get(name))
                ws.cell(row=row_idx, column=col_num).border = border
                col_num += 1
                ws.cell(row=row_idx, column=col_num, value=errors_raw.get(name))
                ws.cell(row=row_idx, column=col_num).border = border
                col_num += 1
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio, f"errores_carga_{job_id}.xlsx"

    def create_job(
        self,
        form_id: UUID,
        entity_name: str,
        file_name: Optional[str],
        total_rows: int,
        created_by: Optional[UUID] = None,
    ) -> BulkUploadJobModel:
        db = self._get_db()
        job = BulkUploadJobModel(
            form_id=form_id,
            entity_name=entity_name,
            file_name=file_name,
            total_rows=total_rows,
            success_count=0,
            error_count=0,
            status=BulkUploadStatus.pending,
            created_by=created_by,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        return job

    def get_template_excel(self, form_id: UUID) -> Tuple[BytesIO, str]:
        """Genera plantilla Excel dinámica desde el schema del formulario. Retorna (bio, form_name)."""
        data_collector = self.container.get("data_collector")
        form_with_schema = data_collector.get_form_by_id(form_id)
        if not form_with_schema or not form_with_schema.schema:
            raise ValueError("Formulario no encontrado o sin schema")
        entity_name = getattr(form_with_schema, "entity_name", None)
        columns = data_collector.get_schema_columns_for_template(
            form_with_schema.schema,
            entity_name=entity_name,
        )
        if not columns:
            raise ValueError("El schema del formulario no tiene columnas para la plantilla")
        form_name = (form_with_schema.name or "datos").strip()
        form_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in form_name) or "datos"
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"[:31]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_num, col in enumerate(columns, 1):
            # Usar display_name del info del modelo (ej: "Código" en vez de "code") para cabeceras legibles
            cell = ws.cell(row=1, column=col_num, value=col.get("display_name") or col["name"])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for col_num, col in enumerate(columns, 1):
            hint = "Identificador" if col.get("is_logical_identifier") else ""
            cell = ws.cell(row=2, column=col_num, value=hint)
            cell.border = border
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio, form_name


    def process_upload_background(self, job_id: UUID, file_content: bytes) -> None:
        """Procesa el archivo Excel en segundo plano: validación, duplicados, relaciones, inserción por lotes."""
        MAX_ROWS = 10000
        errors_list = []
        success_count = 0
        error_count = 0
        column_headers = []

        def _finish_job(status_msg: str = ""):
            """Asegura que el job siempre tenga estado final. Usa sesión nueva por si la anterior se perdió."""
            try:
                db_fresh = self._get_db()
                job_fresh = db_fresh.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
                if job_fresh and job_fresh.status in (BulkUploadStatus.processing, BulkUploadStatus.pending):
                    job_fresh.errors = errors_list
                    job_fresh.success_count = success_count
                    job_fresh.error_count = error_count
                    job_fresh.status = BulkUploadStatus.completed if (error_count == 0 or success_count > 0) else BulkUploadStatus.error
                    job_fresh.finished_at = datetime.utcnow()
                    job_fresh.column_headers = column_headers
                    db_fresh.commit()
                    import sys
                    sys.stderr.write(f"[BULK_UPLOAD] Job {job_id} FINALIZADO: {job_fresh.status.value} ({success_count} ok, {error_count} errores){status_msg}\n")
                    sys.stderr.flush()
            except Exception as e:
                import sys
                sys.stderr.write(f"[BULK_UPLOAD] ERROR al finalizar job {job_id}: {e}\n")
                sys.stderr.flush()

        db = self._get_db()
        data_collector = self.container.get("data_collector")
        try:
            from modules.data_collector.src.models.core_registers import CoreRegisterModel, RegisterStatus
            from modules.data_collector.src.models.forms import FormModel, FormPurpose
            from modules.data_collector.src.resources.register_processor import (
                process_register_to_entity,
                find_model_by_entity_name,
            )
            from modules.data_collector.src.resources.form_auto_creator import get_logical_identifier_field
        except ImportError:
            from backend.modules.data_collector.src.models.core_registers import CoreRegisterModel, RegisterStatus
            from backend.modules.data_collector.src.models.forms import FormModel, FormPurpose
            from backend.modules.data_collector.src.resources.register_processor import (
                process_register_to_entity,
                find_model_by_entity_name,
            )
            from backend.modules.data_collector.src.resources.form_auto_creator import get_logical_identifier_field

        job = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
        if not job:
            return
        job.status = BulkUploadStatus.processing
        db.commit()
        errors_list.extend(job.errors or [])
        import sys
        sys.stderr.write(f"[BULK_UPLOAD] Iniciando procesamiento del job {job_id} (form_id={job.form_id}, entity={job.entity_name})\n")
        sys.stderr.flush()

        current_row_num = None
        current_row_dict = {}
        try:
            form_with_schema = data_collector.get_form_by_id(job.form_id)
            if not form_with_schema or not form_with_schema.schema:
                errors_list.append({"row_index": 0, "column_name": "", "message": "Formulario no encontrado o sin schema", "value": None})
                error_count += 1
                _finish_job(" - Formulario no encontrado")
                return
            columns = data_collector.get_schema_columns_for_template(
            form_with_schema.schema,
            entity_name=getattr(form_with_schema, "entity_name", None),
        )
            if not columns:
                errors_list.append({"row_index": 0, "column_name": "", "message": "Schema sin columnas", "value": None})
                error_count += 1
                _finish_job(" - Schema sin columnas")
                return
            logical_id_field = get_logical_identifier_field(form_with_schema.schema)
            schema_form_id = form_with_schema.schema_id
            if not schema_form_id:
                errors_list.append({"row_index": 0, "column_name": "", "message": "Formulario sin schema_id", "value": None})
                error_count += 1
                _finish_job(" - Sin schema_id")
                return

            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active
            # max_row/max_col pueden fallar; iterar para hallar dimensiones reales
            max_row = ws.max_row if ws.max_row is not None else 0
            max_col = ws.max_column if ws.max_column is not None else 1
            if max_row <= 1:
                actual_last = 0
                for row in ws.iter_rows(min_row=1, max_row=10002):
                    if any(cell.value is not None for cell in row):
                        actual_last = row[0].row
                max_row = max(max_row, actual_last)
            if max_col <= 1:
                for col in range(1, 500):
                    if ws.cell(row=1, column=col).value is not None:
                        max_col = col
            header_row = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
            col_index_to_name = {}
            for idx, header_cell in enumerate(header_row):
                if header_cell is None:
                    continue
                h = str(header_cell).strip()
                for col in columns:
                    if h == (col.get("display_name") or col["name"]) or h == col["name"]:
                        col_index_to_name[idx + 1] = col["name"]
                        break
            if len(col_index_to_name) == 0:
                errors_list.append({"row_index": 1, "column_name": "", "message": "Cabeceras no coinciden con el schema del formulario", "value": None})
                error_count += 1
                wb.close()
                _finish_job(" - Cabeceras no coinciden")
                return
            # Plantilla: fila 1=cabeceras. Fila 2 puede ser hints (Identificador, etc.) o datos.
            # Si fila 2 tiene solo hints/vacío en columnas de datos -> datos desde fila 3. Si no -> desde fila 2.
            data_start_row = 2
            if max_row >= 3:
                row2_vals = [ws.cell(row=2, column=c).value for c in sorted(col_index_to_name.keys())]
                def _is_hint_val(v):
                    if v is None:
                        return True
                    if isinstance(v, (int, float)):
                        return False
                    s = str(v).strip().lower()
                    return s in ("", "identificador")

                row2_looks_like_hints = all(_is_hint_val(v) for v in row2_vals)
                if row2_looks_like_hints:
                    data_start_row = 3
            max_data_row = min(max_row, MAX_ROWS + data_start_row - 1)
            total_to_process = max(0, max_data_row - data_start_row + 1)
            seen_logical_ids = set()
            model_class = find_model_by_entity_name(job.entity_name) if job.entity_name else None
            column_headers = [{"name": c["name"], "display_name": c.get("display_name") or c["name"]} for c in columns]
            PROGRESS_COMMIT_EVERY = 10  # Actualizar progress en BD cada N filas para que polling vea avance

            def _serialize_for_json(obj):
                """Convierte valores a tipos JSON-serializables (UUID, datetime, etc.)."""
                if obj is None:
                    return None
                if isinstance(obj, (UUID,)):
                    return str(obj)
                if hasattr(obj, "isoformat"):  # datetime, date
                    return obj.isoformat()
                if isinstance(obj, dict):
                    return {k: _serialize_for_json(v) for k, v in obj.items()}
                if isinstance(obj, (list, tuple)):
                    return [_serialize_for_json(v) for v in obj]
                return obj

            def _insert_row(row_index: int, values: dict, errors: dict) -> bool:
                """Inserta fila en BD. Retorna True si ok, False si falló (para reintentar con datos mínimos)."""
                try:
                    vals_ser = _serialize_for_json(values) if values else {}
                    row_model = BulkUploadJobRowModel(job_id=job_id, row_index=row_index, values=vals_ser, errors=errors or {})
                    db.add(row_model)
                    db.commit()
                    return True
                except Exception as ins_err:
                    try:
                        db.rollback()
                    except Exception:
                        pass
                    _log(f"[BULK_UPLOAD] ERROR al insertar fila {row_index} en BD: {ins_err}")
                    return False

            def _log(msg: str) -> None:
                import sys
                sys.stderr.write(msg + "\n")
                sys.stderr.flush()

            def _normalize_for_text(val):
                """Excel devuelve números como float (72157860.0). Para campos texto/DNI, quitar .0"""
                if val is None:
                    return None
                if isinstance(val, float) and val == int(val):
                    return str(int(val))
                if isinstance(val, (int, float)):
                    return str(val)
                return val

            COLUMN_FRIENDLY_NAMES = {
                "country_id": "país",
                "department_id": "departamento",
                "province_id": "provincia",
                "district_id": "distrito",
            }

            def _friendly_error_message(exc: Exception, row_dict: dict) -> tuple:
                """Convierte errores técnicos en mensajes amigables para el usuario. Retorna (mensaje, columna, valor)."""
                import re
                err_str = str(exc)
                col_name = ""
                err_value = None
                if "ForeignKeyViolation" in err_str or "foreign key" in err_str.lower():
                    m = re.search(r"Key \(([^)]+)\)=\s*\(([^)]*)\)\s+is not present in table", err_str)
                    if m:
                        col_name = m.group(1)
                        val = m.group(2).strip('"\'')
                        friendly = COLUMN_FRIENDLY_NAMES.get(col_name, col_name)
                        err_value = val if val else (row_dict.get(col_name) or "")
                        if isinstance(err_value, dict) and "id" in err_value:
                            err_value = err_value["id"]
                        return (
                            f"El valor '{err_value}' no existe en {friendly}. "
                            f"Cargue los datos de {friendly} válidos primero o verifique el valor.",
                            col_name,
                            err_value,
                        )
                    m = re.search(r"Key \(([^)]+)\)=", err_str)
                    if m:
                        col_name = m.group(1)
                        err_value = row_dict.get(col_name)
                        if isinstance(err_value, dict) and "id" in err_value:
                            err_value = err_value["id"]
                        friendly = COLUMN_FRIENDLY_NAMES.get(col_name, col_name)
                        return (
                            f"El valor '{err_value}' no existe en {friendly}. Verifique que el dato esté cargado.",
                            col_name,
                            err_value,
                        )
                if "UniqueViolation" in err_str or "unique constraint" in err_str.lower():
                    m = re.search(r"Key \(([^)]+)\)=", err_str)
                    if m:
                        col_name = m.group(1)
                        err_value = row_dict.get(col_name)
                        return (
                            f"El valor '{err_value}' ya existe. Debe ser único.",
                            col_name,
                            err_value,
                        )
                if "InvalidTextRepresentation" in err_str or "invalid input syntax for type uuid" in err_str.lower():
                    m = re.search(r"invalid input syntax for type uuid: ['\"]([^'\"]+)['\"]", err_str)
                    err_value = m.group(1) if m else None
                    m2 = re.search(r"\((\w+_id),", err_str)
                    col_name = m2.group(1) if m2 else ""
                    return (
                        f"El valor '{err_value or '?'}' no es un ID válido. Use el identificador (ej: DNI) del registro existente.",
                        col_name,
                        err_value,
                    )
                if "UndefinedFunction" in err_str or "operator does not exist" in err_str.lower() or "character varying = integer" in err_str.lower():
                    m = re.search(r"WHERE.*?\.(\w+)\s*=", err_str)
                    if m:
                        col_name = m.group(1)
                        err_value = row_dict.get(col_name)
                        return (
                            f"Error de tipo: el campo '{col_name}' espera texto pero recibió número. "
                            f"Verifique que el valor esté en formato correcto.",
                            col_name,
                            err_value,
                        )
                first_line = err_str.split("\n")[0]
                for sep in ("[SQL:", "[parameters:", "(Background"):
                    if sep in first_line:
                        first_line = first_line.split(sep)[0].strip()
                return (first_line[:250] if first_line else "Error al guardar el registro.", col_name, None)

            for row_num in range(data_start_row, max_data_row + 1):
                current_row_num = row_num
                _log(f"[BULK_UPLOAD] Fila {row_num}/{max_data_row} ({row_num - data_start_row + 1}/{total_to_process}): procesando...")
                row_errors = []
                row_dict = {}
                for col_idx, field_name in col_index_to_name.items():
                    cell_val = ws.cell(row=row_num, column=col_idx).value
                    row_dict[field_name] = cell_val
                # Normalizar números a string para campos texto/DNI (Excel devuelve 72157860.0 o 72117500 como int)
                for fn, val in list(row_dict.items()):
                    if val is not None and (isinstance(val, float) and val == int(val) or isinstance(val, int)):
                        row_dict[fn] = _normalize_for_text(val)
                current_row_dict = dict(row_dict)
                try:
                    col_spec = {c["name"]: c for c in columns}
                    for field_name, value in list(row_dict.items()):
                        spec = col_spec.get(field_name)
                        if not spec:
                            continue
                        is_optional = spec.get("is_optional", True)
                        if value is None or (isinstance(value, str) and not value.strip()):
                            if not is_optional:
                                row_errors.append((field_name, "Campo obligatorio vacío", value))
                            continue
                        type_value = spec.get("type_value", "text")
                        if type_value == "entity":
                            related_entity = spec.get("foreign_key_table")
                            if related_entity and value:
                                related_form = data_collector.get_form_by_entity_name(related_entity)
                                rel_model = find_model_by_entity_name(related_entity)
                                row = None
                                if related_form and related_form.schema:
                                    lid_field = get_logical_identifier_field(related_form.schema)
                                    if lid_field and rel_model:
                                        # Normalizar a string para evitar "varchar = integer" (Excel devuelve int)
                                        norm_val = _normalize_for_text(value) or value
                                        row = db.query(rel_model).filter(getattr(rel_model, lid_field) == str(norm_val)).first()
                                if row is None and rel_model:
                                    # Sin formulario: intentar buscar por id, name o code
                                    val_str = str(_normalize_for_text(value) or value).strip()
                                    for attr in ("id", "name", "code"):
                                        if hasattr(rel_model, attr):
                                            try:
                                                row = db.query(rel_model).filter(getattr(rel_model, attr) == val_str).first()
                                                if not row and val_str != value:
                                                    row = db.query(rel_model).filter(getattr(rel_model, attr) == value).first()
                                                if row:
                                                    break
                                            except Exception:
                                                try:
                                                    db.rollback()
                                                except Exception:
                                                    pass
                                                continue
                                if row:
                                    row_dict[field_name] = {"id": str(row.id), "display_name": str(getattr(row, "name", row.id))}
                                elif rel_model:
                                    # Intentar buscar por dni, code, name (identificadores comunes)
                                    val_str = str(_normalize_for_text(value) or value).strip()
                                    row = None
                                    for attr in ("dni", "code", "name", "id"):
                                        if hasattr(rel_model, attr):
                                            try:
                                                row = db.query(rel_model).filter(getattr(rel_model, attr) == val_str).first()
                                                if row:
                                                    row_dict[field_name] = {"id": str(row.id), "display_name": str(getattr(row, "name", row.id))}
                                                    break
                                            except Exception:
                                                try:
                                                    db.rollback()
                                                except Exception:
                                                    pass
                                                continue
                                    if not isinstance(row_dict.get(field_name), dict):
                                        row_errors.append((field_name, f"Entidad {related_entity} no encontrada con valor '{value}'", value))
                                else:
                                    row_errors.append((field_name, f"Entidad {related_entity} no encontrada", value))
                        elif type_value == "number":
                            try:
                                if isinstance(value, (int, float)):
                                    row_dict[field_name] = value
                                else:
                                    row_dict[field_name] = float(value) if "." in str(value) else int(value)
                            except (ValueError, TypeError):
                                row_errors.append((field_name, "Valor numérico inválido", value))
                        elif type_value == "boolean":
                            row_dict[field_name] = str(value).lower() in ("true", "1", "sí", "si", "yes")
                        else:
                            # type text: normalizar 72157860.0 -> "72157860" (Excel devuelve float para números)
                            norm = _normalize_for_text(value)
                            row_dict[field_name] = norm if (norm is not None and (not isinstance(norm, str) or norm.strip())) else None

                    if logical_id_field and model_class:
                        lid_val = row_dict.get(logical_id_field)
                        if lid_val is None or (isinstance(lid_val, str) and not lid_val.strip()):
                            row_errors.append((logical_id_field, "Identificador lógico obligatorio", lid_val))
                        else:
                            lid_str = str(_normalize_for_text(lid_val) or lid_val).strip()
                            if lid_str in seen_logical_ids:
                                row_errors.append((logical_id_field, "Identificador duplicado en el archivo", lid_val))
                            else:
                                existing = db.query(model_class).filter(getattr(model_class, logical_id_field) == lid_str).first()
                                if existing:
                                    row_errors.append((logical_id_field, "Identificador ya existe en el sistema", lid_val))
                                else:
                                    seen_logical_ids.add(lid_str)

                    if row_errors:
                        err_msgs = "; ".join(f"{c}: {m}" for c, m, _ in row_errors)
                        _log(f"[BULK_UPLOAD] Fila {row_num}: ERROR - {err_msgs}")
                        for col_name, msg, val in row_errors:
                            errors_list.append({"row_index": row_num, "column_name": col_name, "message": msg, "value": val})
                        row_errors_dict = {col_name: msg for col_name, msg, _ in row_errors}
                        if not _insert_row(row_num, dict(row_dict), row_errors_dict):
                            _insert_row(row_num, {"_raw": str(row_dict)[:500]}, row_errors_dict)
                        error_count += 1
                        try:
                            db.rollback()
                        except Exception:
                            pass
                        job_ctx = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
                        if job_ctx:
                            job_ctx.success_count = success_count
                            job_ctx.error_count = error_count
                            job_ctx.errors = errors_list
                            db.commit()
                        continue

                    detail = []
                    for col in columns:
                        template_name = col["name"]
                        schema_name = col.get("schema_field_name") or template_name
                        val = row_dict.get(template_name)
                        if val is None and col.get("is_optional", True):
                            continue
                        type_val = col.get("type_value", "text")
                        item = {"name": schema_name, "value": val, "type_value": type_val}
                        if type_val == "entity" and isinstance(val, dict):
                            item["value"] = val
                        detail.append(item)
                    try:
                        register = CoreRegisterModel(
                            form_id=job.form_id,
                            schema_form_id=schema_form_id,
                            detail=detail,
                            status=RegisterStatus.success,
                            entity_name=job.entity_name,
                        )
                        db.add(register)
                        db.commit()
                        db.refresh(register)
                        _log(f"[BULK_UPLOAD] Fila {row_num}: guardando en entidad...")
                        process_register_to_entity(register, db, FormModel, FormPurpose)
                        db.commit()
                        _insert_row(row_num, dict(row_dict), {})
                        success_count += 1
                        _log(f"[BULK_UPLOAD] Fila {row_num}: OK")
                    except Exception as e:
                        try:
                            try:
                                db.rollback()
                                db.expire_all()
                            except Exception:
                                pass
                            _log(f"[BULK_UPLOAD] Fila {row_num}: EXCEPCIÓN - {e}")
                            try:
                                err_msg, col_name, err_value = _friendly_error_message(e, row_dict)
                            except Exception:
                                err_msg, col_name, err_value = str(e)[:250], "", None
                            errors_list.append({"row_index": row_num, "column_name": col_name, "message": err_msg, "value": err_value})
                            err_dict = {col_name: err_msg} if col_name else {"general": str(e)[:250]}
                            if not _insert_row(row_num, dict(row_dict), err_dict):
                                _insert_row(row_num, {"_raw": str(row_dict)[:500]}, err_dict)
                            error_count += 1
                            job_ctx = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
                            if job_ctx:
                                job_ctx.success_count = success_count
                                job_ctx.error_count = error_count
                                job_ctx.errors = errors_list
                                db.commit()
                        except Exception as handler_err:
                            _log(f"[BULK_UPLOAD] Error al registrar excepción de fila {row_num}: {handler_err}")
                            error_count += 1
                            try:
                                db.rollback()
                            except Exception:
                                pass
                        continue

                except Exception as e:
                    """Captura errores de logical_id check, entity resolution o cualquier otro en el procesamiento de la fila."""
                    try:
                        try:
                            db.rollback()
                            db.expire_all()
                        except Exception:
                            pass
                        _log(f"[BULK_UPLOAD] Fila {row_num}: EXCEPCIÓN - {e}")
                        try:
                            err_msg, col_name, err_value = _friendly_error_message(e, row_dict)
                        except Exception:
                            err_msg, col_name, err_value = str(e)[:250], "", None
                        errors_list.append({"row_index": row_num, "column_name": col_name, "message": err_msg, "value": err_value})
                        err_dict = {col_name: err_msg} if col_name else {"general": str(e)[:250]}
                        if not _insert_row(row_num, dict(row_dict), err_dict):
                            _insert_row(row_num, {"_raw": str(row_dict)[:500]}, err_dict)
                        error_count += 1
                        job_ctx = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
                        if job_ctx:
                            job_ctx.success_count = success_count
                            job_ctx.error_count = error_count
                            job_ctx.errors = errors_list
                            db.commit()
                    except Exception as handler_err:
                        _log(f"[BULK_UPLOAD] Error al registrar excepción de fila {row_num}: {handler_err}")
                        error_count += 1
                        try:
                            db.rollback()
                        except Exception:
                            pass
                    continue

                # Actualizar progress periódicamente para que GET /jobs/{id} muestre avance
                processed = success_count + error_count
                if processed > 0 and processed % PROGRESS_COMMIT_EVERY == 0:
                    job_ctx = db.query(BulkUploadJobModel).filter(BulkUploadJobModel.id == job_id).first()
                    if job_ctx:
                        job_ctx.success_count = success_count
                        job_ctx.error_count = error_count
                        job_ctx.errors = errors_list
                        db.commit()
                        import sys
                        sys.stderr.write(f"[BULK_UPLOAD] Procesando fila {row_num} de {max_data_row} ({processed}/{total_to_process} procesadas, {success_count} ok, {error_count} errores)\n")
                        sys.stderr.flush()

            wb.close()
        except Exception as e:
            errors_list.append({"row_index": current_row_num or 0, "column_name": "", "message": f"Error procesando archivo: {e}", "value": None})
            error_count += 1
            try:
                db.rollback()
            except Exception:
                pass
            try:
                if current_row_num is not None or current_row_dict:
                    row_idx = current_row_num if current_row_num is not None else 0
                    vals = current_row_dict if current_row_dict else {}
                    err_dict = {"general": str(e)[:250]}
                    if not _insert_row(row_idx, vals, err_dict):
                        _insert_row(row_idx, {"_raw": str(vals)[:500] if vals else "Error antes de procesar filas"}, err_dict)
            except Exception as ins_err:
                import sys
                sys.stderr.write(f"[BULK_UPLOAD] No se pudo registrar fila con error en BD: {ins_err}\n")
                sys.stderr.flush()
            import sys
            sys.stderr.write(f"[BULK_UPLOAD] EXCEPCIÓN en job {job_id}: {e}\n")
            sys.stderr.flush()
        finally:
            _finish_job()
