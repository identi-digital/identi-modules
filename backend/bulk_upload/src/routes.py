from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse
from starlette.requests import Request
from typing import Optional
from uuid import UUID
from io import BytesIO

from .schemas import (
    PaginatedBulkUploadJobsResponse,
    BulkUploadJobDetailResponse,
    PaginatedBulkUploadErrorsResponse,
    JobHeadersResponse,
    PaginatedJobRowsResponse,
    UploadAcceptedResponse,
)

router = APIRouter(prefix="/bulk-upload", tags=["Bulk Upload"])


def get_funcionalities(request: Request):
    return request.app.state.container.get("bulk_upload")


@router.get("/template")
def get_template(
    form_id: UUID = Query(..., description="ID del formulario"),
    svc=Depends(get_funcionalities),
):
    """Genera plantilla Excel dinámica. Requiere form_id."""
    import sys
    sys.stderr.write(f"[BULK_TEMPLATE] GET /template form_id={form_id}\n")
    sys.stderr.flush()
    try:
        bio, form_name = svc.get_template_excel(form_id=form_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    filename = f"plantilla_{form_name}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



@router.get("/jobs", response_model=PaginatedBulkUploadJobsResponse)
def get_jobs(
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    entity_name: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    svc=Depends(get_funcionalities),
):
    """Listado paginado de cargas masivas (entidad, fecha, total, exitosos, estado)."""
    return svc.get_jobs(page=page, per_page=per_page, entity_name=entity_name, status=status)


@router.get("/jobs/{job_id}", response_model=BulkUploadJobDetailResponse)
def get_job_detail(job_id: UUID, svc=Depends(get_funcionalities)):
    """Detalle de un job: status, total_rows, success_count, error_count, created_at, finished_at."""
    job = svc.get_job_by_id(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return job


@router.get("/jobs/{job_id}/headers", response_model=JobHeadersResponse)
def get_job_headers(job_id: UUID, svc=Depends(get_funcionalities)):
    """Devuelve los nombres de columnas (headers) del job, en el mismo orden que la plantilla."""
    result = svc.get_job_headers(job_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return result


@router.get("/jobs/{job_id}/rows", response_model=PaginatedJobRowsResponse)
def get_job_rows(
    job_id: UUID,
    page: int = Query(1, ge=1, description="Página (1-based)"),
    per_page: int = Query(25, ge=1, le=100, description="Filas por página (máx 100)"),
    errors_only: bool = Query(False, description="Si true, solo filas con al menos un error"),
    svc=Depends(get_funcionalities),
):
    """Filas del job paginadas. Cada fila incluye valor y columna_error por columna. Útil para jobs con muchas filas (hasta 10.000)."""
    result = svc.get_job_rows(job_id, page=page, per_page=per_page, errors_only=errors_only)
    if result is None:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return result


@router.get("/jobs/{job_id}/errors/excel")
def get_job_errors_excel(job_id: UUID, svc=Depends(get_funcionalities)):
    """Descarga Excel solo con filas que tienen errores. Columnas: Fila, Col1, Col1_error, Col2, Col2_error, ..."""
    try:
        bio, filename = svc.get_job_errors_excel(job_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/upload/{form_id}", response_model=UploadAcceptedResponse, status_code=202)
async def upload_file(
    form_id: UUID,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    svc=Depends(get_funcionalities),
):
    """Acepta archivo Excel y encola procesamiento en segundo plano. Límite 10.000 filas. Requiere form_id. El entity_name se obtiene automáticamente del formulario."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Se requiere un archivo .xlsx")
    data_collector = svc.container.get("data_collector")
    form_with_schema = data_collector.get_form_by_id(form_id)
    if not form_with_schema:
        raise HTTPException(status_code=400, detail="Formulario no encontrado. Indique un form_id válido.")
    content = await file.read()
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    # max_row puede fallar en algunos Excel; iterar para hallar la última fila con datos
    max_row = ws.max_row if ws.max_row is not None else 0
    if max_row <= 1:
        actual_last = 0
        for row in ws.iter_rows(min_row=1, max_row=10002):
            if any(cell.value is not None for cell in row):
                actual_last = row[0].row
        max_row = max(max_row, actual_last)
    # Plantilla: fila 1=cabeceras, fila 2=hints, fila 3+=datos. O fila 1=cabeceras, fila 2+=datos (sin hints)
    data_rows = max(0, max_row - 2) if max_row >= 3 else max(0, max_row - 1)
    wb.close()
    if data_rows == 0:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo parece no tener filas de datos (max_row={max_row}). Fila 1=cabeceras; si usa hints, fila 2=hints y desde fila 3 los datos.",
        )
    if data_rows > 10000:
        raise HTTPException(status_code=400, detail="Máximo 10.000 filas de datos permitidas.")
    job = svc.create_job(
        form_id=form_with_schema.id,
        entity_name=form_with_schema.entity_name or "unknown",
        file_name=file.filename,
        total_rows=data_rows,
    )
    background_tasks.add_task(svc.process_upload_background, job.id, content)
    return UploadAcceptedResponse(
        job_id=job.id,
        message=f"Carga aceptada. Procesando {data_rows} filas en segundo plano. Consulte GET /bulk-upload/jobs/{job.id} para ver el progreso (processed_rows, success_count, error_count).",
    )
