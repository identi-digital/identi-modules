from datetime import datetime
from typing import Optional
from uuid import UUID
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, distinct
from .models.deforestation_requests import DeforestationRequestModel, DeforestationRequestStatusEnum
from modules.farmers.src.models.farms import FarmModel
from modules.locations.src.models.countries import CountryModel
from modules.locations.src.models.departments import DepartmentModel
from modules.locations.src.models.provinces import ProvinceModel
from modules.farmers.src.models.farmers import FarmerModel
from modules.locations.src.models.districts import DistrictModel
from .schemas import (
    DeforestationRequestResponse,
    GFWValidationRequest, GFWValidationResponse,
    PaginatedFarmDeforestationResponse, FarmDeforestationResponse,
    FarmDeforestationMetricsResponse, FarmerDeforestationMetricsResponse,
    DeforestationStatusCount, DeforestationStateEnum,
    FarmGeoreferenceMetricsResponse
)
from .resources import geometry_to_geojson

class Funcionalities:
    def __init__(self, container, database_key: str = "core_db"):
        self.container = container
        self.database_key = database_key
        
    def _get_db(self) -> Session:
        """Obtiene la sesión de base de datos desde el container"""
        return self.container.get(self.database_key, "databases")
    
    def get_deforestation_request_by_farm_id(self, farm_id: UUID) -> Optional[DeforestationRequestResponse]:
        """Obtiene el registro de deforestación de una farm específica"""
        db = self._get_db()
        request = db.query(DeforestationRequestModel).filter(
            DeforestationRequestModel.farm_id == farm_id,
            DeforestationRequestModel.disabled_at.is_(None)
        ).first()
        
        if not request:
            return None
        
        return DeforestationRequestResponse.model_validate(request)
    
    def validate_gfw_request(self, validation_data: GFWValidationRequest, token: Optional[str] = None) -> GFWValidationResponse:
        """
        Valida el estado de una solicitud de análisis de deforestación en GFW.
        
        Args:
            validation_data: Objeto con request_id, api_url (opcional)
            token: Authorization token
            
        Returns:
            GFWValidationResponse con el estado actual de la solicitud
        """
        from .environment import GFW_API_URL
        from .services.gfw import GeoJSONTransformer
        
        # Usar valores del request o valores por defecto de environment
        api_url = validation_data.api_url or GFW_API_URL
        
        if not token:
            raise ValueError("Authorization token is required. Provide it in the request header.")
        
        print(f"🔍 Validando request de GFW: {validation_data.request_id}")
        
        gfw_service = GeoJSONTransformer()
        validation_response = gfw_service.request_validation(
            request_id=validation_data.request_id,
            api_url=api_url,
            token=token
        )
        
        if not validation_response:
            raise ValueError(f"No se pudo obtener validación de GFW para request_id: {validation_data.request_id}")
        
        return GFWValidationResponse(**validation_response)
    
    # ========== DEFORESTATION STATUS METHODS ==========
    def get_farms_deforestation_paginated(
        self,
        page: int = 1,
        per_page: int = 10,
        status: Optional[str] = None,  # "baja", "media", "alta", "sin_datos"
        sort_by: Optional[str] = None,
        order: Optional[str] = "asc",
        search: str = ""
    ) -> PaginatedFarmDeforestationResponse:
        """
        Obtiene una lista paginada de farms con estado de deforestación.
        
        IMPORTANTE: Solo retorna parcelas que tienen un deforestation_request 
        con status COMPLETED (análisis completado exitosamente).
        
        El estado se determina basado en natural_forest_loss_ha:
        - "baja/nula": loss === 0
        - "parcial": 0 < loss <= 0.2
        - "crítica": loss > 0.2
        
        Args:
            page: Número de página
            per_page: Elementos por página
            status: Filtro por estado ("baja/nula", "parcial", "crítica")
            sort_by: Campo para ordenar
            order: Orden ascendente o descendente
            search: Búsqueda por nombre de parcela o productor
        
        Returns:
            PaginatedFarmDeforestationResponse con farms y sus estados de deforestación
        """
        db = self._get_db()
        
        # Query base con joins
        # IMPORTANTE: Solo incluimos parcelas que tengan un deforestation_request con status COMPLETED
        query = db.query(
            FarmModel,
            FarmerModel,
            DistrictModel,
            DeforestationRequestModel
        ).join(
            DeforestationRequestModel, 
            and_(
                DeforestationRequestModel.farm_id == FarmModel.id,
                DeforestationRequestModel.disabled_at.is_(None),
                DeforestationRequestModel.status == DeforestationRequestStatusEnum.COMPLETED
            )
        ).outerjoin(
            FarmerModel, FarmModel.farmer_id == FarmerModel.id
        ).outerjoin(
            DistrictModel, FarmModel.district_id == DistrictModel.id
        ).filter(
            FarmModel.disabled_at.is_(None)
        )
        
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(
                or_(
                    FarmModel.name.ilike(f"%{search}%"),
                    FarmerModel.first_name.ilike(f"%{search}%"),
                    FarmerModel.last_name.ilike(f"%{search}%")
                )
            )
        
        # Aplicar filtro por estado de deforestación si se proporciona
        if status:
            if status == "baja/nula":
                query = query.filter(
                    and_(
                        DeforestationRequestModel.natural_forest_loss_ha.isnot(None),
                        DeforestationRequestModel.natural_forest_loss_ha == 0
                    )
                )
            elif status == "parcial":
                query = query.filter(
                    and_(
                        DeforestationRequestModel.natural_forest_loss_ha > 0,
                        DeforestationRequestModel.natural_forest_loss_ha <= 0.2
                    )
                )
            elif status == "crítica":
                query = query.filter(
                    DeforestationRequestModel.natural_forest_loss_ha > 0.2
                )
        
        # Aplicar ordenamiento
        if sort_by:
            if sort_by == "code":
                sort_column = FarmModel.name
            elif sort_by == "producer_full_name":
                sort_column = FarmerModel.first_name
            elif sort_by == "natural_forest_loss_ha":
                sort_column = DeforestationRequestModel.natural_forest_loss_ha
            elif sort_by == "state_deforesting":
                # Ordenar por cobertura para ordenar indirectamente por estado
                sort_column = DeforestationRequestModel.natural_forest_loss_ha
            else:
                sort_column = getattr(FarmModel, sort_by, None)
            
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            query = query.order_by(FarmModel.created_at.desc())
        
        # Calcular totales y paginación
        total = query.count()
        total_pages = (total + per_page - 1) // per_page
        offset = (page - 1) * per_page
        
        # Obtener resultados paginados
        results = query.offset(offset).limit(per_page).all()
        
        # Construir respuesta
        items = []
        for farm, farmer, district, deforestation_request in results:
            # Calcular estado de deforestación
            # Como solo incluimos parcelas con deforestation_request COMPLETED,
            # siempre tendremos un deforestation_request válido
            loss = float(deforestation_request.natural_forest_loss_ha) if deforestation_request.natural_forest_loss_ha is not None else 0
            natural_forest_loss_ha = loss
            
            if loss == 0:
                state_deforesting = DeforestationStateEnum.BAJA_NULA
            elif 0 < loss <= 0.2:
                state_deforesting = DeforestationStateEnum.PARCIAL
            else:  # loss > 0.2
                state_deforesting = DeforestationStateEnum.CRITICA
            
            # Obtener descripción del distrito
            district_description = district.description if district else None
            # obtengo la provinvincia
            province = db.query(ProvinceModel).filter(ProvinceModel.id == farm.province_id).first()
            country = db.query(CountryModel).filter(CountryModel.id == farm.country_id).first()
            department = db.query(DepartmentModel).filter(DepartmentModel.id == farm.department_id).first()
            items.append(FarmDeforestationResponse(
                id=farm.id,
                code=farm.name,  # Usando name como code
                farmer=farmer,
                district=district,
                province=province,
                country=country,
                department=department,
                total_area=farm.total_area,
                geometry=geometry_to_geojson(farm.geometry, db) if farm.geometry else None,
                district_description=district_description,
                state_deforesting=state_deforesting,
                natural_forest_loss_ha=natural_forest_loss_ha,
                deforestation_request=deforestation_request,
                created_at=farm.created_at
            ))
        
        return PaginatedFarmDeforestationResponse(
            items=items,
            total=total,
            page=page,
            page_size=per_page,
            total_pages=total_pages
        )
    
    def get_farm_deforestation_metrics(self) -> FarmDeforestationMetricsResponse:
        """
        Obtiene métricas de evaluación de deforestación de parcelas.
        
        Solo considera parcelas con deforestation_request status COMPLETED.
        Calcula:
        - Total de hectáreas evaluadas (suma de farm.total_area)
        - Total de parcelas evaluadas
        - Cantidad y porcentaje por cada estado (baja/nula, parcial, crítica)
        
        Returns:
            FarmDeforestationMetricsResponse con las métricas de deforestación
        """
        db = self._get_db()
        
        try:
            # Query base: parcelas con deforestation_request COMPLETED
            query = db.query(
                FarmModel,
                DeforestationRequestModel
            ).join(
                DeforestationRequestModel,
                and_(
                    DeforestationRequestModel.farm_id == FarmModel.id,
                    DeforestationRequestModel.disabled_at.is_(None),
                    DeforestationRequestModel.status == DeforestationRequestStatusEnum.COMPLETED
                )
            ).filter(
                FarmModel.disabled_at.is_(None)
            )
            
            # Obtener todas las parcelas evaluadas
            evaluated_farms = query.all()
            
            # Inicializar contadores
            total_farms = len(evaluated_farms)
            total_hectares = 0.0
            count_baja_nula = 0
            count_parcial = 0
            count_critica = 0
            
            # Procesar cada parcela
            for farm, deforestation_request in evaluated_farms:
                # Sumar hectáreas
                if farm.total_area:
                    total_hectares += float(farm.total_area)
                
                # Clasificar por estado de deforestación
                if deforestation_request.natural_forest_loss_ha is not None:
                    loss = float(deforestation_request.natural_forest_loss_ha)
                    
                    if loss == 0:
                        count_baja_nula += 1
                    elif 0 < loss <= 0.2:
                        count_parcial += 1
                    else:  # loss > 0.2
                        count_critica += 1
            
            # Calcular porcentajes
            if total_farms > 0:
                percentage_baja_nula = round((count_baja_nula / total_farms) * 100, 2)
                percentage_parcial = round((count_parcial / total_farms) * 100, 2)
                percentage_critica = round((count_critica / total_farms) * 100, 2)
            else:
                percentage_baja_nula = 0.0
                percentage_parcial = 0.0
                percentage_critica = 0.0
            
            return FarmDeforestationMetricsResponse(
                total_hectares_evaluated=round(total_hectares, 2),
                total_farms_evaluated=total_farms,
                baja_nula=DeforestationStatusCount(
                    count=count_baja_nula,
                    percentage=percentage_baja_nula
                ),
                parcial=DeforestationStatusCount(
                    count=count_parcial,
                    percentage=percentage_parcial
                ),
                critica=DeforestationStatusCount(
                    count=count_critica,
                    percentage=percentage_critica
                )
            )
            
        except Exception as e:
            print(f"❌ Error al obtener métricas de deforestación: {e}")
            raise e
    
    def get_farmer_deforestation_metrics(self) -> FarmerDeforestationMetricsResponse:
        """
        Obtiene métricas de evaluación de deforestación de productores.
        
        Solo considera productores que tienen al menos una farm con deforestation_request status COMPLETED.
        Clasifica a cada productor según el promedio de natural_forest_loss_ha de todas sus farms evaluadas.
        
        Calcula:
        - Total de productores evaluados
        - Cantidad y porcentaje por cada estado según promedio (baja/nula, parcial, crítica)
        
        Returns:
            FarmerDeforestationMetricsResponse con las métricas de productores
        """
        db = self._get_db()
        
        try:
            # Obtener todos los productores con al menos una farm con deforestation_request COMPLETED
            # Agrupamos por farmer_id para obtener lista de farms evaluadas por productor
            
            # Obtener IDs de productores que tienen farms evaluadas
            farmer_ids_query = db.query(
                distinct(FarmModel.farmer_id)
            ).join(
                DeforestationRequestModel,
                and_(
                    DeforestationRequestModel.farm_id == FarmModel.id,
                    DeforestationRequestModel.disabled_at.is_(None),
                    DeforestationRequestModel.status == DeforestationRequestStatusEnum.COMPLETED
                )
            ).filter(
                FarmModel.disabled_at.is_(None),
                FarmModel.farmer_id.isnot(None)
            ).all()
            
            farmer_ids = [farmer_id[0] for farmer_id in farmer_ids_query]
            
            # Inicializar contadores
            total_farmers = len(farmer_ids)
            count_baja_nula = 0
            count_parcial = 0
            count_critica = 0
            
            # Para cada productor, calcular el promedio de pérdida de bosque
            for farmer_id in farmer_ids:
                # Obtener todas las farms evaluadas de este productor
                farms_query = db.query(
                    DeforestationRequestModel.natural_forest_loss_ha
                ).join(
                    FarmModel,
                    and_(
                        FarmModel.id == DeforestationRequestModel.farm_id,
                        FarmModel.farmer_id == farmer_id
                    )
                ).filter(
                    FarmModel.disabled_at.is_(None),
                    DeforestationRequestModel.disabled_at.is_(None),
                    DeforestationRequestModel.status == DeforestationRequestStatusEnum.COMPLETED,
                    DeforestationRequestModel.natural_forest_loss_ha.isnot(None)
                ).all()
                
                # Calcular promedio
                if farms_query:
                    loss_values = [float(loss[0]) for loss in farms_query]
                    avg_loss = sum(loss_values) / len(loss_values)
                    
                    # Clasificar al productor según el promedio
                    if avg_loss == 0:
                        count_baja_nula += 1
                    elif 0 < avg_loss <= 0.2:
                        count_parcial += 1
                    else:  # avg_loss > 0.2
                        count_critica += 1
            
            # Calcular porcentajes
            if total_farmers > 0:
                percentage_baja_nula = round((count_baja_nula / total_farmers) * 100, 2)
                percentage_parcial = round((count_parcial / total_farmers) * 100, 2)
                percentage_critica = round((count_critica / total_farmers) * 100, 2)
            else:
                percentage_baja_nula = 0.0
                percentage_parcial = 0.0
                percentage_critica = 0.0
            
            return FarmerDeforestationMetricsResponse(
                total_farmers_evaluated=total_farmers,
                baja_nula=DeforestationStatusCount(
                    count=count_baja_nula,
                    percentage=percentage_baja_nula
                ),
                parcial=DeforestationStatusCount(
                    count=count_parcial,
                    percentage=percentage_parcial
                ),
                critica=DeforestationStatusCount(
                    count=count_critica,
                    percentage=percentage_critica
                )
            )
            
        except Exception as e:
            print(f"❌ Error al obtener métricas de productores: {e}")
            raise e
    
    def export_farms_deforestation_to_excel(
        self,
        status: Optional[str] = None,
        sort_by: Optional[str] = None,
        order: Optional[str] = "asc",
        search: str = ""
    ):
        """
        Exporta parcelas con deforestación a Excel.
        
        Usa los mismos filtros que la API paginada.
        Solo incluye parcelas con deforestation_request status COMPLETED.
        
        Args:
            status: Filtro por estado ("baja/nula", "parcial", "crítica")
            sort_by: Campo para ordenar
            order: Orden ascendente o descendente
            search: Búsqueda por nombre de parcela o productor
            
        Returns:
            BytesIO con el archivo Excel generado
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        db = self._get_db()
        
        try:
            # Query base (misma lógica que la paginación)
            query = db.query(
                FarmModel,
                FarmerModel,
                DistrictModel,
                DeforestationRequestModel
            ).join(
                DeforestationRequestModel, 
                and_(
                    DeforestationRequestModel.farm_id == FarmModel.id,
                    DeforestationRequestModel.disabled_at.is_(None),
                    DeforestationRequestModel.status == DeforestationRequestStatusEnum.COMPLETED
                )
            ).outerjoin(
                FarmerModel, FarmModel.farmer_id == FarmerModel.id
            ).outerjoin(
                DistrictModel, FarmModel.district_id == DistrictModel.id
            ).filter(
                FarmModel.disabled_at.is_(None)
            )
            
            # Aplicar búsqueda
            if search:
                query = query.filter(
                    or_(
                        FarmModel.name.ilike(f"%{search}%"),
                        FarmerModel.first_name.ilike(f"%{search}%"),
                        FarmerModel.last_name.ilike(f"%{search}%")
                    )
                )
            
            # Aplicar filtro por estado
            if status:
                if status == "baja/nula":
                    query = query.filter(
                        and_(
                            DeforestationRequestModel.natural_forest_loss_ha.isnot(None),
                            DeforestationRequestModel.natural_forest_loss_ha == 0
                        )
                    )
                elif status == "parcial":
                    query = query.filter(
                        and_(
                            DeforestationRequestModel.natural_forest_loss_ha > 0,
                            DeforestationRequestModel.natural_forest_loss_ha <= 0.2
                        )
                    )
                elif status == "crítica":
                    query = query.filter(
                        DeforestationRequestModel.natural_forest_loss_ha > 0.2
                    )
            
            # Aplicar ordenamiento
            if sort_by:
                if sort_by == "code":
                    sort_column = FarmModel.name
                elif sort_by == "producer_full_name":
                    sort_column = FarmerModel.first_name
                elif sort_by == "natural_forest_loss_ha":
                    sort_column = DeforestationRequestModel.natural_forest_loss_ha
                elif sort_by == "state_deforesting":
                    sort_column = DeforestationRequestModel.natural_forest_loss_ha
                else:
                    sort_column = getattr(FarmModel, sort_by, None)
                
                if sort_column is not None:
                    if order and order.lower() == "desc":
                        query = query.order_by(sort_column.desc())
                    else:
                        query = query.order_by(sort_column.asc())
            else:
                query = query.order_by(FarmModel.created_at.desc())
            
            # Obtener todos los resultados (sin paginación)
            results = query.all()
            
            # Crear el libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Parcelas Deforestación"
            
            # Estilo para el encabezado
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Encabezados
            headers = [
                "ID Parcela",
                "Código/Nombre Parcela",
                "Productor",
                "Distrito",
                "Estado Deforestación",
                "Pérdida Bosque Natural (ha)",
                "ID Solicitud Deforestación",
                "Fecha Creación Parcela"
            ]
            ws.append(headers)
            
            # Aplicar estilo al encabezado
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Agregar datos
            for farm, farmer, district, deforestation_request in results:
                # Calcular estado de deforestación
                loss = float(deforestation_request.natural_forest_loss_ha) if deforestation_request.natural_forest_loss_ha is not None else 0
                
                if loss == 0:
                    state_deforesting = "baja/nula"
                elif 0 < loss <= 0.2:
                    state_deforesting = "parcial"
                else:
                    state_deforesting = "crítica"
                
                # Construir nombre completo del productor
                producer_full_name = ""
                if farmer:
                    first_name = farmer.first_name or ""
                    last_name = farmer.last_name or ""
                    producer_full_name = f"{first_name} {last_name}".strip()
                
                # Agregar fila
                ws.append([
                    str(farm.id),
                    farm.name or "",
                    producer_full_name,
                    district.description if district else "",
                    state_deforesting,
                    float(loss) if loss is not None else 0.0,
                    str(deforestation_request.id) if deforestation_request else "",
                    farm.created_at.strftime("%Y-%m-%d %H:%M:%S") if farm.created_at else ""
                ])
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 38,  # ID Parcela (UUID)
                'B': 25,  # Código/Nombre Parcela
                'C': 30,  # Productor
                'D': 25,  # Distrito
                'E': 20,  # Estado Deforestación
                'F': 25,  # Pérdida Bosque Natural
                'G': 38,  # ID Solicitud
                'H': 20   # Fecha Creación
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Guardar en BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
            
        except Exception as e:
            print(f"❌ Error al exportar parcelas con deforestación a Excel: {e}")
            raise e
    
    def get_farm_georeference_metrics(self) -> FarmGeoreferenceMetricsResponse:
        """
        Obtiene métricas de georreferenciación de parcelas (farms).
        
        Calcula:
        - Número y porcentaje de parcelas con geometría
        - Número y porcentaje de parcelas sin geometría
        
        Solo considera parcelas activas (disabled_at IS NULL).
        
        Returns:
            FarmGeoreferenceMetricsResponse con las métricas
        """
        db = self._get_db()
        
        try:
            # Contar parcelas activas totales
            total_farms = db.query(FarmModel).filter(
                FarmModel.disabled_at.is_(None)
            ).count()
            
            # Contar parcelas con geometría
            farms_with_geometry = db.query(FarmModel).filter(
                FarmModel.disabled_at.is_(None),
                FarmModel.geometry.isnot(None)
            ).count()
            
            # Contar parcelas sin geometría
            farms_without_geometry = total_farms - farms_with_geometry
            
            # Calcular porcentajes
            if total_farms > 0:
                coverage_with = round((farms_with_geometry / total_farms) * 100, 2)
                coverage_without = round((farms_without_geometry / total_farms) * 100, 2)
            else:
                coverage_with = 0.0
                coverage_without = 0.0
            
            return FarmGeoreferenceMetricsResponse(
                farm_georefrence_count=farms_with_geometry,
                farm_georefrence_coverage=coverage_with,
                farm_wh_georefeence_count=farms_without_geometry,
                farm_wh_georefeence_coverage=coverage_without
            )
            
        except Exception as e:
            print(f"Error al obtener métricas de georreferenciación: {e}")
            import traceback
            traceback.print_exc()
            raise e
