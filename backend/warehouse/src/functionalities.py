from datetime import datetime
from typing import Optional
from uuid import UUID
from sqlalchemy.orm import Session
from sqlalchemy import func, case, cast, Date

from modules.gathering.src.models.lots import LotModel

from .models.store_centers import StoreCenterModel
from .models.store_movement import StoreMovementModel
from .schemas import (
    StoreCenterCreate, StoreCenterUpdate, StoreCenterResponse, PaginatedStoreCenterResponse,
    StoreMovementCreate, StoreMovementUpdate, StoreMovementResponse, PaginatedStoreMovementResponse,
    WarehouseSummaryResponse
)

class Funcionalities:
    def __init__(self, container, database_key: str = "core_db"):
        self.container = container
        self.database_key = database_key
        
    def _get_db(self) -> Session:
        """Obtiene la sesión de base de datos desde el container"""
        return self.container.get(self.database_key, "databases")
    
    # ========== STORE CENTER METHODS ==========
    def create_store_center(self, center_data: StoreCenterCreate) -> StoreCenterResponse:
        """Crea un nuevo centro de almacenamiento"""
        db = self._get_db()
        try:
            center = StoreCenterModel(**center_data.model_dump(exclude_none=True))
            db.add(center)
            db.commit()
            db.refresh(center)
            return StoreCenterResponse.model_validate(center)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_store_centers_paginated(self, page: int = 1, per_page: int = 10, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = "") -> PaginatedStoreCenterResponse:
        """Obtiene centros de almacenamiento paginados (solo los no deshabilitados)"""
        db = self._get_db()

        lots_count_sq = (
            db.query(
                LotModel.current_store_center_id.label("gc_id"),
                func.count(LotModel.id).label("lots_count"),
            )
            .filter(LotModel.disabled_at.is_(None))
            .group_by(LotModel.current_store_center_id)
            .subquery()
        )
        
        query = (
            db.query(
                StoreCenterModel,
                func.coalesce(lots_count_sq.c.lots_count, 0).label("lots_count")
            )
            .outerjoin(
                lots_count_sq,
                lots_count_sq.c.gc_id == StoreCenterModel.id,
            )
            .filter(StoreCenterModel.disabled_at.is_(None)))
        
        # Aplicar búsqueda si se proporciona
        if search:
            query = query.filter(
                (StoreCenterModel.name.ilike(f"%{search}%")) |
                (StoreCenterModel.code.isnot(None) & StoreCenterModel.code.ilike(f"%{search}%"))
            )
        
        # Aplicar ordenamiento
      
        if sort_by == "lots_count":
            sort_column = lots_count_sq.c.lots_count
        elif sort_by:
            sort_column = getattr(StoreCenterModel, sort_by, None)
        else:
            sort_column = None

        if sort_column is not None:
            query = query.order_by(
                sort_column.desc() if order == "desc" else sort_column.asc()
            )
        total = (
            db.query(func.count(StoreCenterModel.id))
            .filter(StoreCenterModel.disabled_at.is_(None))
            .scalar()
        )
        
        offset = (page - 1) * per_page
        total_pages = (total + per_page - 1) // per_page
        
        centers = query.offset(offset).limit(per_page).all()

        items = []
        for center, lots_count in centers:
            data = StoreCenterResponse.model_validate(center)
            data.lots_count = lots_count
            items.append(data)
        
        
        return PaginatedStoreCenterResponse(
            items=items,
            total=total,
            page=page,
            page_size=per_page,
            total_pages=total_pages
        )
    
    def get_store_center_by_id(self, center_id: UUID) -> Optional[StoreCenterResponse]:
        """Obtiene un centro de almacenamiento específico por ID (solo si no está deshabilitado)"""
        db = self._get_db()
        center = db.query(StoreCenterModel).filter(
            StoreCenterModel.id == center_id,
            StoreCenterModel.disabled_at.is_(None)
        ).first()
        
        if not center:
            return None
        
        return StoreCenterResponse.model_validate(center)
    
    def update_store_center(self, center_id: UUID, center_data: StoreCenterUpdate) -> Optional[StoreCenterResponse]:
        """Actualiza un centro de almacenamiento existente"""
        db = self._get_db()
        try:
            center = db.query(StoreCenterModel).filter(
                StoreCenterModel.id == center_id,
                StoreCenterModel.disabled_at.is_(None)
            ).first()
            
            if not center:
                return None
            
            update_data = center_data.model_dump(exclude_none=True)
            for key, value in update_data.items():
                setattr(center, key, value)
            
            center.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(center)
            return StoreCenterResponse.model_validate(center)
        except Exception as e:
            db.rollback()
            raise e
    
    def disable_store_center(self, center_id: UUID) -> Optional[StoreCenterResponse]:
        """Deshabilita un centro de almacenamiento"""
        db = self._get_db()
        try:
            center = db.query(StoreCenterModel).filter(
                StoreCenterModel.id == center_id,
                StoreCenterModel.disabled_at.is_(None)
            ).first()
            
            if not center:
                return None
            
            center.disabled_at = datetime.utcnow()
            center.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(center)
            
            print(f"✓ Centro de almacenamiento {center.name} deshabilitado")
            return StoreCenterResponse.model_validate(center)
        except Exception as e:
            db.rollback()
            raise e
    
    def restore_store_center(self, center_id: UUID) -> Optional[StoreCenterResponse]:
        """Restaura un centro de almacenamiento deshabilitado"""
        db = self._get_db()
        try:
            center = db.query(StoreCenterModel).filter(
                StoreCenterModel.id == center_id,
                StoreCenterModel.disabled_at.isnot(None)
            ).first()
            
            if not center:
                return None
            
            center.disabled_at = None
            center.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(center)
            
            print(f"✓ Centro de almacenamiento {center.name} restaurado")
            return StoreCenterResponse.model_validate(center)
        except Exception as e:
            db.rollback()
            raise e
    
    def export_store_centers_to_excel(self, sort_by: Optional[str] = None, order: Optional[str] = "asc", search: str = ""):
        """Exporta todos los centros de almacenamiento a Excel con los mismos datos que la paginación"""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        db = self._get_db()
        
        lots_count_sq = (
            db.query(
                LotModel.current_store_center_id.label("gc_id"),
                func.count(LotModel.id).label("lots_count"),
            )
            .filter(LotModel.disabled_at.is_(None))
            .group_by(LotModel.current_store_center_id)
            .subquery()
        )
        
        query = (
            db.query(
                StoreCenterModel,
                func.coalesce(lots_count_sq.c.lots_count, 0).label("lots_count")
            )
            .outerjoin(
                lots_count_sq,
                lots_count_sq.c.gc_id == StoreCenterModel.id,
            )
            .filter(StoreCenterModel.disabled_at.is_(None)))
        
        # Aplicar búsqueda si existe
        if search:
            query = query.filter(
                (StoreCenterModel.name.ilike(f"%{search}%")) |
                (StoreCenterModel.code.isnot(None) & StoreCenterModel.code.ilike(f"%{search}%"))
            )
        
        # Aplicar ordenamiento
        if sort_by == "lots_count":
            sort_column = lots_count_sq.c.lots_count
        elif sort_by:
            sort_column = getattr(StoreCenterModel, sort_by, None)
        else:
            sort_column = None

        if sort_column is not None:
            query = query.order_by(
                sort_column.desc() if order == "desc" else sort_column.asc()
            )
        
        # Obtener todos los resultados (sin paginación)
        centers = query.all()
        
        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Centros de Almacenamiento"
        
        # Estilo para el encabezado
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados (SIN IDs, solo nombres)
        headers = [
            "Nombre",
            "Código",
            "Capacidad (kg)",
            "Cantidad de Lotes",
            "Fecha de Creación"
        ]
        ws.append(headers)
        
        # Aplicar estilo al encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Agregar datos (SIN IDs)
        for center, lots_count in centers:
            ws.append([
                center.name,
                center.code or "",
                float(center.capacity_kg) if center.capacity_kg else 0.0,
                lots_count,
                center.created_at.strftime("%Y-%m-%d %H:%M:%S") if center.created_at else ""
            ])
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def get_warehouse_summary(self, store_center_id: Optional[UUID] = None) -> WarehouseSummaryResponse:
        """
        Obtiene el resumen de almacén con métricas de lotes.
        Si se proporciona store_center_id, devuelve el resumen de ese almacén específico.
        Si no, devuelve el resumen de todos los almacenes.
        """
        db = self._get_db()
        
        # Query base para lotes en almacenes
        query = db.query(LotModel).filter(
            LotModel.current_store_center_id.isnot(None),
            LotModel.disabled_at.is_(None)
        )
        
        # Filtrar por almacén específico si se proporciona
        if store_center_id:
            query = query.filter(LotModel.current_store_center_id == store_center_id)
        
        # Lotes activos (current_status = "activo")
        active_lots = query.filter(LotModel.current_status == "activo").count()
        
        # Lotes en stock (current_status = "en stock")
        stock_lots = query.filter(LotModel.current_status == "en stock").count()
        
        # Total de lotes en almacenes
        total_lots = query.count()
        
        # Último lote registrado (por fecha de creación)
        last_lot_query = query.order_by(LotModel.created_at.desc()).first()
        
        # Sumatoria de kilos (net_weight)
        kg_total_result = query.with_entities(
            func.coalesce(func.sum(LotModel.net_weight), 0)
        ).scalar()
        kg_total = float(kg_total_result) if kg_total_result else 0.0
        
        # Total de costos de los lotes (sumatoria del costo de las compras)
        # Importar PurchaseModel para calcular el costo total
        from modules.gathering.src.models.purchases import PurchaseModel
        
        # Obtener los IDs de lotes que están en almacenes
        lot_ids_query = query.with_entities(LotModel.id).all()
        lot_ids = [lot_id[0] for lot_id in lot_ids_query]
        
        # Calcular el costo total de las compras de esos lotes
        total_cost = 0.0
        if lot_ids:
            cost_result = db.query(
                func.coalesce(func.sum(PurchaseModel.quantity * PurchaseModel.price), 0)
            ).filter(
                PurchaseModel.lot_id.in_(lot_ids),
                PurchaseModel.disabled_at.is_(None)
            ).scalar()
            total_cost = float(cost_result) if cost_result else 0.0
        
        return WarehouseSummaryResponse(
            active_lots=active_lots,
            last_lot=last_lot_query,
            stock_lots=stock_lots,
            total_lots=total_lots,
            kg_total=kg_total,
            total=total_cost
        )
    
    # ========== STORE MOVEMENT METHODS ==========
    def create_store_movement(self, movement_data: StoreMovementCreate) -> StoreMovementResponse:
        """Crea un nuevo movimiento de almacén"""
        db = self._get_db()
        try:
            movement = StoreMovementModel(**movement_data.model_dump(exclude_none=True))
            db.add(movement)
            db.commit()
            db.refresh(movement)
            return StoreMovementResponse.model_validate(movement)
        except Exception as e:
            db.rollback()
            raise e
    
    def get_store_movements_paginated(
        self, 
        page: int = 1, 
        per_page: int = 10,
        sort_by: Optional[str] = None,
        order: Optional[str] = "desc",
        search: str = ""
    ) -> PaginatedStoreMovementResponse:
        """Obtiene movimientos de almacén paginados"""
        db = self._get_db()
        
        query = db.query(StoreMovementModel)
        
        # Aplicar búsqueda si se proporciona (buscar por UUIDs de lot_id, store_center_id, identity_id)
        if search:
            try:
                search_uuid = UUID(search)
                query = query.filter(
                    (StoreMovementModel.lot_id == search_uuid) |
                    (StoreMovementModel.store_center_id == search_uuid) |
                    (StoreMovementModel.identity_id == search_uuid)
                )
            except ValueError:
                # Si no es un UUID válido, no aplicar filtro de búsqueda
                pass
        
        # Aplicar ordenamiento
        if sort_by and isinstance(sort_by, str):
            sort_column = getattr(StoreMovementModel, sort_by, None)
            if sort_column is not None:
                if order and order.lower() == "desc":
                    query = query.order_by(sort_column.desc())
                else:
                    query = query.order_by(sort_column.asc())
        else:
            # Ordenamiento por defecto
            query = query.order_by(StoreMovementModel.created_at.desc())
        
        total = query.count()
        offset = (page - 1) * per_page
        total_pages = (total + per_page - 1) // per_page
        
        movements = query.offset(offset).limit(per_page).all()
        
        return PaginatedStoreMovementResponse(
            items=[StoreMovementResponse.model_validate(movement) for movement in movements],
            total=total,
            page=page,
            page_size=per_page,
            total_pages=total_pages
        )
    
    def get_store_movement_by_id(self, movement_id: UUID) -> Optional[StoreMovementResponse]:
        """Obtiene un movimiento de almacén específico por ID"""
        db = self._get_db()
        movement = db.query(StoreMovementModel).filter(
            StoreMovementModel.id == movement_id
        ).first()
        
        if not movement:
            return None
        
        return StoreMovementResponse.model_validate(movement)
    
    def update_store_movement(self, movement_id: UUID, movement_data: StoreMovementUpdate) -> Optional[StoreMovementResponse]:
        """Actualiza un movimiento de almacén existente"""
        db = self._get_db()
        try:
            movement = db.query(StoreMovementModel).filter(
                StoreMovementModel.id == movement_id
            ).first()
            
            if not movement:
                return None
            
            update_data = movement_data.model_dump(exclude_none=True)
            for key, value in update_data.items():
                setattr(movement, key, value)
            
            db.commit()
            db.refresh(movement)
            return StoreMovementResponse.model_validate(movement)
        except Exception as e:
            db.rollback()
            raise e
